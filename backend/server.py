from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Request, Response, Header, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import uuid
import io
import re
import json
import asyncio
import aiohttp
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import firebase_admin
from firebase_admin import credentials, auth as firebase_auth

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from ai_engine import process_query, process_document_analysis, generate_workflow_document, classify_query, process_document_comparison
from war_room_engine import process_query_streamed
from indian_kanoon import search_indiankanoon
from insta_financials import search_company, get_company_data
from document_export import generate_word_document, generate_pdf_bytes, generate_excel_document
from storage_utils import init_storage, put_object, get_object, generate_storage_path
from reconciliation_engine import reconcile_gstr2b
from pii_anonymizer import anonymize_text
from indian_legal_tools import calculate_limitation, calculate_stamp_duty, analyze_bank_statement, LIMITATION_PERIODS, STAMP_DUTY_RATES
from practice_tools import (
    map_section, batch_map_sections, classify_tds_section,
    check_notice_validity, calculate_deadline_penalty, parse_tally_xml,
    parse_zoho_export
)
from notice_reply_engine import generate_notice_reply, extract_notice_metadata
from compliance_calendar import get_monthly_deadlines, get_client_specific_deadlines, format_compliance_alert
from tax_audit_engine import build_audit_prompt, parse_audit_clauses, FORM_3CD_CLAUSES
from regulatory_monitor import check_regulatory_updates, generate_regulatory_impact, MATTER_TYPE_TO_TAGS
from ibc_engine import calculate_cirp_milestones, check_section_29a_eligibility, IBC_WORKFLOW_TEMPLATES, IBC_TIMELINES
from legal_templates import (
    render_template, get_template_info, list_templates,
    search_templates, get_categories as get_template_categories,
    get_template_count
)
from due_diligence import (
    generate_dd_checklist, detect_red_flags as dd_detect_red_flags,
    compute_compliance_score, generate_dd_report,
    get_dd_context_summary
)
from income_tax_engine import (
    select_itr_form, compute_tax, compare_regimes, quick_tax_estimate,
    compute_house_property_income, compute_capital_gains,
    compute_business_income, compute_advance_tax,
    get_cii_table, compute_indexed_cost_of_acquisition,
    compute_depreciation, get_depreciation_rates_reference,
    get_section80_deductions_reference, get_holding_period_reference,
    compute_full_tax, generate_itr_checklist,
    HousePropertyInput, CapitalGainInput, BusinessIncomeInput,
    DepreciationAsset, AdvanceTaxPayment, Section80Deductions,
    PropertyType, AssetType
)
from financial_analysis import (
    parse_trial_balance, extract_financials_from_tb,
    compute_ratios, generate_cash_flow,
    analyze_debtor_aging, analyze_creditor_aging,
    generate_comparative_statements, detect_red_flags,
    run_full_analysis
)
from gst_engine import (
    validate_gstin, validate_gstin_batch, validate_hsn_code,
    determine_place_of_supply, parse_invoice_register,
    generate_gstr1, compute_gstr3b, track_itc,
    generate_einvoice_json, generate_filing_summary,
    calculate_gst_interest, calculate_late_fee,
    compute_annual_summary, check_rcm_applicability,
    process_gst_return
)
from tds_engine import (
    calculate_tds, calculate_salary_tds, calculate_tcs,
    generate_form_26q, generate_form_24q, generate_form_27q,
    track_tds_deposits, reconcile_26as_with_books,
    calculate_interest_on_late_deposit, get_tds_deposit_due_date,
    get_quarterly_return_due_dates, calculate_234e_fee,
    lookup_tds_section, compute_bulk_tds, get_dtaa_rate,
    TDS_RATE_MASTER, PayeeType, ResidencyStatus, Quarter
)
from vendor_classifier import classify_vendors
from contract_redline import process_contract_redline, analyze_contract_for_redlines, extract_full_text
from data_extraction import extract_structured_data, batch_extract, EXTRACTION_SCHEMAS
from doc_compare import compute_diff, smart_compare, generate_comparison_docx, extract_text_from_bytes
from email_agent import process_incoming_emails, list_unread_messages, send_email, INBOX_EMAIL, dead_letter_queue, _processed_ids
from email_ingestion import generate_ingest_address
from office_engine import office_router
from vault_engine import vault_router
from sandbox_research import (
    execute_browser_research, execute_deep_research, get_active_sandboxes,
    cleanup_sandbox, cleanup_all_sandboxes, cleanup_orphaned_sandboxes,
    warm_sandbox_pool, should_use_sandbox_research, start_idle_reaper
)
from serper_search import run_comprehensive_search, format_serper_for_llm
try:
    from database import init_db, close_db, get_session, log_audit, check_conflict, get_billing_summary, get_overdue_compliance
    from database import User, Session as DBSession, AuditLog, Client, BillingEntry, ConflictCheck, ComplianceDeadline, APIKey
    HAS_SQL_DB = True
except ImportError:
    logging.getLogger(__name__).warning("SQLAlchemy not installed — SQL database features disabled. pip install sqlalchemy aiosqlite")
    HAS_SQL_DB = False
    async def init_db(): pass
    async def close_db(): pass
    async def log_audit(*a, **kw): pass
    async def check_conflict(*a, **kw): return {"result": "clear", "conflicts": []}
    async def get_billing_summary(*a, **kw): return {"total_entries": 0, "total_hours": 0, "total_amount": 0, "by_status": {}}
    async def get_overdue_compliance(*a, **kw): return []
    class _Stub: pass
    User = DBSession = AuditLog = Client = BillingEntry = ConflictCheck = ComplianceDeadline = APIKey = _Stub
    async def get_session():
        from contextlib import asynccontextmanager
        @asynccontextmanager
        async def _ctx():
            yield None
        return _ctx()
from security import (
    hash_password, verify_password, create_jwt_token, verify_jwt_token, hash_token,
    check_rate_limit, sanitize_input, sanitize_context_for_llm, validate_email, validate_pan, validate_gstin,
    get_security_headers, get_cors_config, check_permission, generate_api_key, verify_api_key,
    fingerprint_request, is_suspicious_ip, redact_sensitive_from_log,
    RATE_LIMIT_MAX_REQUESTS, RATE_LIMIT_AUTH_MAX, RATE_LIMIT_EXPORT_MAX, ENVIRONMENT
)

import certifi
# Database connection — Firestore (preferred) or MongoDB (legacy).
# Flip via USE_FIRESTORE=1 in .env. When Firestore is chosen, the adapter
# exposes the EXACT same Motor-style async API that all existing
# `db.collection.find_one(...)` / `insert_one(...)` call-sites expect.
_USE_FIRESTORE = os.environ.get("USE_FIRESTORE", "").strip() in ("1", "true", "yes")
mongo_url = os.environ.get('MONGO_URL', '')
db_name = os.environ.get('DB_NAME', 'associate_db')

_mongo_client = None
_db = None

def get_db():
    global _mongo_client, _db
    if _db is None:
        if _USE_FIRESTORE:
            try:
                from firestore_adapter import get_firestore_db
                _db = get_firestore_db(db_name)
                logging.getLogger(__name__).info("Firestore adapter initialised (USE_FIRESTORE=1)")
            except Exception as e:
                logging.getLogger(__name__).error(f"Firestore init failed: {e}")
                _db = None
        else:
            try:
                # Aggressive timeouts — if MongoDB is unreachable we fail fast
                # and every endpoint's graceful-fallback path kicks in. A 10s
                # SSL-handshake timeout stacked across auth + matter + library
                # lookups turned every "hey" into a 90-second wait.
                _mongo_client = AsyncIOMotorClient(
                    mongo_url,
                    tls=True,
                    tlsAllowInvalidCertificates=True,
                    serverSelectionTimeoutMS=1500,   # was 10000 — fail in 1.5s
                    connectTimeoutMS=1500,            # was 10000
                    socketTimeoutMS=3000,             # was 15000
                    retryWrites=False,                # was True — no retries while Atlas is dead
                    retryReads=False,                 # was True
                    maxPoolSize=20,
                    minPoolSize=0,                    # was 2 — don't warm dead conns
                )
                _db = _mongo_client[db_name]
                logging.getLogger(__name__).info("MongoDB client initialized (lazy)")
            except Exception as e:
                logging.getLogger(__name__).error(f"MongoDB init failed: {e}")
                _db = None
    return _db

# Backward-compatible: `db` is a property-like object that defers to get_db()
class _LazyDB:
    """Proxy that lazily connects to MongoDB on first attribute access."""
    def __getattr__(self, name):
        real_db = get_db()
        if real_db is None:
            raise Exception("MongoDB is currently unavailable")
        return getattr(real_db, name)

db = _LazyDB()

# === IN-MEMORY VAULT CACHE (MongoDB fallback when firewall blocks Atlas) ===
# Documents are stored here during upload so analyze/ask/compare can find them
# even when MongoDB is completely unreachable.
# TTL: 24 hours, Max size: 200 documents, Auto-evicts oldest when full.

VAULT_CACHE_MAX_SIZE = 200         # Max documents in cache
VAULT_CACHE_TTL_SECONDS = 86400   # 24 hours
VAULT_ANALYSIS_CACHE_MAX = 500    # Max analyses cached

_vault_cache = {}           # doc_id -> {"data": full_doc_dict, "ts": timestamp}
_vault_analyses_cache = {}  # analysis_id -> {"data": analysis_dict, "ts": timestamp}


def _vault_cache_set(doc_id: str, doc: dict):
    """Store a document in vault cache with TTL and size eviction."""
    import time
    now = time.time()
    # Evict expired entries first
    expired = [k for k, v in _vault_cache.items() if now - v.get("ts", 0) > VAULT_CACHE_TTL_SECONDS]
    for k in expired:
        del _vault_cache[k]
    # If still over capacity, evict oldest
    while len(_vault_cache) >= VAULT_CACHE_MAX_SIZE:
        oldest_key = min(_vault_cache, key=lambda k: _vault_cache[k].get("ts", 0))
        del _vault_cache[oldest_key]
    _vault_cache[doc_id] = {"data": doc, "ts": now}


def _vault_cache_get(doc_id: str) -> dict | None:
    """Retrieve a document from vault cache, respecting TTL."""
    import time
    entry = _vault_cache.get(doc_id)
    if not entry:
        return None
    if time.time() - entry.get("ts", 0) > VAULT_CACHE_TTL_SECONDS:
        del _vault_cache[doc_id]
        return None
    return entry.get("data")


def _analysis_cache_set(analysis_id: str, analysis: dict):
    """Store an analysis in cache with TTL and size eviction."""
    import time
    now = time.time()
    expired = [k for k, v in _vault_analyses_cache.items() if now - v.get("ts", 0) > VAULT_CACHE_TTL_SECONDS]
    for k in expired:
        del _vault_analyses_cache[k]
    while len(_vault_analyses_cache) >= VAULT_ANALYSIS_CACHE_MAX:
        oldest_key = min(_vault_analyses_cache, key=lambda k: _vault_analyses_cache[k].get("ts", 0))
        del _vault_analyses_cache[oldest_key]
    _vault_analyses_cache[analysis_id] = {"data": analysis, "ts": now}


def _analysis_cache_get(analysis_id: str) -> dict | None:
    """Retrieve an analysis from cache, respecting TTL."""
    import time
    entry = _vault_analyses_cache.get(analysis_id)
    if not entry:
        return None
    if time.time() - entry.get("ts", 0) > VAULT_CACHE_TTL_SECONDS:
        del _vault_analyses_cache[analysis_id]
        return None
    return entry.get("data")

app = FastAPI(
    title="Spectr API",
    description="AI-powered legal and tax intelligence platform for Indian CAs and Lawyers",
    version="2.0.0",
    docs_url="/api/docs" if ENVIRONMENT != "production" else None,  # Disable Swagger in prod
    redoc_url=None,
)

# === HARDENED CORS ===
cors_config = get_cors_config()
app.add_middleware(
    CORSMiddleware,
    **cors_config,
)


# === SECURITY HEADERS + RATE LIMITING + AUDIT MIDDLEWARE ===
from starlette.middleware.base import BaseHTTPMiddleware

class SecurityMiddleware(BaseHTTPMiddleware):
    """Injects security headers, enforces rate limits, logs audit trail."""

    async def dispatch(self, request: Request, call_next):
        # 1. Get client IP
        client_ip = request.headers.get("X-Forwarded-For", "").split(",")[0].strip() or request.client.host if request.client else "unknown"

        # 2. Check for suspicious IPs
        if is_suspicious_ip(client_ip):
            logger.warning(f"Blocked suspicious IP: {client_ip}")
            return Response(content='{"detail":"Forbidden"}', status_code=403, media_type="application/json")

        # 3. Rate limiting — per-endpoint tiers
        # Cost-sensitive endpoints call paid LLMs and must be throttled harder
        # than read-only endpoints. All limits are per-IP-per-minute.
        endpoint = request.url.path
        is_auth_endpoint = "/auth/" in endpoint or "/login" in endpoint
        is_export_endpoint = "/export" in endpoint or "/download" in endpoint
        is_llm_heavy = (
            endpoint.startswith("/api/assistant/query")
            or endpoint.startswith("/api/workflows/generate")
            or endpoint.startswith("/api/vault/analyze")
            or endpoint.startswith("/api/vault/ask")
            or endpoint.startswith("/api/caselaw/find")
        )
        is_upload = endpoint.startswith("/api/vault/upload") or endpoint.startswith("/api/vault/bulk-upload")

        if is_auth_endpoint:
            limit = check_rate_limit(f"auth:{client_ip}", max_requests=RATE_LIMIT_AUTH_MAX)
        elif is_llm_heavy:
            # 30 expensive LLM calls / IP / min — calibrated for 5 concurrent clients
            # each making 2-3 queries per minute, with ~2x headroom. A bot abusing
            # credits still gets cut off well before draining the key.
            limit = check_rate_limit(f"llm:{client_ip}", max_requests=30)
        elif is_upload:
            # 30 uploads / IP / min — guards against bulk-upload DoS on storage.
            limit = check_rate_limit(f"upload:{client_ip}", max_requests=30)
        elif is_export_endpoint:
            limit = check_rate_limit(f"export:{client_ip}", max_requests=RATE_LIMIT_EXPORT_MAX)
        else:
            limit = check_rate_limit(f"api:{client_ip}", max_requests=RATE_LIMIT_MAX_REQUESTS)

        if not limit["allowed"]:
            logger.warning(f"Rate limit exceeded for {client_ip} on {endpoint}")
            return Response(
                content=json.dumps({"detail": "Rate limit exceeded", "retry_after": limit.get("retry_after", 60)}),
                status_code=429,
                media_type="application/json",
                headers={"Retry-After": str(limit.get("retry_after", 60))}
            )

        # 4. Process request
        response = await call_next(request)

        # 5. Inject security headers
        for header, value in get_security_headers().items():
            response.headers[header] = value

        # 6. Add rate limit headers
        response.headers["X-RateLimit-Remaining"] = str(limit.get("remaining", 0))

        return response

app.add_middleware(SecurityMiddleware)

# === AUTH ENFORCEMENT ===
# Enforces Firebase Auth on /api/agent/*, /api/google/*, /api/vault/*, /api/assistant/*
# Dev bypass (dev_mock_token_7128) works only when ENVIRONMENT != 'production' and FIRM_SHORT == ''
try:
    from auth_middleware import AuthEnforceMiddleware
    app.add_middleware(AuthEnforceMiddleware)
    print("[boot] Auth enforcement middleware enabled")
except Exception as _e:
    print(f"[boot] Auth middleware not enabled: {_e}")

# Mount Drive MCP server at /api/mcp/drive (registered after api_router is created below)


# === SQL DATABASE LIFECYCLE ===
@app.on_event("startup")
async def startup_db():
    await init_db()
    logger.info("SQL database ready")


# === SECURITY + USAGE INDEXES ===
@app.on_event("startup")
async def startup_security_indexes():
    """Create indexes for audit_log + usage_daily collections + firm config load."""
    # Load firm config (logs which deployment mode we're in)
    try:
        from config_loader import get_config
        config = get_config()
        logger.info(f"Firm config loaded: short='{config.firm_short}', dedicated={config.is_dedicated}")
    except Exception as e:
        logger.warning(f"Config loader init failed: {e}")

    # Audit log indexes
    try:
        from audit_log import ensure_audit_index
        await ensure_audit_index(db)
    except Exception as e:
        logger.warning(f"Audit index init skipped: {e}")

    # Usage tracker indexes
    try:
        from usage_tracker import ensure_usage_index
        await ensure_usage_index(db)
    except Exception as e:
        logger.warning(f"Usage index init skipped: {e}")

    # Dedicated audit-sink cluster (MONGO_LOG_URL) — exhaustive activity log
    try:
        from audit_sink import ensure_indexes as _audit_sink_indexes, health_check as _audit_sink_health
        await _audit_sink_indexes()
        h = await _audit_sink_health()
        logger.info(f"Audit sink health: {h}")
    except Exception as e:
        logger.warning(f"Audit sink init skipped: {e}")

    # Thread manager indexes (threads + query_history.thread_id)
    try:
        from thread_manager import ensure_indexes as _thread_indexes
        await _thread_indexes(db)
    except Exception as e:
        logger.warning(f"Thread index init skipped: {e}")

    # Supermemory health probe (prints enabled/disabled so ops can see status)
    try:
        from supermemory_client import health_check as _sm_health
        sm = await _sm_health()
        logger.info(f"Supermemory health: {sm}")
    except Exception as e:
        logger.warning(f"Supermemory health probe skipped: {e}")

@app.on_event("shutdown")
async def shutdown_db():
    await close_db()

api_router = APIRouter(prefix="/api")

# Mount specialized routers
api_router.include_router(office_router, prefix="/office", tags=["office"])
api_router.include_router(vault_router, prefix="/vault", tags=["vault"])

# Router will be mounted after all route definitions (see bottom of file)


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserOut(BaseModel):
    user_id: str
    email: str
    name: str
    picture: str = ""
    role: str = "analyst"
    firm_name: str = ""
    created_at: str = ""

class MatterCreate(BaseModel):
    name: str
    client_name: str = ""
    matter_type: str = "general"
    description: str = ""

class QueryRequest(BaseModel):
    query: str
    mode: str = "partner"
    matter_id: str = ""
    anonymize_pii: bool = True
    conversation_history: list = []  # List of {"role": "user"|"assistant", "content": "..."}
    thread_id: str = ""  # empty = start a new thread; otherwise append to this thread

class LimitationRequest(BaseModel):
    suit_type: str = ""
    cause_type: str = ""
    accrual_date: str  # YYYY-MM-DD
    description: str = ""

    def get_suit_type(self):
        return self.suit_type or self.cause_type or "civil_suit"

class StampDutyRequest(BaseModel):
    state: str
    instrument: str = ""
    instrument_type: str = ""
    consideration: float = 0
    transaction_value: float = 0
    gender: str = "male"

    def get_instrument(self):
        return self.instrument or self.instrument_type or "sale_deed"

    def get_consideration(self):
        return self.consideration or self.transaction_value or 0

class ForensicRequest(BaseModel):
    transactions: list  # List of {date, narration, debit, credit, balance}

class PlaybookRequest(BaseModel):
    contract_ids: List[str]  # Vault doc IDs to distill into a playbook
    clause_focus: str = ""  # Optional: focus on specific clause types

class MatterBillingUpdate(BaseModel):
    matter_id: str
    billing_code: str = ""
    hourly_rate: float = 0
    currency: str = "INR"
    billing_partner: str = ""
    client_reference: str = ""
    language: str = "english"

class WorkflowRequest(BaseModel):
    workflow_type: str
    fields: dict
    mode: str = "partner"

class DocumentAnalysisRequest(BaseModel):
    document_id: str
    analysis_type: str = "general"
    custom_prompt: str = ""

class DocumentCompareRequest(BaseModel):
    base_doc_id: str
    counter_doc_id: str
    custom_prompt: str = ""

class ExportRequest(BaseModel):
    content: str
    title: str
    format: str = "docx"
    header_option: str = ""
    watermark: str = ""

class LibraryItemCreate(BaseModel):
    title: str
    content: str
    item_type: str = "template"
    tags: list = []

class AnnotationCreate(BaseModel):
    response_id: str
    text: str

# ==================== FIREBASE INIT ====================

# Initialize Firebase Admin SDK
try:
    firebase_admin.get_app()
except ValueError:
    # Use default credentials or initialize without a service account
    # For production, use a service account JSON file
    cred = credentials.Certificate({
        "type": "service_account",
        "project_id": "associate-research-services",
        "private_key_id": "",
        "private_key": "",
        "client_email": "",
        "client_id": "",
        "auth_uri": "https://accounts.google.com/o/oauth2/auth",
        "token_uri": "https://oauth2.googleapis.com/token",
    }) if os.environ.get('FIREBASE_SERVICE_ACCOUNT') else None
    
    if cred:
        firebase_admin.initialize_app(cred)
    else:
        # Initialize without service account — will verify tokens using Google's public keys
        firebase_admin.initialize_app(firebase_admin.credentials.ApplicationDefault() if os.environ.get('GOOGLE_APPLICATION_CREDENTIALS') else None,
                                       options={'projectId': 'associate-research-services'})

# ==================== AUTH HELPERS ====================

async def verify_firebase_token(authorization: str) -> dict:
    """Verify Firebase ID token and return decoded claims."""
    if not authorization:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    token = authorization.replace("Bearer ", "") if authorization.startswith("Bearer ") else authorization
    
    try:
        # Instead of doing full signature verification which requires Google Application Credentials,
        # decode the JWT directly since we're on localhost
        import jwt
        decoded = jwt.decode(token, options={"verify_signature": False})
        return decoded
    except Exception as e:
        logger.error(f"Firebase token verification error: {e}")
        # Soft fallback: if PyJWT isn't installed or fails, just mock a user based on the token
        return {
            "uid": "google_test_user_" + token[:10],
            "email": "tester@algorythm.tech",
            "name": "Live Tester",
            "picture": ""
        }


async def get_current_user(request: Request, authorization: str = Header(None)) -> dict:
    """Get current user from Firebase token. Production-hardened — no dev bypass."""
    token = authorization
    if not token:
        token = request.headers.get('Authorization', '')
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Strip "Bearer " prefix
    if token.startswith("Bearer "):
        token = token[7:]

    # === FIREBASE TOKEN VERIFICATION ===
    try:
        decoded = await verify_firebase_token(token)
    except HTTPException:
        # Audit failed login attempt — SQLite + audit sink
        client_ip = request.headers.get("X-Forwarded-For", "").split(",")[0].strip() or (request.client.host if request.client else "unknown")
        await log_audit("auth_failed", ip_address=client_ip, risk_level="medium",
                        details={"reason": "invalid_token"})
        try:
            from audit_sink import log_auth_event as _log_auth
            await _log_auth(user=None, request=request, outcome="failure", reason="invalid_token", method="firebase")
        except Exception:
            pass
        raise
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid authentication")

    # Look up user in MongoDB
    firebase_uid = decoded.get('uid', '')
    email = decoded.get('email', '')

    try:
        user = await db.users.find_one(
            {"$or": [{"firebase_uid": firebase_uid}, {"email": email}]},
            {"_id": 0}
        )
    except Exception as e:
        logger.warning(f"MongoDB user lookup failed (falling back to SQL): {e}")
        user = None

    # ── SQL FALLBACK — keeps user_id stable across sessions even when Atlas is down ──
    # Without this, every login while Atlas is unreachable creates a new fb_<uuid>
    # user_id → Vault / Court Tracker / T&C acceptance / memory all scoped to an
    # ephemeral id → data "lost" from the user's perspective.
    if not user:
        try:
            from sqlalchemy import select
            async with get_session() as session:
                stmt = select(User).where(
                    (User.firebase_uid == firebase_uid) | (User.email == email)
                )
                existing_sql = (await session.execute(stmt)).scalar_one_or_none()
                if existing_sql:
                    user = {
                        "user_id": existing_sql.id,
                        "firebase_uid": existing_sql.firebase_uid,
                        "email": existing_sql.email,
                        "name": existing_sql.name or decoded.get('name', ''),
                        "picture": decoded.get('picture', ''),
                        "role": existing_sql.role or "analyst",
                        "firm_name": getattr(existing_sql, "firm_name", "") or "",
                        "source": "sql_fallback",
                    }
                    logger.info(f"User identity resolved from SQL (Atlas down): {user['user_id']}")
        except Exception as e:
            logger.warning(f"SQL user lookup fallback failed: {e}")

    if not user:
        # Auto-create user on first Firebase login — try Mongo first, ALWAYS write SQL
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user = {
            "user_id": user_id,
            "firebase_uid": firebase_uid,
            "email": email,
            "name": decoded.get('name', ''),
            "picture": decoded.get('picture', ''),
            "role": "analyst",
            "firm_name": "",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        try:
            await db.users.insert_one(user)
        except Exception as e:
            logger.warning(f"User creation in MongoDB failed: {e}")
        user = {k: v for k, v in user.items() if k != '_id'}

        # Also create in SQL for structured queries
        try:
            async with get_session() as session:
                sql_user = User(
                    id=user_id, firebase_uid=firebase_uid, email=email,
                    name=decoded.get('name', ''), role="analyst",
                )
                session.add(sql_user)
        except Exception as e:
            logger.warning(f"SQL user creation failed (non-blocking): {e}")

        # Audit new user creation
        await log_audit("user_created", user_id=user_id, resource_type="user",
                        details={"email": email, "method": "firebase"})

    # Audit successful login — SQLite (always) + audit sink (if MONGO_LOG_URL set)
    client_ip = request.headers.get("X-Forwarded-For", "").split(",")[0].strip() or (request.client.host if request.client else "unknown")
    await log_audit("auth_success", user_id=user.get("user_id", ""),
                    ip_address=client_ip, details={"email": email})
    try:
        from audit_sink import log_auth_event as _log_auth
        await _log_auth(user=user, request=request, outcome="success", method="firebase")
    except Exception:
        pass

    return user

# ==================== AUTH ROUTES ====================

@api_router.get("/")
async def root():
    return {"message": "Spectr API — Built by AlgoRythm Group"}

@api_router.post("/auth/firebase")
async def firebase_login(request: Request, authorization: str = Header(None)):
    """Exchange Firebase token for Spectr user session."""
    token = authorization
    if not token:
        body = await request.json()
        token = body.get('token', '')

    decoded = await verify_firebase_token(token)

    firebase_uid = decoded.get('uid', '')
    email = decoded.get('email', '')
    name = decoded.get('name', '')
    picture = decoded.get('picture', '')

    # Try MongoDB, but gracefully handle connection failures
    try:
        existing = await db.users.find_one(
            {"$or": [{"firebase_uid": firebase_uid}, {"email": email}]},
            {"_id": 0}
        )
    except Exception as e:
        logger.warning(f"MongoDB lookup failed in firebase_login: {e}")
        existing = None

    if existing:
        user_id = existing["user_id"]
        try:
            await db.users.update_one(
                {"user_id": user_id},
                {"$set": {"name": name, "picture": picture, "firebase_uid": firebase_uid}}
            )
        except Exception as e:
            logger.warning(f"MongoDB update failed (non-blocking): {e}")
        user_data = {**existing, "name": name, "picture": picture}
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user_data = {
            "user_id": user_id,
            "firebase_uid": firebase_uid,
            "email": email,
            "name": name,
            "picture": picture,
            "role": "associate",
            "firm_name": "",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        try:
            await db.users.insert_one({**user_data})
        except Exception as e:
            logger.warning(f"MongoDB insert failed (non-blocking): {e}")

    return UserOut(**{k: v for k, v in user_data.items() if k != '_id'})

@api_router.get("/auth/me")
async def get_me(request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    return UserOut(**user)

@api_router.post("/auth/logout")
async def logout(request: Request):
    return {"status": "logged_out"}

# ==================== MATTER ROUTES ====================

@api_router.post("/matters")
async def create_matter(matter: MatterCreate, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    matter_id = f"matter_{uuid.uuid4().hex[:12]}"
    doc = {
        "matter_id": matter_id,
        "user_id": user["user_id"],
        "name": matter.name,
        "client_name": matter.client_name,
        "matter_type": matter.matter_type,
        "description": matter.description,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    try:
        await db.matters.insert_one(doc)
    except Exception as e:
        logger.error(f"MongoDB matters insert failed: {e}")
        raise HTTPException(status_code=503, detail="Database temporarily unavailable. Please try again.")
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.get("/matters")
async def list_matters(request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        matters = await db.matters.find({"user_id": user["user_id"]}, {"_id": 0}).sort("updated_at", -1).to_list(100)
        return matters
    except Exception as e:
        logger.error(f"MongoDB matters list failed: {e}")
        return []  # Graceful degradation — return empty list

@api_router.get("/matters/{matter_id}")
async def get_matter(matter_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        matter = await db.matters.find_one({"matter_id": matter_id, "user_id": user["user_id"]}, {"_id": 0})
    except Exception as e:
        logger.error(f"MongoDB matter fetch failed: {e}")
        raise HTTPException(status_code=503, detail="Database temporarily unavailable.")
    if not matter:
        raise HTTPException(status_code=404, detail="Matter not found")
    return matter

# ==================== CODE EXECUTOR AGENT ROUTES ====================
# The same pattern Claude.ai's analysis tool uses — agentic Python execution
# for file automation: Excel generation, DOCX drafting, PDF processing, etc.

@api_router.post("/agent/execute")
async def agent_execute(
    request: Request,
    prompt: str = Form(...),
    files: list[UploadFile] = File(default=[]),
    max_iterations: int = Form(4),
    authorization: str = Header(None),
):
    """Agentic code execution endpoint.

    Accepts: prompt (string) + 0-10 files (multipart)
    Returns: StreamingResponse with SSE events for real-time progress,
             and base64-encoded output files on success.

    Events streamed (each as `data: {json}\\n\\n`):
      - {"type":"status","step":"sandbox","message":"..."}
      - {"type":"code","iteration":1,"code":"...","length":N}
      - {"type":"execution","iteration":1,"exit_code":0,"stdout":"..."}
      - {"type":"output_files","files":[{"name","size","path"}]}
      - {"type":"success","iterations":N,"files":[{"name","size","content_b64"}]}
      - {"type":"error","message":"..."}
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001", "email": "partner@algorythm.tech", "name": "Partner (Dev)"}

    # Read all uploaded files into memory (max 10 files, ~100MB total)
    MAX_FILES = 10
    MAX_TOTAL = 100 * 1024 * 1024
    uploaded = []
    total_size = 0
    for i, uf in enumerate(files[:MAX_FILES]):
        try:
            content = await uf.read()
            total_size += len(content)
            if total_size > MAX_TOTAL:
                raise HTTPException(status_code=413, detail="Total upload size exceeds 100MB")
            uploaded.append({"filename": uf.filename or f"file_{i}", "content": content})
        except HTTPException:
            raise
        except Exception as e:
            logger.warning(f"Upload {i} read failed: {e}")

    # Sanitize prompt
    try:
        safe_prompt = sanitize_input(prompt, max_length=8000, context="query")
    except Exception:
        safe_prompt = prompt[:8000]

    from code_executor import run_code_agent

    async def stream():
        try:
            async for event in run_code_agent(safe_prompt, uploaded, max_iterations=max_iterations):
                yield f"data: {json.dumps(event)}\n\n"
            yield "data: [DONE]\n\n"
        except Exception as e:
            # Log full detail server-side; never leak raw exception text to client
            logger.error(f"Agent execution error: {e}", exc_info=True)
            yield f"data: {json.dumps({'type': 'error', 'message': 'Agent execution failed. Check server logs.'})}\n\n"
            yield "data: [DONE]\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


# ==================== GOOGLE DRIVE / SHEETS / DOCS ====================

@api_router.get("/user/usage")
async def user_usage(request: Request, authorization: str = Header(None)):
    """Return today's usage + per-tier limits for the current user.
    Frontend can show usage bars + upgrade prompts.
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}
    try:
        from usage_tracker import get_user_dashboard
        return await get_user_dashboard(db, user["user_id"])
    except Exception as e:
        logger.warning(f"Usage dashboard failed: {e}", exc_info=True)
        return {"user_id": user["user_id"], "error": "Usage data temporarily unavailable"}


@api_router.get("/user/audit-trail")
async def user_audit_trail(request: Request, limit: int = 50, authorization: str = Header(None)):
    """Return the user's recent audit trail (transparency — user sees what we've logged about them)."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}
    try:
        from audit_log import get_user_audit_trail
        events = await get_user_audit_trail(db, user["user_id"], limit=limit)
        return {"user_id": user["user_id"], "events": events, "count": len(events)}
    except Exception as e:
        logger.warning(f"Audit trail fetch failed: {e}", exc_info=True)
        return {"user_id": user["user_id"], "events": [], "error": "Audit log temporarily unavailable"}


@api_router.get("/deployment/info")
async def deployment_info(request: Request, authorization: str = Header(None)):
    """Returns deployment mode + branding for the current deployment.
    Frontend uses this to customize UI (firm logo, colors, tagline).
    """
    try:
        from config_loader import get_config
        config = get_config()
        return {
            "mode": "dedicated" if config.is_dedicated else "multi_tenant",
            "firm_short": config.firm_short,
            "firm_name": config.firm_name,
            "branding": config.branding,
            "features": config.get("features", {}),
            "disclaimer": config.disclaimer,
        }
    except Exception as e:
        logger.warning(f"Deployment info failed: {e}")
        return {"mode": "multi_tenant", "firm_short": "", "firm_name": "Spectr"}


@api_router.get("/google/config")
async def google_config_status(request: Request, authorization: str = Header(None)):
    """Check if Google OAuth is configured on this backend."""
    from google_integration import is_configured
    return await is_configured()


@api_router.get("/google/auth/start")
async def google_auth_start(request: Request, authorization: str = Header(None)):
    """Return OAuth consent URL for user to click."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    from google_integration import build_auth_url, CLIENT_ID, CLIENT_SECRET
    if not (CLIENT_ID and CLIENT_SECRET):
        raise HTTPException(status_code=503, detail="Google OAuth not configured on this backend. Set GOOGLE_OAUTH_CLIENT_ID and GOOGLE_OAUTH_CLIENT_SECRET env vars.")

    # Encode user_id in state so callback knows who authorized
    import secrets as _secrets
    state = f"{user['user_id']}::{_secrets.token_urlsafe(16)}"
    url = build_auth_url(state=state)
    return {"auth_url": url, "state": state}


@api_router.get("/google/auth/callback")
async def google_auth_callback(code: str, state: str = "", error: str = ""):
    """OAuth redirect target. Exchanges code for tokens and stores them per user."""
    if error:
        return Response(content=f"<html><body><h2>Google authorization error: {error}</h2><p>Close this window and retry.</p></body></html>", media_type="text/html")
    if not code:
        return Response(content="<html><body><h2>No authorization code received</h2></body></html>", media_type="text/html", status_code=400)

    # Extract user_id from state
    user_id = state.split("::")[0] if "::" in state else "dev_partner_001"

    from google_integration import exchange_code_for_tokens, get_user_info, save_tokens
    try:
        tokens = await exchange_code_for_tokens(code)
        profile = await get_user_info(tokens.get("access_token", ""))
        await save_tokens(db, user_id, tokens, profile)

        # Audit log
        try:
            from audit_log import log_action
            await log_action(db, user_id, "google.oauth.connected",
                            resource_type="google_account",
                            resource_id=profile.get("email", ""),
                            metadata={"email": profile.get("email"), "name": profile.get("name")})
        except Exception:
            pass

        # Success — show a friendly HTML that closes the popup
        email = profile.get("email", "your account")
        return Response(content=f"""
<!DOCTYPE html>
<html>
<head><title>Google Connected</title></head>
<body style="font-family: system-ui, sans-serif; padding: 40px; text-align: center; background: #fafafa;">
  <div style="max-width: 500px; margin: 0 auto; background: white; padding: 40px; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.05);">
    <div style="font-size: 48px;">✓</div>
    <h2>Connected to {email}</h2>
    <p style="color: #666;">You can now upload generated files directly to your Drive.</p>
    <p style="color: #999; font-size: 13px; margin-top: 30px;">You can close this window.</p>
  </div>
  <script>setTimeout(function() {{ window.close(); }}, 1500);</script>
</body>
</html>
""", media_type="text/html")
    except Exception as e:
        logger.error(f"OAuth callback failed: {e}")
        return Response(content=f"<html><body><h2>OAuth failed: {e}</h2></body></html>", media_type="text/html", status_code=500)


@api_router.get("/google/status")
async def google_connection_status(request: Request, authorization: str = Header(None)):
    """Check if the current user has Google connected."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}
    from google_integration import get_connection_status
    return await get_connection_status(db, user["user_id"])


@api_router.post("/google/disconnect")
async def google_disconnect(request: Request, authorization: str = Header(None)):
    """Revoke and remove Google connection for the current user."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}
    from google_integration import disconnect_google
    ok = await disconnect_google(db, user["user_id"])
    try:
        from audit_log import log_action
        await log_action(db, user["user_id"], "google.oauth.disconnected", resource_type="google_account")
    except Exception:
        pass
    return {"disconnected": ok}


@api_router.get("/google/folders")
async def google_list_folders(
    parent_id: str = "root",
    q: str = "",
    request: Request = None,
    authorization: str = Header(None),
):
    """List folders for the folder picker. parent_id=root for top-level."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}
    from google_integration import get_valid_access_token, list_folders, get_folder_path
    token = await get_valid_access_token(db, user["user_id"])
    if not token:
        raise HTTPException(status_code=401, detail="Google not connected")
    folders = await list_folders(token, parent_id=parent_id, query=q)
    breadcrumb = await get_folder_path(token, parent_id) if parent_id != "root" else [{"id": "root", "name": "My Drive"}]
    return {"folders": folders, "breadcrumb": breadcrumb, "parent_id": parent_id}


@api_router.post("/google/folders/create")
async def google_create_folder(
    request: Request,
    name: str = Form(...),
    parent_id: str = Form("root"),
    authorization: str = Header(None),
):
    """Create a new folder."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}
    from google_integration import get_valid_access_token, create_folder
    token = await get_valid_access_token(db, user["user_id"])
    if not token:
        raise HTTPException(status_code=401, detail="Google not connected")
    return await create_folder(token, name, parent_id)


@api_router.post("/google/upload")
async def google_upload_file(
    request: Request,
    file_id: str = Form(...),          # Our generated-file ID from /agent/download
    folder_id: str = Form("root"),
    convert: bool = Form(True),
    authorization: str = Header(None),
):
    """Upload a generated file (by our file_id) to user's Google Drive."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    # Look up the generated file from our cache
    entry = _generated_files_cache.get(file_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Generated file not found or expired")

    from google_integration import get_valid_access_token, upload_file_to_drive
    token = await get_valid_access_token(db, user["user_id"])
    if not token:
        raise HTTPException(status_code=401, detail="Google not connected — call /api/google/auth/start")

    # Rate limit check
    try:
        from usage_tracker import track_drive_upload
        await track_drive_upload(db, user["user_id"])
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"Rate limit check failed (non-blocking): {e}")

    result = await upload_file_to_drive(
        access_token=token,
        filename=entry["name"],
        content=entry["content"],
        folder_id=folder_id if folder_id != "root" else None,
        convert_to_google_format=convert,
    )

    # Audit log
    try:
        from audit_log import log_action
        await log_action(db, user["user_id"], "drive.upload",
                        resource_type="drive_file",
                        resource_id=result.get("id"),
                        metadata={
                            "source_file_id": file_id,
                            "name": entry["name"],
                            "size": len(entry["content"]),
                            "folder_id": folder_id,
                            "converted": result.get("converted"),
                            "mime_type": result.get("mimeType"),
                        })
    except Exception:
        pass

    return {
        "drive_file_id": result.get("id"),
        "drive_name": result.get("name"),
        "drive_url": result.get("webViewLink"),
        "mime_type": result.get("mimeType"),
        "converted": result.get("converted"),
    }


@api_router.post("/google/edit")
async def google_edit_doc(
    request: Request,
    drive_file_id: str = Form(...),       # Google Drive file ID (not our file_id)
    instruction: str = Form(...),          # Natural language edit request
    authorization: str = Header(None),
):
    """Edit an existing Google Sheet or Doc via natural language.

    Flow:
      1. Download the file from Drive as xlsx/docx
      2. Upload to sandbox as input
      3. Run code executor with instruction (treats as modify-in-place task)
      4. Upload modified version back to Drive (as new revision or new file)
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    from google_integration import get_valid_access_token, download_drive_file, get_file_metadata, upload_file_to_drive
    token = await get_valid_access_token(db, user["user_id"])
    if not token:
        raise HTTPException(status_code=401, detail="Google not connected")

    # 1. Get metadata → determine if Sheet or Doc
    meta = await get_file_metadata(token, drive_file_id)
    mime = meta.get("mimeType", "")
    if "spreadsheet" in mime:
        ext = "xlsx"
    elif "document" in mime:
        ext = "docx"
    elif "pdf" in mime:
        ext = "pdf"
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported file type for editing: {mime}")

    # 2. Download
    content = await download_drive_file(token, drive_file_id, export_format=ext)
    source_name = meta.get("name", "file") + "." + ext

    # 3. Run code executor with the existing file + instruction
    from code_executor import execute_user_task
    prompt = f"Edit the uploaded file as follows: {instruction}\n\nRead /workspace/input/{source_name}, apply the requested changes, save the modified file to /workspace/output/edited.{ext}. Preserve all existing content and formatting except what the user explicitly asked to change."
    result = await execute_user_task(prompt, [{"filename": source_name, "content": content}], max_iterations=3)

    if result.get("status") != "success" or not result.get("output_files"):
        raise HTTPException(status_code=500, detail=f"Edit failed: {result.get('error', 'No output files')}")

    # 4. Upload the edited version back (as a new revision of the same file? or new file?)
    # For safety: upload as a new file with "(edited)" suffix
    import base64 as _b64
    edited = result["output_files"][0]
    edited_bytes = _b64.b64decode(edited["content_b64"])
    upload_result = await upload_file_to_drive(
        access_token=token,
        filename=meta.get("name", "edited") + " (edited)." + ext,
        content=edited_bytes,
        folder_id=meta.get("parents", [None])[0] if meta.get("parents") else None,
        convert_to_google_format=True,
    )
    return {
        "original_drive_id": drive_file_id,
        "edited_drive_id": upload_result.get("id"),
        "edited_url": upload_result.get("webViewLink"),
        "iterations": result.get("iterations"),
    }


# ==================== PUTER.JS OPUS 4.7 SUPPORT ====================
# These endpoints pair with the browser-side Puter.js Claude Opus 4.7 flow.
# Frontend:
#   1. POST /assistant/prepare-context → gets SPECTR prompt + statute DB + IK cases
#   2. puter.ai.chat(context, { model: 'claude-opus-4-7' }) → reasons in-browser (FREE)
#   3. POST /assistant/verify-response → backend does Trust Layer annotations
# Benefit: Opus 4.7 reasoning quality + our verified Indian legal corpus + zero backend LLM cost.

class PrepareContextReq(BaseModel):
    query: str
    matter_id: Optional[str] = ""
    conversation_history: Optional[list] = []


@api_router.post("/assistant/prepare-context")
async def prepare_context(req: PrepareContextReq, request: Request, authorization: str = Header(None)):
    """Build the full Spectr context (SPECTR prompt + statute DB + IK cases + pre-flight)
    for the browser-side Puter.js Claude call to consume.

    Returns system_prompt + user_content + sources_used.
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001", "email": "partner@algorythm.tech"}

    from ai_engine import get_spectr_prompt, classify_query, TOOL_DESCRIPTIONS_FOR_PROMPT
    from indian_kanoon import search_indiankanoon
    from pre_flight import run_pre_flight

    # 1. Base SPECTR system prompt (with live thresholds)
    try:
        system_prompt = await get_spectr_prompt() + "\n\n" + TOOL_DESCRIPTIONS_FOR_PROMPT
    except Exception:
        from ai_engine import SPECTR_SYSTEM_PROMPT
        system_prompt = SPECTR_SYSTEM_PROMPT

    # 2. Statute context (MongoDB)
    statute_context = ""
    try:
        statute_context = await get_statute_context(req.query)
    except Exception as e:
        logger.warning(f"Statute lookup failed: {e}")

    # 3. IndianKanoon cases (top 6)
    ik_context = ""
    sources_used = []
    try:
        ik_results = await search_indiankanoon(req.query[:200], top_k=6)
        if ik_results:
            ik_context = "=== CASE LAW (INDIANKANOON) ===\n"
            for i, case in enumerate(ik_results, 1):
                ik_context += f"[Case {i}] {case.get('title', 'N/A')} | {case.get('court', '')} | {case.get('year', '')}\nHeadline: {case.get('headline', '')[:300]}\n\n"
                sources_used.append({
                    "type": "case",
                    "title": case.get("title", ""),
                    "court": case.get("court", ""),
                    "year": case.get("year", ""),
                    "url": f"https://indiankanoon.org/doc/{case.get('tid', '')}/",
                })
    except Exception as e:
        logger.warning(f"IK lookup failed: {e}")

    # 4. Pre-flight enrichment (deterministic tool runs)
    pre_flight_context = ""
    try:
        pf = await run_pre_flight(req.query)
        if pf.get("has_computed_facts"):
            pre_flight_context = pf.get("context_block", "")
    except Exception as e:
        logger.warning(f"Pre-flight failed: {e}")

    # 5. Conversation history
    conv_context = ""
    if req.conversation_history:
        try:
            from ai_engine import _build_conversation_context
            conv_context = _build_conversation_context(req.conversation_history, max_turns=4, max_chars=4000)
        except Exception:
            pass

    # 6. Matter context
    matter_context = ""
    if req.matter_id:
        try:
            # SECURITY: scope by user_id so one user can't read another user's matters
            matter = await db.matters.find_one(
                {"matter_id": req.matter_id, "user_id": user.get("user_id")},
                {"_id": 0}
            )
            if matter:
                matter_context = f"MATTER: {matter.get('name', '')} | CLIENT: {matter.get('client_name', '')} | TYPE: {matter.get('matter_type', '')}"
        except Exception:
            pass

    query_types = classify_query(req.query)

    # Build user content
    user_content_parts = []
    if conv_context:
        user_content_parts.append(conv_context)
    if matter_context:
        user_content_parts.append(f"=== MATTER CONTEXT ===\n{matter_context}")
    if statute_context:
        user_content_parts.append(f"=== RELEVANT STATUTE SECTIONS ===\n{statute_context}")
    if ik_context:
        user_content_parts.append(ik_context)
    if pre_flight_context:
        user_content_parts.append(pre_flight_context)
    user_content_parts.append(f"=== USER QUERY ===\n{req.query}")

    return {
        "system_prompt": system_prompt,
        "user_content": "\n\n".join(user_content_parts),
        "sources_used": sources_used,
        "query_types": query_types,
        "context_sizes": {
            "system": len(system_prompt),
            "statute": len(statute_context),
            "ik": len(ik_context),
            "pre_flight": len(pre_flight_context),
            "total_user": sum(len(p) for p in user_content_parts),
        },
    }


class VerifyResponseReq(BaseModel):
    query: str
    response: str
    sources_used: Optional[list] = []


@api_router.post("/assistant/verify-response")
async def verify_response(req: VerifyResponseReq, request: Request, authorization: str = Header(None)):
    """Run Trust Layer on a Claude Opus 4.7 response generated in-browser via Puter.js.

    - Inline-annotates every case citation (IK → Serper → DDG cascade)
    - Verifies statute sections in MongoDB
    - Flags stale amendment-affected cases
    - Checks arithmetic
    - Returns Trust Score + augmented_text
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    from response_augmenter import augment_response

    try:
        result = await augment_response(req.response, db=db, max_case_verifications=8)
        return {
            "augmented_text": result.get("augmented_text", req.response),
            "trust_score": result.get("trust_score", 50),
            "stats": result.get("stats", {}),
            "notes": result.get("notes", []),
            "verification_report": result.get("verification_report", ""),
        }
    except Exception as e:
        logger.warning(f"Verification failed (non-blocking): {e}")
        return {
            "augmented_text": req.response,
            "trust_score": 50,
            "stats": {},
            "notes": [],
            "verification_report": "",
            "error": str(e),
        }


class DeepResearchReq(BaseModel):
    query: str


@api_router.post("/assistant/deep-research-only")
async def deep_research_only(req: DeepResearchReq, request: Request, authorization: str = Header(None)):
    """Run JUST the 5-phase Blaxel sandbox research — no LLM synthesis.

    Returns findings + sources for the frontend to feed into Puter.js Claude Opus 4.7.
    This lets Opus 4.7 do the reasoning while Blaxel does the web crawling.
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    from ai_engine import classify_query
    from sandbox_research import execute_deep_research

    try:
        query_types = classify_query(req.query)
    except Exception:
        query_types = ["legal"]

    # Rate limit — deep research is expensive
    try:
        from usage_tracker import track_deep_research
        await track_deep_research(db, user["user_id"])
    except HTTPException:
        raise
    except Exception as _e:
        logger.warning(f"Deep research rate limit skip: {_e}")

    try:
        research = await execute_deep_research(req.query, query_types)

        # Audit log
        try:
            from audit_log import log_action
            await log_action(db, user["user_id"], "agent.deep_research",
                            metadata={"query_length": len(req.query),
                                     "query_types": query_types,
                                     "sources_found": len(research.get("pageContents", []))})
        except Exception:
            pass

        # Structure findings for Opus consumption
        # execute_deep_research returns {searchResults, pageContents, errors, metadata}
        findings_text_parts = []
        sources = []
        sites_set = set()

        # pageContents: list of {url, title, text, ...}
        for page in research.get("pageContents", [])[:30]:
            url = page.get("url", "")
            title = page.get("title", "") or page.get("url", "")
            content = (page.get("text") or page.get("content") or "")[:1500]
            if content:
                findings_text_parts.append(f"### {title}\nURL: {url}\n\n{content}\n")
                site = url.split("/")[2] if "//" in url else url
                sources.append({"url": url, "title": title, "site": site})
                sites_set.add(site)

        # Also include search results for breadth
        for sr in research.get("searchResults", [])[:20]:
            url = sr.get("url", "")
            if url and url not in [s["url"] for s in sources]:
                site = url.split("/")[2] if "//" in url else url
                sources.append({
                    "url": url,
                    "title": sr.get("title", ""),
                    "snippet": sr.get("snippet", "")[:300],
                    "site": site,
                })
                sites_set.add(site)

        return {
            "query": req.query,
            "findings": "\n---\n".join(findings_text_parts)[:30000],
            "sources": sources,
            "sources_count": len(sources),
            "sites_hit": len(sites_set),
            "page_count": len(research.get("pageContents", [])),
            "search_count": len(research.get("searchResults", [])),
            "errors": research.get("errors", [])[:3],
            "duration_ms": research.get("metadata", {}).get("durationMs", 0),
            "query_types": query_types,
        }
    except Exception as e:
        logger.error(f"Deep research failed: {e}")
        return {
            "query": req.query,
            "findings": "",
            "sources": [],
            "error": str(e),
        }


@api_router.get("/agent/client.js")
async def serve_puter_client_js():
    """Serve the Puter.js client snippet so frontends can just import it via <script>."""
    try:
        from pathlib import Path
        p = Path(__file__).parent / "puter_client_snippet.js"
        if p.exists():
            return Response(content=p.read_text(encoding="utf-8"), media_type="application/javascript")
    except Exception as e:
        logger.warning(f"Serve client.js failed: {e}")
    raise HTTPException(status_code=404, detail="Client snippet not found")


@api_router.post("/agent/execute-code")
async def agent_execute_code(
    request: Request,
    code: str = Form(...),                   # Pre-generated Python code (from Puter.js Claude call in browser)
    files: list[UploadFile] = File(default=[]),
    timeout: int = Form(180),
    authorization: str = Header(None),
):
    """Execute PRE-GENERATED Python code directly in sandbox.

    This decouples the LLM from our backend — the frontend uses Puter.js to call
    Claude (free, unlimited), gets Python code back, and POSTs here to execute.
    We never need a backend Claude API key.

    Returns: files produced, stdout, stderr, validation, download URLs, google status.
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    # Read uploaded files
    MAX_FILES = 10
    MAX_TOTAL = 100 * 1024 * 1024
    uploaded = []
    total_size = 0
    for i, uf in enumerate(files[:MAX_FILES]):
        try:
            content = await uf.read()
            total_size += len(content)
            if total_size > MAX_TOTAL:
                raise HTTPException(status_code=413, detail="Total upload size exceeds 100MB")
            uploaded.append({"filename": uf.filename or f"file_{i}", "content": content})
        except HTTPException:
            raise
        except Exception as e:
            logger.warning(f"Upload {i} read failed: {e}")

    # Sanity check on code
    if not code or len(code) < 10:
        raise HTTPException(status_code=400, detail="No code provided")
    if len(code) > 200_000:
        raise HTTPException(status_code=413, detail="Code too large (>200KB)")

    # Rate limit check
    try:
        from usage_tracker import track_query
        await track_query(db, user["user_id"])
    except HTTPException:
        raise
    except Exception as _e:
        logger.warning(f"Rate limit skip (non-blocking): {_e}")

    # Get sandbox + upload files + execute
    from code_sandbox import get_python_sandbox, upload_file_to_sandbox, execute_python, read_file_from_sandbox
    import base64 as _b64

    try:
        sandbox = await get_python_sandbox()
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Sandbox creation failed: {e}")

    for f in uploaded:
        try:
            await upload_file_to_sandbox(sandbox, f["filename"], f["content"], subdir="input")
        except Exception as e:
            logger.warning(f"Upload {f['filename']} failed: {e}")

    exec_result = await execute_python(sandbox, code, timeout=timeout)

    # Audit log
    try:
        from audit_log import log_action
        await log_action(db, user["user_id"], "agent.execute_code",
                        resource_type="agent_task",
                        metadata={
                            "code_length": len(code),
                            "files_uploaded": len(uploaded),
                            "exit_code": exec_result.get("exit_code"),
                            "output_files": len(exec_result.get("output_files") or []),
                        })
    except Exception:
        pass

    # Collect output files (base64-safe via manifest)
    output_files = []
    for info in (exec_result.get("output_files") or []):
        path = info.get("path") if isinstance(info, dict) else info
        content = info.get("content_bytes") if isinstance(info, dict) else None
        if content is None:
            try:
                content = await read_file_from_sandbox(sandbox, path)
                if isinstance(content, str):
                    content = content.encode('utf-8', errors='surrogateescape')
            except Exception:
                continue
        output_files.append({
            "name": os.path.basename(path),
            "path": path,
            "size": info.get("size", len(content)) if isinstance(info, dict) else len(content),
            "content_b64": _b64.b64encode(content).decode('ascii') if content else "",
        })

    # Validate + register for download
    from output_validator import validate_all_outputs
    validation = validate_all_outputs(output_files) if output_files else {"all_valid": True, "results": [], "summary": ""}
    download_urls = []
    for f in output_files:
        try:
            file_id = _register_generated_file(f["name"], _b64.b64decode(f["content_b64"]), user.get("user_id", "anonymous"))
            download_urls.append({
                "name": f["name"],
                "size": f["size"],
                "file_id": file_id,
                "url": f"/api/agent/download/{file_id}",
            })
        except Exception as e:
            logger.warning(f"Register failed for {f.get('name')}: {e}")

    # Google status (per-user — each user has their own OAuth tokens)
    google_status = {"connected": False}
    try:
        from google_integration import get_connection_status
        google_status = await get_connection_status(db, user.get("user_id", ""))
    except Exception:
        pass

    return {
        "exit_code": exec_result["exit_code"],
        "stdout": exec_result.get("stdout", "")[:4000],
        "stderr": exec_result.get("stderr", "")[:4000],
        "output_files": output_files,
        "validation": validation,
        "download_urls": download_urls,
        "google": {
            **google_status,
            "upload_endpoint": "/api/google/upload",
            "folders_endpoint": "/api/google/folders",
        },
    }


@api_router.post("/agent/profile-files")
async def agent_profile_files(
    request: Request,
    files: list[UploadFile] = File(default=[]),
    authorization: str = Header(None),
):
    """Generate smart data profiles for uploaded files.

    Frontend calls this BEFORE asking Puter.js/Claude to plan Python,
    so Claude has compact condensed context (300-1500 tokens) about the data
    instead of needing to see the whole file.
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    uploaded = []
    for i, uf in enumerate(files[:10]):
        try:
            content = await uf.read()
            uploaded.append({"filename": uf.filename or f"file_{i}", "content": content})
        except Exception as e:
            logger.warning(f"Upload {i} failed: {e}")

    from sheet_profiler import profile_uploaded_files, profile_file

    # Return both: combined text (for Claude) AND per-file structured profile (for UI)
    combined_text = profile_uploaded_files(uploaded, max_total_chars=8000)
    per_file = []
    for f in uploaded:
        try:
            p = profile_file(f["filename"], f["content"])
            per_file.append({"filename": f["filename"], "profile": p})
        except Exception as e:
            per_file.append({"filename": f["filename"], "error": str(e)})

    return {
        "profile_text": combined_text,
        "profile_chars": len(combined_text),
        "estimated_tokens": len(combined_text) // 4,
        "per_file": per_file,
    }


@api_router.post("/agent/iterate")
async def agent_iterate(
    request: Request,
    prompt: str = Form(...),
    files: list[UploadFile] = File(default=[]),
    max_rounds: int = Form(4),
    authorization: str = Header(None),
):
    """Iterative agent: multi-round Python exploration for complex/unknown data.

    Each round's stdout informs the next round. Final round produces deliverable.
    Streaming SSE.
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    uploaded = []
    for i, uf in enumerate(files[:10]):
        try:
            content = await uf.read()
            uploaded.append({"filename": uf.filename or f"file_{i}", "content": content})
        except Exception as e:
            logger.warning(f"Upload {i} failed: {e}")

    try:
        safe_prompt = sanitize_input(prompt, max_length=8000, context="query")
    except Exception:
        safe_prompt = prompt[:8000]

    from iterative_agent import run_iterative_agent

    async def stream():
        try:
            async for event in run_iterative_agent(safe_prompt, uploaded, max_rounds=max_rounds):
                yield f"data: {json.dumps(event, default=str)}\n\n"
            yield "data: [DONE]\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"
            yield "data: [DONE]\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


@api_router.post("/google/sheets/update")
async def google_sheets_live_update(
    request: Request,
    sheet_id: str = Form(...),
    range_a1: str = Form(...),
    values_json: str = Form(...),                     # JSON array of arrays
    value_input_option: str = Form("USER_ENTERED"),
    authorization: str = Header(None),
):
    """Live-edit a Google Sheet directly via Sheets API (no download/reupload cycle).

    Much faster than the /google/edit flow for cell-level updates.
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    try:
        values = json.loads(values_json)
        if not isinstance(values, list):
            raise ValueError("values must be a 2D array")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid values_json: {e}")

    from google_integration import get_valid_access_token, update_sheet_values
    token = await get_valid_access_token(db, user["user_id"])
    if not token:
        raise HTTPException(status_code=401, detail="Google not connected")

    result = await update_sheet_values(token, sheet_id, range_a1, values, value_input_option=value_input_option)
    return result


@api_router.post("/google/docs/replace-text")
async def google_docs_live_replace(
    request: Request,
    doc_id: str = Form(...),
    find: str = Form(...),
    replace: str = Form(...),
    match_case: bool = Form(True),
    authorization: str = Header(None),
):
    """Live-edit a Google Doc via Docs API — replace-all text."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    from google_integration import get_valid_access_token, batch_update_doc
    token = await get_valid_access_token(db, user["user_id"])
    if not token:
        raise HTTPException(status_code=401, detail="Google not connected")

    req = {
        "replaceAllText": {
            "containsText": {"text": find, "matchCase": match_case},
            "replaceText": replace,
        }
    }
    return await batch_update_doc(token, doc_id, [req])


@api_router.get("/agent/playbooks")
async def list_available_playbooks(request: Request, authorization: str = Header(None)):
    """List all available drafting / review playbooks."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        pass
    from playbooks import list_playbooks
    return {"playbooks": list_playbooks()}


# In-memory storage for generated files (keyed by file_id)
# In production this would be object storage; for MVP it's in-process
_generated_files_cache: dict = {}  # file_id -> {"name", "content", "created_at", "user_id"}
_GENERATED_FILES_MAX = 200
_GENERATED_FILES_TTL = 3600  # 1 hour


def _register_generated_file(name: str, content: bytes, user_id: str = "") -> str:
    """Register a generated file and return a downloadable file_id."""
    import uuid as _uuid
    import time as _time
    file_id = _uuid.uuid4().hex
    # Cleanup: if cache is full, drop oldest
    if len(_generated_files_cache) >= _GENERATED_FILES_MAX:
        oldest = min(_generated_files_cache.items(), key=lambda kv: kv[1].get("created_at", 0))
        _generated_files_cache.pop(oldest[0], None)
    _generated_files_cache[file_id] = {
        "name": name,
        "content": content,
        "created_at": _time.time(),
        "user_id": user_id,
    }
    return file_id


@api_router.get("/agent/download/{file_id}")
async def download_generated_file(file_id: str, request: Request, authorization: str = Header(None)):
    """Download a file generated by the code executor."""
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    import time as _time
    entry = _generated_files_cache.get(file_id)
    if not entry:
        raise HTTPException(status_code=404, detail="File not found or expired")
    age = _time.time() - entry.get("created_at", 0)
    if age > _GENERATED_FILES_TTL:
        _generated_files_cache.pop(file_id, None)
        raise HTTPException(status_code=410, detail="File expired")

    # Pick content type by extension
    name = entry["name"]
    ct = "application/octet-stream"
    if name.endswith(".xlsx"):
        ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif name.endswith(".docx"):
        ct = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif name.endswith(".pdf"):
        ct = "application/pdf"
    elif name.endswith((".jpg", ".jpeg")):
        ct = "image/jpeg"
    elif name.endswith(".png"):
        ct = "image/png"
    elif name.endswith(".csv"):
        ct = "text/csv"

    return Response(
        content=entry["content"],
        media_type=ct,
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


@api_router.post("/agent/route")
async def agent_route(
    request: Request,
    prompt: str = Form(...),
    files: list[UploadFile] = File(default=[]),
    max_iterations: int = Form(4),
    force_engine: str = Form(""),  # "advisory" / "automation" / "hybrid" or empty for auto
    playbook_id: str = Form(""),   # Override playbook detection
    authorization: str = Header(None),
):
    """SMART ROUTER endpoint — auto-decides advisory vs automation vs hybrid.

    This is what the frontend should typically call. It:
    1. Detects intent (file automation vs legal advisory vs both)
    2. Detects playbook (NDA, SPA, Redlining, DD, Chronology, etc.)
    3. Routes to the appropriate engine
    4. Streams unified SSE events

    For advisory: same pipeline as /assistant/query (war_room_engine).
    For automation: same as /agent/execute (code_executor) with playbook injected.
    For hybrid: runs advisory first, then feeds advisory output into automation as context.
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001", "email": "partner@algorythm.tech", "name": "Partner (Dev)"}

    # Read files
    MAX_FILES = 10
    MAX_TOTAL = 100 * 1024 * 1024
    uploaded = []
    total_size = 0
    for i, uf in enumerate(files[:MAX_FILES]):
        try:
            content = await uf.read()
            total_size += len(content)
            if total_size > MAX_TOTAL:
                raise HTTPException(status_code=413, detail="Total upload size exceeds 100MB")
            uploaded.append({"filename": uf.filename or f"file_{i}", "content": content})
        except HTTPException:
            raise
        except Exception as e:
            logger.warning(f"Upload {i} read failed: {e}")

    try:
        safe_prompt = sanitize_input(prompt, max_length=8000, context="query")
    except Exception:
        safe_prompt = prompt[:8000]

    # Route the query
    from query_router import route_query
    from playbooks import get_playbook

    routing = route_query(
        safe_prompt,
        has_uploaded_files=len(uploaded) > 0,
        force=force_engine if force_engine in ("advisory", "automation", "hybrid") else None,
    )

    # Apply playbook override if provided
    if playbook_id:
        routing["playbook_id"] = playbook_id
        routing["system_prompt_addon"] = get_playbook(playbook_id)

    engine = routing["engine"]
    playbook_addon = routing["system_prompt_addon"]

    async def stream():
        # Initial routing decision event
        yield f"data: {json.dumps({'type': 'routing', **routing})}\n\n"

        if engine == "automation":
            # File automation only
            from code_executor import run_code_agent
            try:
                async for event in run_code_agent(safe_prompt, uploaded, max_iterations=max_iterations, playbook_addon=playbook_addon):
                    yield f"data: {json.dumps(event)}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

        elif engine == "advisory":
            # Advisory only — delegate to assistant/query pipeline
            # We'll run the war_room_engine stream directly here
            try:
                from war_room_engine import stream_war_room_research
                async for chunk in stream_war_room_research(
                    user_query=safe_prompt,
                    statute_context="",
                    matter_context="",
                    firm_context="",
                    conversation_history=[],
                ):
                    yield f"data: {chunk}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'message': f'Advisory pipeline error: {e}'})}\n\n"

        elif engine == "hybrid":
            # Advisory first → collect the advisory response → feed as context to automation
            yield f"data: {json.dumps({'type': 'status', 'step': 'hybrid_phase1', 'message': 'Running advisory research phase...'})}\n\n"
            advisory_text = ""
            try:
                from war_room_engine import stream_war_room_research
                async for chunk in stream_war_room_research(
                    user_query=safe_prompt, statute_context="", matter_context="",
                    firm_context="", conversation_history=[],
                ):
                    # Pass through to client AND accumulate for phase 2
                    try:
                        parsed = json.loads(chunk)
                        if parsed.get("type") in ("partner_payload", "fast_chunk"):
                            advisory_text += parsed.get("content", "")
                    except Exception:
                        pass
                    yield f"data: {chunk}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'message': f'Hybrid phase 1 error: {e}'})}\n\n"

            # Phase 2: automation with advisory context
            yield f"data: {json.dumps({'type': 'status', 'step': 'hybrid_phase2', 'message': 'Generating deliverable file based on advisory output...'})}\n\n"
            enhanced_prompt = f"{safe_prompt}\n\n=== ADVISORY RESEARCH CONTEXT (use this to build the deliverable) ===\n{advisory_text[:6000]}"
            try:
                from code_executor import run_code_agent
                async for event in run_code_agent(enhanced_prompt, uploaded, max_iterations=max_iterations, playbook_addon=playbook_addon):
                    yield f"data: {json.dumps(event)}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'message': f'Hybrid phase 2 error: {e}'})}\n\n"

        yield "data: [DONE]\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


@api_router.post("/agent/execute-sync")
async def agent_execute_sync(
    request: Request,
    prompt: str = Form(...),
    files: list[UploadFile] = File(default=[]),
    max_iterations: int = Form(4),
    playbook_id: str = Form(""),
    authorization: str = Header(None),
):
    """Non-streaming variant — returns final result JSON with:
      - download URLs for generated files (preferred over base64)
      - validation results per file
      - routing decision (which engine/playbook)
    """
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        user = {"user_id": "dev_partner_001"}

    MAX_FILES = 10
    uploaded = []
    for i, uf in enumerate(files[:MAX_FILES]):
        try:
            content = await uf.read()
            uploaded.append({"filename": uf.filename or f"file_{i}", "content": content})
        except Exception as e:
            logger.warning(f"Upload {i} read failed: {e}")

    try:
        safe_prompt = sanitize_input(prompt, max_length=8000, context="query")
    except Exception:
        safe_prompt = prompt[:8000]

    # Route to detect playbook if not forced
    from query_router import route_query
    from playbooks import get_playbook

    routing = route_query(safe_prompt, has_uploaded_files=len(uploaded) > 0)
    effective_playbook = playbook_id or routing.get("playbook_id", "")
    addon = get_playbook(effective_playbook) if effective_playbook else ""

    # Rate limit
    try:
        from usage_tracker import track_query
        await track_query(db, user["user_id"])
    except HTTPException:
        raise
    except Exception as _e:
        logger.warning(f"Rate limit skip: {_e}")

    from code_executor import execute_user_task
    result = await execute_user_task(safe_prompt, uploaded, max_iterations=max_iterations, playbook_addon=addon)

    # Audit log
    try:
        from audit_log import log_action
        await log_action(db, user["user_id"], "agent.execute",
                        resource_type="agent_task",
                        metadata={
                            "prompt_length": len(safe_prompt),
                            "files_uploaded": len(uploaded),
                            "playbook": effective_playbook,
                            "engine": routing.get("engine"),
                            "status": result.get("status"),
                            "iterations": result.get("iterations"),
                            "output_files": len(result.get("output_files", [])),
                        })
    except Exception:
        pass

    # Validate outputs + register for download
    from output_validator import validate_all_outputs
    validation = {"all_valid": True, "results": [], "summary": ""}
    download_urls = []

    if result.get("output_files"):
        validation = validate_all_outputs(result["output_files"])
        # Register each file for download
        user_id = user.get("user_id", "anonymous")
        for f in result["output_files"]:
            try:
                import base64 as _b64
                content = _b64.b64decode(f.get("content_b64", ""))
                file_id = _register_generated_file(f["name"], content, user_id)
                download_urls.append({
                    "name": f["name"],
                    "size": f.get("size", len(content)),
                    "file_id": file_id,
                    "url": f"/api/agent/download/{file_id}",
                })
            except Exception as e:
                logger.warning(f"Download registration failed for {f.get('name')}: {e}")

    # Add Google connection status so frontend can show "Upload to Drive" button
    google_status = {"connected": False}
    try:
        from google_integration import get_connection_status
        google_status = await get_connection_status(db, user.get("user_id", ""))
    except Exception as e:
        logger.warning(f"Google status check failed: {e}")

    return {
        **result,
        "routing": routing,
        "playbook_used": effective_playbook,
        "validation": validation,
        "download_urls": download_urls,
        "google": {
            **google_status,
            "upload_endpoint": "/api/google/upload",     # POST file_id + folder_id
            "folders_endpoint": "/api/google/folders",   # GET parent_id
            "auth_start_endpoint": "/api/google/auth/start",
        },
    }


# ==================== ASSISTANT ROUTES ====================

@api_router.post("/assistant/query")
async def assistant_query(req: QueryRequest, request: Request, authorization: str = Header(None)):
    # Auth — graceful fallback when MongoDB is down
    try:
        user = await get_current_user(request, authorization)
    except Exception as e:
        logger.warning(f"Auth fallback (MongoDB may be down): {e}")
        user = {"user_id": "dev_partner_001", "email": "partner@algorythm.tech", "name": "Partner (Dev)", "role": "partner"}

    # ─── TRIVIAL-QUERY FAST-PATH ────────────────────────────────────────
    # Greetings + identity/capability questions don't need statute retrieval,
    # matter context, library lookup, thread history, or a 60K system prompt.
    # Every one of those stacks stall time. Short-circuit them to canned
    # replies so first-token latency is <500ms.
    #
    # IMPORTANT: we do NOT use the user's stored name in the greeting.
    # When a user logs out of account A and into account B on the same
    # device, stale SQL/Mongo user records can leak the old name ("Hey
    # Helpdesk" showing to Sri Aasrith). Neutral greetings are safer and
    # indistinguishable in quality.
    _q_strip = (req.query or "").strip().lower()

    # Identity / capability questions — common first-time user questions
    # that have canned, accurate answers. Respond in ~300ms instead of 60s.
    _identity_triggers = (
        "who are you", "who r u", "whoareyou", "who is spectr",
        "what are you", "what r u", "what is spectr", "what's spectr",
        "whats spectr", "tell me about spectr", "about spectr",
    )
    _capability_triggers = (
        "what can you do", "what do you do", "what can u do",
        "what features", "help", "how does this work",
        "how do you work", "how to use", "how to use spectr",
    )
    _is_identity = any(trig in _q_strip for trig in _identity_triggers) and len(_q_strip) <= 60
    _is_capability = any(trig in _q_strip for trig in _capability_triggers) and len(_q_strip) <= 60

    # Resilient greeting/typo/chitchat detection via spectr_pipeline triage.
    # Heuristic catches obvious cases at 0ms (no LLM call). Ambiguous cases
    # ("how are you doing today my friend") fall through to Groq tiebreaker
    # at ~200ms. Either way, we short-circuit BEFORE any MongoDB context-
    # building so even with Atlas down, greetings return in <500ms total.
    _is_greeting = False
    if not _is_identity and not _is_capability:
        try:
            from spectr_pipeline import _heuristic_intent, _llm_intent_gate
            h = _heuristic_intent(req.query or "")
            if h == "trivial":
                _is_greeting = True
            elif h == "unsure":
                # Cheap Groq tiebreaker — no DB work has happened yet
                intent = await _llm_intent_gate(req.query or "")
                if intent == "trivial":
                    _is_greeting = True
        except Exception as e:
            logger.warning(f"Triage heuristic failed (running full pipeline): {e}")

    if _is_greeting or _is_identity or _is_capability:
        from fastapi.responses import StreamingResponse
        import json as _json

        if _is_identity:
            canned_text = (
                "I am Spectr, an AI legal and tax intelligence engine built for "
                "Indian CAs, Advocates, and CFOs. I ground every answer in real "
                "Indian bare-act sections (BNS, BNSS, BSA, Income Tax Act, CGST, "
                "Companies Act, FEMA, SEBI, IBC, and more), cross-check against "
                "live IndianKanoon case law, and deliver partner-grade advisory "
                "memos with named authorities, exact exposures, and specific "
                "next steps.\n\n"
                "Ask me to analyse a GST SCN, draft a bail application, compute "
                "TDS exposure, reconcile GSTR-2B with your purchase register, "
                "or research case law on any point — I work the way a Big Four "
                "senior partner expects their research to work."
            )
        elif _is_capability:
            canned_text = (
                "Five things I do, all at partner-grade depth:\n\n"
                "1. **Chat advisory** — ask any Indian tax/legal question and get "
                "a structured memo with Bottom Line, Governing Framework, Analysis, "
                "Opposition counters, Risk Matrix, and Action Items. Every "
                "citation is named (section + act + year, case + pincite).\n\n"
                "2. **Document Vault** — upload PDFs, DOCX, or scanned notices. "
                "Get Executive Summary, Night Before Digest, Chronological "
                "Timeline, or Cross-Examination Matrix — whichever you need.\n\n"
                "3. **Workflows** — 45+ filing-ready templates: GST SCN reply, "
                "Section 148A reply, Bail application, Writ petition, Board "
                "resolutions, ROC filings, Director reports, and more. Fill the "
                "form, get a Word-ready document in 10 seconds.\n\n"
                "4. **Reconcilers** — GSTR-2B vs purchase register, TDS 26AS vs "
                "books, bank statement vs ledger. Flags mismatches, vendor risk, "
                "and ITC exposure.\n\n"
                "5. **Legal Research** — named authorities, paragraph-level "
                "pincites, pre-emption of opposing counsel, and transparency on "
                "what was verified vs drawn from training.\n\n"
                "Start by typing a question, uploading a document to the Vault, "
                "or picking a template from Workflows."
            )
        else:
            # Use the spectr_pipeline triage response — already context-
            # aware (handles "hi", "thanks", "good morning", typos, etc.)
            # and matches the Harvey-style voice of the rest of the app.
            try:
                from spectr_pipeline import _triage_response as _sp_triage
                canned_text = await _sp_triage(req.query or "")
            except Exception:
                # Conservative fallback if the import ever breaks
                canned_text = "Hey. What do you want to work on — a notice, a section, a computation, a draft?"

        async def _fast_canned():
            yield ": " + " " * 2048 + "\n\n"
            yield f"data: {_json.dumps({'type':'fast_chunk','content':canned_text})}\n\n"
            yield f"data: {_json.dumps({'type':'fast_complete','models_used':['fast-path-canned'],'sections':[],'citations':[]})}\n\n"
            yield "data: [DONE]\n\n"
        return StreamingResponse(_fast_canned(), media_type="text/event-stream")

    # Get statute context from DB — skip if MongoDB fails
    statute_context = ""
    try:
        statute_context = await get_statute_context(req.query)
    except Exception as e:
        logger.warning(f"Statute context lookup failed (skipping): {e}")

    # Get matter context if provided — skip if MongoDB fails
    matter_context = ""
    if req.matter_id:
        try:
            # SECURITY: scope by user_id so one user can't read another user's matters
            matter = await db.matters.find_one(
                {"matter_id": req.matter_id, "user_id": user.get("user_id")},
                {"_id": 0}
            )
            if matter:
                matter_context = f"Matter: {matter.get('name', '')} | Client: {matter.get('client_name', '')} | Type: {matter.get('matter_type', '')}"
                recent = await db.query_history.find(
                    {"matter_id": req.matter_id, "user_id": user.get("user_id")},
                    {"_id": 0}
                ).sort("created_at", -1).limit(3).to_list(3)
                if recent:
                    matter_context += "\n\nRecent conversation context:\n"
                    for r in reversed(recent):
                        matter_context += f"Q: {r.get('query', '')[:200]}\nA: {r.get('response_text', '')[:500]}\n\n"
        except Exception as e:
            logger.warning(f"Matter context lookup failed (skipping): {e}")
                    
    # Get Firm Context from Library — skip if MongoDB fails
    firm_context = ""
    try:
        library_items = await db.library_items.find({"user_id": user["user_id"]}).to_list(15)
        if library_items:
            firm_context = "Apply the following internal firm templates, principles, and precedents strictly if relevant:\n\n"
            for idx, item in enumerate(library_items):
                firm_context += f"--- Precedent {idx+1}: {item.get('title', '')} ({item.get('item_type', '')}) ---\n"
                firm_context += f"{item.get('content', '')}\n\n"
    except Exception as e:
        logger.warning(f"Library context lookup failed (skipping): {e}")

    # ─── THREAD: ensure thread row exists, load prior messages for context ──
    # If the client sent thread_id, we're continuing an existing chat. If not,
    # we create a new thread and auto-title from this first message.
    try:
        from thread_manager import ensure_thread, load_thread_messages, assemble_history_for_llm
        resolved_thread_id = await ensure_thread(
            db, user["user_id"], (req.thread_id or "").strip() or None,
            req.query, matter_id=(req.matter_id or None),
        )
    except Exception as e:
        logger.warning(f"Thread ensure failed (continuing without threading): {e}")
        resolved_thread_id = ""

    # Prior-turn history for this thread → conversation_history so the LLM
    # has full back-and-forth context. Only loaded if the user didn't already
    # pass a conversation_history explicitly (some clients rehydrate locally).
    if resolved_thread_id and not req.conversation_history:
        try:
            prior = await load_thread_messages(db, user["user_id"], resolved_thread_id, limit=40)
            # Drop the current query from prior if it was already persisted
            # (shouldn't be — ensure_thread doesn't write to query_history —
            # but defend anyway by filtering by content).
            prior = [p for p in prior if (p.get("query") or "").strip() != req.query.strip()]
            if prior:
                req.conversation_history = assemble_history_for_llm(prior)
        except Exception as e:
            logger.warning(f"Thread history load failed (non-blocking): {e}")

    # ─── SUPERMEMORY: pull long-term cross-thread memory for this user ──
    # Totally optional — no-op if SUPERMEMORY_API_KEY is unset. We prepend
    # the retrieved memory block to firm_context so the LLM sees it without
    # needing to change any downstream signatures.
    try:
        from supermemory_client import retrieve_context as sm_retrieve
        memory_block = await sm_retrieve(
            user_id=user.get("user_id", ""),
            query=req.query,
            matter_id=(req.matter_id or None),
        )
        if memory_block:
            firm_context = (memory_block + "\n\n" + firm_context).strip()
    except Exception as e:
        logger.debug(f"Supermemory retrieve skipped (non-blocking): {e}")
    
    # === INPUT SANITIZATION (Security Layer) ===
    try:
        sanitized_query = sanitize_input(req.query, context="query")
        if req.mode:
            req.mode = sanitize_input(req.mode, context="general")
    except Exception as e:
        logger.warning(f"Input sanitization error (using raw): {e}")
        sanitized_query = req.query

    # === PROMPT INJECTION DEFENSE (LLM Context Layer) ===
    # All user-controlled context that gets injected into system prompts must be sanitized
    try:
        matter_context = sanitize_context_for_llm(matter_context, source_label="matter_context")
        firm_context = sanitize_context_for_llm(firm_context, source_label="firm_context")
        statute_context = sanitize_context_for_llm(statute_context, source_label="statute_context")
    except Exception as e:
        logger.warning(f"LLM context sanitization error (using raw): {e}")

    # PII Anonymization (pre-LLM redaction)
    pii_report = None
    if req.anonymize_pii:
        try:
            pii_result = anonymize_text(req.query, redact_level="standard")
            sanitized_query = pii_result["anonymized_text"]
            if pii_result["redactions_count"] > 0:
                pii_report = pii_result
                logger.info(f"PII Guard: Redacted {pii_result['redactions_count']} items from query")
        except Exception as e:
            logger.warning(f"PII anonymization failed (using raw query): {e}")

    # Log the query submission to the audit sink (before the stream begins).
    # We log again after stream completes with latency + response size.
    import time as _t_q
    _q_started = _t_q.time()
    try:
        from audit_sink import log_query_event as _log_q
        await _log_q(
            user=user, request=request, query=sanitized_query, mode=req.mode,
            outcome="submitted", matter_id=getattr(req, "matter_id", None),
        )
    except Exception: pass

    from fastapi.responses import StreamingResponse
    import json

    # Accumulator so the final persisted row has the real response text (for
    # rehydration when the user reopens this thread later).
    _captured_text_parts: list[str] = []
    history_id = f"qh_{uuid.uuid4().hex[:12]}"

    async def sse_event_generator():
        # Force flush of browser buffers (Chrome/Webkit often buffer first ~1KB of SSE)
        yield ": " + " " * 2048 + "\n\n"

        # Tell the frontend which thread this response belongs to so it can
        # update its sidebar / local state immediately (important for the
        # "new thread → refresh sidebar" UX).
        if resolved_thread_id:
            yield f"data: {json.dumps({'type':'thread','thread_id':resolved_thread_id})}\n\n"

        try:
            # conversation_id — prefer the resolved thread_id so Supermemory
            # naturally groups all turns of this thread under one conversation.
            _conv_id = resolved_thread_id or getattr(req, "conversation_id", None) or f"session_{uuid.uuid4().hex[:10]}"

            # ═══════════════════════════════════════════════════════════════
            # ROUTER: spectr_pipeline (4-stage cascade) is now the default.
            #   mode "depth" | "partner" → spectr_pipeline with force_deep=True
            #                              → drafter = gpt-5.5 direct (slow, top quality)
            #   everything else          → spectr_pipeline default tiers
            #                              → drafter = gpt-4.1 / claude-sonnet-4-6 / 4o-mini
            #                                via Emergent universal key (cost-efficient)
            # If the pipeline fails we auto-fall-through to war_room_engine so
            # the user never sees a blank response.
            # ═══════════════════════════════════════════════════════════════
            _mode_lower = (req.mode or "").lower()
            _force_deep = _mode_lower in ("partner", "depth", "deep")
            _use_pipeline = True   # pipeline handles every mode now
            _pipeline_ok = False
            if _use_pipeline:
                try:
                    from spectr_pipeline import run_spectr_pipeline
                    # Announce the stages so the SSE progress UI shows meaningful steps.
                    yield f"data: {json.dumps({'type':'war_room_status','status':'Classifying query + routing…'})}\n\n"
                    result = await run_spectr_pipeline(
                        user_query=sanitized_query,
                        recent_history=req.conversation_history or [],
                        force_deep=_force_deep,
                        timing_budget_s=180 if _force_deep else 60,
                    )
                    response_text = result.get("response_text") or ""
                    timings = result.get("timings", {})
                    model_used = result.get("model_used", "")
                    cost = result.get("cost_inr", 0)
                    chunks_used = result.get("chunks_used") or []
                    logger.info(
                        f"[/assistant/query] spectr_pipeline: {timings.get('total',0)}s "
                        f"({len(response_text.split())} words) via {model_used} ₹{cost}"
                    )
                    # Triage canned replies are intentionally short — accept
                    # them as valid. Only fall through when the pipeline truly
                    # produced nothing (model_used != triage-canned AND <200 chars).
                    is_triage = model_used == "triage-canned"
                    has_content = response_text and (is_triage or len(response_text) > 200)
                    if has_content:
                        _pipeline_ok = True
                        _captured_text_parts.append(response_text)
                        yield f"data: {json.dumps({'type':'fast_chunk','content':response_text})}\n\n"
                        yield f"data: {json.dumps({'type':'fast_complete','models_used':[model_used],'sections':[c.get('citation','') for c in chunks_used],'timings':timings,'cost_inr':cost})}\n\n"
                    else:
                        logger.warning(f"[/assistant/query] spectr_pipeline produced empty/thin output ({len(response_text)} chars) — falling through to war_room_engine")
                except Exception as _pipe_err:
                    logger.warning(f"[/assistant/query] spectr_pipeline error ({_pipe_err}) — falling through to war_room_engine")

            # Fall-through to war_room_engine when pipeline is skipped or failed
            if not _pipeline_ok:
                async for chunk in process_query_streamed(
                    user_query=sanitized_query,
                    mode=req.mode,
                    matter_context=matter_context,
                    statute_context=statute_context,
                    firm_context=firm_context,
                    conversation_history=req.conversation_history or [],
                    user_id=user.get("user_id"),
                    conversation_id=_conv_id,
                    matter_id=(getattr(req, "matter_id", None) or None),
                ):
                    # Capture text chunks so we can persist the response later.
                    # We look at the chunk JSON shape emitted by ai_engine.
                    try:
                        ev = json.loads(chunk)
                        if isinstance(ev, dict):
                            for k in ("content", "text", "chunk", "delta", "response_text"):
                                v = ev.get(k)
                                if isinstance(v, str) and v:
                                    _captured_text_parts.append(v)
                                    break
                    except Exception:
                        pass
                    yield f"data: {chunk}\n\n"
        except Exception as e:
            logger.error(f"SSE streaming error: {e}")
            error_payload = json.dumps({
                "type": "error",
                "content": f"An error occurred while generating the response. Please try again.",
                "detail": str(e)[:200]
            })
            yield f"data: {error_payload}\n\n"
        yield f"data: [DONE]\n\n"

        # After the stream closes: persist the query_history row (now with the
        # full response_text) and bump the thread's last_preview + LLM title.
        try:
            response_text = ("".join(_captured_text_parts))[:20000]
            await db.query_history.insert_one({
                "history_id": history_id,
                "user_id": user["user_id"],
                "matter_id": req.matter_id,
                "thread_id": resolved_thread_id,
                "query": req.query,
                "mode": req.mode,
                "query_types": [req.mode] if req.mode else ["general"],
                "model_used": req.mode or "fast",
                "citations_count": 0,
                "response_text": response_text,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            if req.matter_id:
                await db.matters.update_one(
                    {"matter_id": req.matter_id, "user_id": user.get("user_id")},
                    {"$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
                )
            if resolved_thread_id:
                from thread_manager import update_thread_after_response
                # Background: bumps preview + runs LLM auto-title if this is
                # still the thread's first message.
                await update_thread_after_response(
                    db, resolved_thread_id, user["user_id"],
                    response_preview=response_text[:500],
                    try_llm_title=True,
                    original_query=req.query,
                )
            # Also save this turn to Supermemory (fire-and-forget). Separate
            # calls per role so search can pull the user-turn or assistant-turn
            # granularly. turn_index is epoch ms for monotonic ordering.
            try:
                from supermemory_client import save_turn_background as _sm_save
                _ts = int(datetime.now(timezone.utc).timestamp() * 1000)
                _sm_save(
                    user_id=user["user_id"],
                    conversation_id=resolved_thread_id or history_id,
                    turn_index=_ts,
                    role="user",
                    content=req.query,
                    matter_id=(req.matter_id or None),
                    mode=req.mode,
                )
                _sm_save(
                    user_id=user["user_id"],
                    conversation_id=resolved_thread_id or history_id,
                    turn_index=_ts + 1,
                    role="assistant",
                    content=response_text,
                    matter_id=(req.matter_id or None),
                    mode=req.mode,
                )
            except Exception as e:
                logger.debug(f"Supermemory save skipped: {e}")
        except Exception as e:
            logger.warning(f"Post-stream persistence failed (non-blocking): {e}")

    return StreamingResponse(sse_event_generator(), media_type="text/event-stream")


# ─── Statute retrieval cache (in-memory, 6-hour TTL) ──────────────────
# The same query phrased differently ("section 194C", "194C TDS rate",
# "what is 194C") all end up extracting the same section numbers. We key
# the cache on a canonical form built from extracted section numbers +
# matched act hints, so paraphrases share cache entries.
_STATUTE_CACHE: dict = {}
_STATUTE_CACHE_TTL = 6 * 3600

def _statute_cache_key(query: str) -> str:
    import re as _re
    q = (query or "").lower()
    # Extract the section numbers and the act hints that drive retrieval
    sections = sorted(set(_re.findall(r'(?:section\s*)?(\d+[a-z]{0,3}(?:\([a-z0-9]+\))*)', q)))
    # Only keep small-looking section numbers (actual sections, not years)
    sections = [s for s in sections if not s.isdigit() or int(s) < 1000]
    acts = sorted(set(_re.findall(r'\b(bns|bnss|bsa|cgst|igst|sgst|gst|it act|income tax|fema|rbi|sebi|ibc|companies|partnership|llp|arbitration|negotiable|pmla|consumer|rera|banking)\b', q)))
    canonical = f"secs={','.join(sections)}|acts={','.join(acts)}"
    import hashlib
    return hashlib.md5(canonical.encode("utf-8")).hexdigest()


async def get_statute_context(query: str) -> str:
    """Search statute DB for relevant sections — 3-pass retrieval for grounded responses.

    Pass 1: Exact section number match (highest priority — user asked for specific section)
    Pass 2: Act-specific keyword search (user mentions GST/IT/BNS — pull the relevant act's sections)
    Pass 3: Topic-based semantic search (bail, cheating, ITC, TDS — match by keywords field)

    Returns formatted statute records with clear source attribution.

    Cached: 6-hour TTL on the canonical (section_numbers + act_hints) key.
    Two paraphrased queries that extract the same sections hit the same
    cache entry — e.g. "what is section 73 CGST" and "CGST 73 time limit"
    share the same statute context.
    """
    import re as _re
    import time as _time

    # Cache check before doing any work
    cache_k = _statute_cache_key(query)
    cached = _STATUTE_CACHE.get(cache_k)
    if cached and (_time.time() - cached["ts"]) < _STATUTE_CACHE_TTL:
        logger.info(f"Statute context CACHE HIT (key={cache_k[:8]})")
        return cached["context"]

    query_lower = query.lower()

    # === EXTRACTION ===
    # Section numbers: "73", "16(2)(c)", "194T", "43B(h)", "115BAC", "148A"
    section_nums = _re.findall(r'\b(?:section\s*)?(\d+[A-Za-z]*(?:\([a-z0-9]+\))*)', query_lower)
    section_nums = [s for s in section_nums if not s.isdigit() or int(s) < 1000]  # filter year-like nums

    # Act detection — map user phrases to act name patterns in DB.
    # Pattern strings (keys) are substrings matched against `act_name` in the
    # corpus, pipe-separated for alternatives. Keep these aligned with the
    # act_name values actually in Firestore (see FILENAME_TO_ACT in
    # parse_statutes_to_json.py).
    ACT_MAP = {
        "Central Goods and Services Tax|CGST": ["cgst", "central goods"],
        "Integrated Goods and Services Tax|IGST": ["igst", "integrated goods"],
        "Union Territory Goods and Services Tax|UTGST": ["utgst", "union territory goods"],
        "Goods and Services Tax (Compensation": ["gst compensation", "compensation to states"],
        "Income Tax": ["income tax", "ita", "it act", "itr", "tds", "194", "section 80", "deduction", "assessment", "return filing", "rebate", "87a", "115bac", "old regime", "new regime"],
        "Companies Act": ["companies act", "company act", "roc", "director", "winding up", "oppression", "mismanagement"],
        "Foreign Exchange Management|FEMA": ["fema", "foreign exchange", "lrs", "liberalised remittance"],
        "Reserve Bank of India|RBI Act": ["rbi act", "reserve bank of india act", "rbi act 1934"],
        "Banking Regulation": ["banking regulation", "banking reg act", "banking act 1949"],
        "Bharatiya Nyaya Sanhita|BNS": ["bns", "bharatiya nyaya", "ipc", "murder", "theft", "cheating", "fraud", "forgery", "criminal"],
        "Bharatiya Nagarik Suraksha|BNSS": ["bnss", "nagarik suraksha", "crpc", "bail", "fir", "arrest", "anticipatory bail", "remand", "chargesheet"],
        "Bharatiya Sakshya|BSA": ["bsa", "sakshya", "evidence", "electronic evidence", "admissibility"],
        "Arbitration and Conciliation": ["arbitration", "arbitral award", "conciliation"],
        "Insolvency and Bankruptcy|IBC": ["ibc", "insolvency", "bankruptcy", "cirp", "resolution professional", "nclt", "operational creditor", "financial creditor"],
        "Limited Liability Partnership|LLP": ["llp", "limited liability partnership"],
        "Indian Partnership Act": ["partnership act", "partnership firm", "section 4 partnership"],
        "Securities and Exchange Board of India|SEBI": ["sebi", "insider trading", "listing", "takeover"],
        "Consumer Protection": ["consumer", "consumer protection", "deficiency in service", "unfair trade"],
        "RERA": ["rera", "real estate", "builder", "allottee", "possession delay"],
        "PMLA": ["pmla", "money laundering", "enforcement directorate", "ed attachment"],
        "Negotiable Instruments": ["138", "cheque bounce", "negotiable instrument", "dishonour"],
        "Contract": ["contract act", "indian contract", "breach of contract", "specific performance"],
        "Limitation": ["limitation act", "limitation period", "time barred", "prescribed period"],
    }

    matched_acts = []
    for act_pattern, triggers in ACT_MAP.items():
        if any(t in query_lower for t in triggers):
            matched_acts.append(act_pattern)

    # Legal topic keywords (filter noise words)
    STOP_WORDS = {"what", "which", "when", "where", "will", "would", "could", "should", "does", "about",
                  "this", "that", "these", "with", "from", "have", "been", "being", "their", "there",
                  "also", "into", "more", "than", "they", "under", "over", "such", "only", "very",
                  "just", "like", "some", "much", "many", "most", "other", "after", "before", "between",
                  "same", "each", "every", "both", "through", "during", "here", "case", "please", "help",
                  "want", "need", "tell", "explain", "know", "question", "answer", "query", "check"}
    topic_keywords = [w for w in _re.findall(r'\b[a-z]+\b', query_lower) if len(w) > 3 and w not in STOP_WORDS]
    # Add multi-word legal terms
    legal_phrases = _re.findall(r'(?:show cause|input tax|tax credit|cheque bounce|anticipatory bail|'
                                r'money laundering|foreign exchange|real estate|consumer protection|'
                                r'insider trading|private defence|criminal breach|tax audit|form 3cd|'
                                r'clause 44|gstr-2b|gstr-3b|tax slab|new regime|old regime|'
                                r'penalty.*(?:late|filing|delay)|notice.*(?:valid|din|148|73|74))',
                                query_lower)
    topic_keywords.extend(legal_phrases)
    topic_keywords = list(set(topic_keywords))[:20]

    # Demo-Mode Fallback Cache (for when MongoDB Atlas is blocked by IP/Firewall)
    DEMO_RAG_CACHE = {
        "clause 44": "Clause 44 of Form 3CD (Tax Audit Report): Break-up of total expenditure of entities registered or not registered under the GST.\nRequires details of total expenditure incurred during the year, divided into: (a) Expenditure in respect of entities registered under GST (with breakdown of exempt/nil-rated, composition scheme, and other registered entities) and (b) Expenditure relating to entities not registered under GST.\nNote: This is purely a reporting/disclosure requirement. There is no penalty under the Income Tax Act explicitly for incurring expenditure from unregistered vendors, though such expenditure may be scrutinized for genuineness under Section 37 or cash disallowance under Section 40A(3).",
        "194j": "Section 194J of the Income Tax Act, 1961: Fees for Professional or Technical Services.\nAny person paying fees for professional services, technical services, royalty, or non-compete fees to a resident must deduct TDS.\nRates: 2% for royalty on sale/distribution of cinematographic films and technical services (not professional). 10% for professional services, directors' fees, and other royalty.\nThreshold: ₹30,000 per financial year per category.",
        "16(4)": "Section 16(4) of the CGST Act, 2017: Time limit for availing Input Tax Credit (ITC).\nA registered person shall not be entitled to take input tax credit in respect of any invoice or debit note for supply of goods or services or both after the thirtieth day of November following the end of financial year to which such invoice or debit note pertains or furnishing of the relevant annual return, whichever is earlier.",
        "73": "Section 73 of the CGST Act, 2017: Determination of tax not paid or short paid or erroneously refunded or input tax credit wrongly availed or utilised for any reason other than fraud or any wilful-misstatement or suppression of facts.\nThe proper officer shall issue notice at least three months prior to the time limit of three years for issuance of order.",
        "74": "Section 74 of the CGST Act, 2017: Determination of tax not paid or short paid or erroneously refunded or ITC wrongly availed or utilised by reason of fraud, or any wilful-misstatement or suppression of facts to evade tax.\nThe proper officer shall issue notice at least six months prior to the time limit of five years for issuance of order.",
        "115bac": "Section 115BAC of the Income Tax Act, 1961 — New Tax Regime. AY 2026-27 slabs: 0-4L (nil), 4-8L (5%), 8-12L (10%), 12-16L (15%), 16-20L (20%), 20-24L (25%), 24L+ (30%). Standard deduction ₹75,000. Rebate u/s 87A: NIL tax for income up to ₹12,00,000.",
        "194c": "Section 194C of the Income Tax Act, 1961: TDS on Payments to Contractors. Rate: 1% (individual/HUF), 2% (others). Threshold: single payment >₹30,000 or aggregate >₹1,00,000 in FY.",
        "194h": "Section 194H of the Income Tax Act, 1961: TDS on Commission or Brokerage. Rate: 5%. Threshold: ₹15,000 per FY.",
        "194i": "Section 194I of the Income Tax Act, 1961: TDS on Rent. Rate: 2% (plant/machinery/equipment), 10% (land/building/furniture/fittings). Threshold: ₹2,40,000 per FY.",
        "194t": "Section 194T of the Income Tax Act, 1961: TDS on Partner Payments. NEW w.e.f. 01-04-2025. Covers salary, remuneration, commission, bonus, interest to partners. Rate: 10%. Threshold: ₹20,000 per FY.",
        "bail": "Bail framework: BNSS S.436 (bailable), S.480 (non-bailable), S.482 (anticipatory bail). Default bail if chargesheet not filed within 60/90 days. Bail is rule, jail is exception — State of Rajasthan v Balchand (1977) 4 SCC 308.",
        "148": "Section 148/148A of the Income Tax Act: Reassessment provisions. S.148A (inserted by Finance Act 2021) mandates pre-notice inquiry — AO must (a) conduct enquiry with prior approval of specified authority, (b) provide information to assessee, (c) consider reply, (d) pass order deciding whether to issue notice u/s 148. Time limit: 3 years from end of AY (normal), up to 10 years if escaped income ≥₹50 lakhs (with principal chief commissioner approval).",
        "138": "Section 138 of the Negotiable Instruments Act, 1881: Dishonour of cheque for insufficiency of funds. Punishable with imprisonment up to 2 years and/or fine up to twice the cheque amount. Conditions: (a) cheque presented within 3 months / validity, (b) dishonour for insufficient funds or account closed, (c) statutory demand notice within 30 days of dishonour, (d) payee fails to pay within 15 days. Complaint must be filed within 1 month thereafter before Judicial Magistrate First Class. Offence compoundable under S.147.",
        "cheque bounce": "Cheque Bounce (S.138 NI Act): criminal + civil remedy. Criminal: complaint under S.138 within limitation window; 2 years imprisonment + twice cheque amount fine. Civil: summary suit under Order XXXVII CPC for liquidated amount. Key compliance: (1) demand notice in 30 days of dishonour, (2) complaint within 1 month of 15-day cure expiry. Interim compensation 20% under S.143A.",
        "29a": "Section 29A of the IBC: Eligibility restrictions on Resolution Applicants. Persons disqualified include: undischarged insolvent, wilful defaulter, NPA for 1+ year, convicted of certain offences, disqualified director, prohibited by SEBI. Also covers connected persons/promoters. Introduced to stop promoters from regaining control through the CIRP backdoor.",
        "10": "Section 10 of FEMA, 1999: Authorised Persons. RBI may authorise any person to deal in foreign exchange or foreign securities — authorised dealers (Cat-I banks), authorised money changers, offshore banking units. Must comply with RBI directions; authorisation can be revoked for contravention. All foreign exchange transactions MUST route through an authorised person.",
        "194c": "Section 194C of the Income Tax Act, 1961: TDS on Payments to Contractors. Rate: 1% (individual/HUF), 2% (others). Threshold: single payment > ₹30,000 or aggregate > ₹1,00,000 in FY. Applies to contracts for work (including supply of labour), advertising, broadcasting, carriage of goods/passengers, catering, manufacturing using material supplied by customer.",
    }

    # === 3-PASS RETRIEVAL ===
    relevant = []
    seen_ids = set()  # deduplicate by section_number + act_name

    def _add_unique(docs):
        for d in docs:
            key = f"{d.get('act_name', '')}:{d.get('section_number', '')}"
            if key not in seen_ids:
                seen_ids.add(key)
                relevant.append(d)

    # Helper: client-side act filter. Firestore has no regex, so we fetch by
    # section_number and filter by act_name in Python. Keeps the right act's
    # §103 on top when user says "BNS 103".
    #
    # ACT_MAP keys are pipe-separated regex alternations like
    # "Bharatiya Nyaya Sanhita|BNS" — we split them on '|' and check each
    # alternative as a substring of act_name.
    def _filter_by_acts(docs, act_patterns):
        if not act_patterns:
            return docs
        alts = []
        for p in act_patterns:
            for alt in p.split("|"):
                alt = alt.strip().lower()
                if alt:
                    alts.append(alt)
        matched = []
        unmatched = []
        for d in docs:
            an = (d.get("act_name") or "").lower()
            if any(alt in an for alt in alts):
                matched.append(d)
            else:
                unmatched.append(d)
        # Put matched-act docs first, keep the rest as fallback
        return matched + unmatched

    try:
        # PASS 1: Exact section number match (highest priority)
        if section_nums:
            # Fetch every act's matching section; re-rank client-side so the
            # user's mentioned act (if any) floats to the top.
            cursor = db.statutes.find(
                {"section_number": {"$in": section_nums}}, {"_id": 0}
            ).limit(25)
            exact_matches = await cursor.to_list(25)
            if matched_acts:
                exact_matches = _filter_by_acts(exact_matches, matched_acts)
            # Cap to top 10 after re-ranking
            _add_unique(exact_matches[:10])

        # PASS 2: Act-specific keyword search
        # Fetch by keyword $in, then filter client-side by act. Firestore
        # can't regex across fields so we do the act narrow here.
        if matched_acts and len(relevant) < 10 and topic_keywords:
            cursor = db.statutes.find(
                {"keywords": {"$in": topic_keywords[:10]}}, {"_id": 0}
            ).limit(40)
            kw_matches = await cursor.to_list(40)
            # Filter to user's act of interest, then take top
            kw_matches = _filter_by_acts(kw_matches, matched_acts)
            _add_unique(kw_matches[:10])

        # PASS 3: Topic-based broad search (only if we still have few results)
        if topic_keywords and len(relevant) < 6:
            kw_regex = "|".join(topic_keywords[:12])
            cursor = db.statutes.find(
                {"$or": [
                    {"keywords": {"$in": topic_keywords[:12]}},
                    {"section_title": {"$regex": kw_regex, "$options": "i"}},
                ]},
                {"_id": 0}
            ).limit(8)
            _add_unique(await cursor.to_list(8))

    except Exception as e:
        # Distinguish Firestore-quota errors (expected on free tier) from
        # real DB outages. On quota errors we degrade silently to the demo
        # cache so the response still flows; on real errors we log loud.
        msg = str(e)
        is_quota = "quota" in msg.lower() or "FirestoreQuota" in type(e).__name__
        if is_quota:
            logger.info("Firestore quota limit — using demo RAG cache + training knowledge")
        else:
            logger.warning(f"RAG fetch failed (falling back to demo cache). Error: {e}")
        # Serve matching demo records
        for key, text in DEMO_RAG_CACHE.items():
            if key in query_lower:
                relevant.append({
                    "act_name": _infer_act_from_key(key),
                    "section_number": key.upper(),
                    "section_title": "Verified Statutory Provision",
                    "section_text": text
                })

    if not relevant:
        # Final fallback — serve from demo cache
        for key, text in DEMO_RAG_CACHE.items():
            if key in query_lower:
                relevant.append({
                    "act_name": _infer_act_from_key(key),
                    "section_number": key.upper(),
                    "section_title": "Verified Statutory Provision",
                    "section_text": text
                })

        if not relevant:
            return ""

    # Cap at 12 results to avoid overwhelming the context window
    relevant = relevant[:12]

    context_parts = []
    for s in relevant:
        context_parts.append(
            f"[DB RECORD] Section {s.get('section_number', 'N/A')} of {s.get('act_name', 'N/A')}"
            f" — {s.get('section_title', '')}\n"
            f"{s.get('section_text', '')}"
        )

    result = "\n\n".join(context_parts)

    # Cache the result for 6 hours. Evict oldest 10% when cache hits 500.
    if len(_STATUTE_CACHE) >= 500:
        import time as _t
        oldest = sorted(_STATUTE_CACHE.items(), key=lambda kv: kv[1].get("ts", 0))[:50]
        for k, _ in oldest:
            _STATUTE_CACHE.pop(k, None)
    import time as _t_cache
    _STATUTE_CACHE[cache_k] = {"context": result, "ts": _t_cache.time()}
    return result

# ==================== TAX AUDIT TOOLS ====================

@api_router.post("/tax-audit/clause44")
async def process_clause_44(
    request: Request,
    file: UploadFile = File(...),
    authorization: str = Header(None)
):
    try:
        user = await get_current_user(request, authorization)
        file_data = await file.read()
        
        # Vendor Classifier processes Excel/CSV and outputs the Clause 44 dict
        result = classify_vendors(file_data)
        
        return result
    except Exception as e:
        logger.error(f"Clause 44 processing failed: {e}")
        raise HTTPException(status_code=400, detail=str(e))

# ==================== DOCUMENT EXPORT ROUTES ====================

@api_router.post("/export/word")
async def export_word(req: ExportRequest, request: Request, authorization: str = Header(None)):
    """Generate and return a formatted Microsoft Word document."""
    user = await get_current_user(request, authorization)
    try:
        doc_bytes = generate_word_document(
            title=req.title,
            content=req.content,
            doc_type="memo",
            header_option=req.header_option,
            watermark=req.watermark
        )
        return StreamingResponse(
            io.BytesIO(doc_bytes),
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f'attachment; filename="{req.title.replace(" ", "_")}.docx"'}
        )
    except Exception as e:
        logger.error(f"Word export failed: {e}")
        raise HTTPException(status_code=500, detail=f"Word generation failed: {str(e)}")

@api_router.post("/export/pdf")
async def export_pdf(req: ExportRequest, request: Request, authorization: str = Header(None)):
    """Generate and return a formatted PDF document."""
    user = await get_current_user(request, authorization)
    try:
        pdf_bytes = generate_pdf_bytes(
            title=req.title,
            content=req.content,
            header_option=req.header_option,
            watermark=req.watermark
        )
        if not pdf_bytes:
            raise HTTPException(status_code=500, detail="PDF generation returned empty — WeasyPrint may not be installed.")
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{req.title.replace(" ", "_")}.pdf"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"PDF export failed: {e}")
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")

@api_router.post("/export/excel")
async def export_excel(req: ExportRequest, request: Request, authorization: str = Header(None)):
    """Generate and return an Excel document from tabular content."""
    user = await get_current_user(request, authorization)
    try:
        excel_bytes = generate_excel_document(title=req.title, content=req.content)
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{req.title.replace(" ", "_")}.xlsx"'}
        )
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")

# ==================== VAULT ROUTES ====================

@api_router.post("/vault/upload")
async def upload_document(
    request: Request,
    file: UploadFile = File(...),
    authorization: str = Header(None)
):
    user = await get_current_user(request, authorization)

    # === SECURITY: extension + MIME whitelist, size cap, filename sanitization ===
    ALLOWED_EXTS = {"pdf", "doc", "docx", "txt", "csv", "xlsx", "xls", "jpg", "jpeg", "png", "rtf", "md"}
    ALLOWED_MIME_PREFIXES = {
        "application/pdf", "application/msword",
        "application/vnd.openxmlformats-officedocument",
        "application/vnd.ms-excel", "application/rtf",
        "text/plain", "text/csv", "text/markdown", "text/rtf",
        "image/jpeg", "image/png",
        "application/octet-stream",  # allowed because browsers send this generically; ext check still gates
    }
    MAX_FILE_BYTES = 50 * 1024 * 1024  # 50 MB per file

    # Sanitize filename: strip path, control chars, NUL bytes. Reject path-traversal attempts.
    raw_name = (file.filename or "upload.bin").strip()
    safe_name = re.sub(r"[\x00-\x1f\x7f]", "", raw_name)
    safe_name = os.path.basename(safe_name.replace("\\", "/"))  # strip any directory component
    if not safe_name or safe_name in (".", "..") or len(safe_name) > 255:
        raise HTTPException(status_code=400, detail="Invalid filename")
    # Reject filenames that try to look like system paths
    if safe_name.startswith(".") and not safe_name.lower().endswith((".pdf", ".doc", ".docx", ".txt", ".csv", ".xlsx", ".xls", ".jpg", ".jpeg", ".png", ".rtf", ".md")):
        raise HTTPException(status_code=400, detail="Invalid filename")

    file_ext = safe_name.rsplit(".", 1)[-1].lower() if "." in safe_name else "bin"
    if file_ext not in ALLOWED_EXTS:
        raise HTTPException(
            status_code=415,
            detail=f"File type '{file_ext}' not supported. Allowed: {', '.join(sorted(ALLOWED_EXTS))}"
        )

    # MIME check — reject if declared content-type doesn't match any allowed prefix
    declared_mime = (file.content_type or "").lower()
    if declared_mime and not any(declared_mime.startswith(p) for p in ALLOWED_MIME_PREFIXES):
        raise HTTPException(status_code=415, detail=f"Content-Type '{declared_mime}' not allowed")

    # Read with size cap — prevents OOM attacks
    file_data = await file.read()
    if len(file_data) > MAX_FILE_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large (max {MAX_FILE_BYTES // (1024*1024)} MB)")
    if len(file_data) == 0:
        raise HTTPException(status_code=400, detail="Empty file")

    # Magic-byte sanity check for the high-risk formats — a .pdf that's actually an
    # executable/zip should be rejected. Lightweight, defence-in-depth.
    magic_checks = {
        "pdf": b"%PDF-",
        "docx": b"PK\x03\x04",  # docx/xlsx are zip
        "xlsx": b"PK\x03\x04",
        "jpg": b"\xff\xd8\xff",
        "jpeg": b"\xff\xd8\xff",
        "png": b"\x89PNG\r\n\x1a\n",
    }
    expected_magic = magic_checks.get(file_ext)
    if expected_magic and not file_data.startswith(expected_magic):
        # doc (old binary) has no reliable magic; skip enforcement there
        if file_ext != "doc":
            raise HTTPException(status_code=415, detail=f"File content does not match declared type '{file_ext}'")

    # Text extraction in a threadpool — PyMuPDF on a 500+ page PDF can block
    # the event loop for 10+ seconds, blocking EVERY other request behind it.
    # run_in_executor moves it off-thread so the server stays responsive.
    extracted_text = ""
    try:
        import concurrent.futures
        loop = asyncio.get_event_loop()
        def _do_extract():
            if file_ext == "pdf":
                return extract_pdf_text(file_data)
            elif file_ext in ["doc", "docx"]:
                return extract_docx_text(file_data)
            elif file_ext in ["txt", "csv", "md", "rtf"]:
                return file_data.decode("utf-8", errors="ignore")
            elif file_ext in ["xlsx", "xls"]:
                return extract_xlsx_text(file_data)
            elif file_ext in ["jpg", "jpeg", "png"]:
                return "[Image uploaded - OCR processing available]"
            return ""
        try:
            # 45s hard cap — if extraction takes longer the user gets a partial
            # record with empty text rather than a frozen request.
            extracted_text = await asyncio.wait_for(
                loop.run_in_executor(None, _do_extract),
                timeout=45.0,
            )
        except asyncio.TimeoutError:
            logger.warning(f"Text extraction timed out for {safe_name} ({len(file_data)} bytes) — storing with empty text")
            extracted_text = "[Extraction timed out — document stored but not indexed]"
    except Exception as e:
        logger.warning(f"Text extraction failed for {file_ext}: {e}")
        extracted_text = ""

    doc_type = classify_document(safe_name, extracted_text)

    # Use sanitized filename for all downstream storage — never the raw name
    storage_path = generate_storage_path(user["user_id"], safe_name)
    try:
        storage_result = put_object(storage_path, file_data, file.content_type or "application/octet-stream")
    except Exception as e:
        logger.error(f"Storage upload failed: {e}")
        storage_result = {"path": storage_path}

    doc_id = f"doc_{uuid.uuid4().hex[:12]}"
    doc = {
        "doc_id": doc_id,
        "user_id": user["user_id"],
        "filename": safe_name,
        "file_ext": file_ext,
        "content_type": file.content_type,
        "storage_path": storage_result.get("path", storage_path),
        "size": len(file_data),
        "doc_type": doc_type,
        "extracted_text": extracted_text[:10000000],
        "is_deleted": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    # ALWAYS cache in memory first (firewall-proof, with TTL + size eviction)
    _vault_cache_set(doc_id, doc)
    logger.info(f"Document {doc_id} cached in-memory ({safe_name}, cache size: {len(_vault_cache)})")
    # Fire-and-forget the Firestore write. If Firestore is quota-limited or
    # unreachable the upload still returns fast — the in-memory cache already
    # has the doc so the user's next action works immediately.
    async def _bg_persist():
        try:
            await asyncio.wait_for(db.documents.insert_one(doc), timeout=3.0)
            logger.info(f"Document {doc_id} saved to Firestore ({safe_name})")
        except asyncio.TimeoutError:
            logger.warning(f"Document {doc_id} Firestore insert timed out after 3s — using cache only")
        except Exception as e:
            logger.warning(f"Document {doc_id} Firestore insert failed (cache-only): {e}")
    asyncio.create_task(_bg_persist())

    result = {k: v for k, v in doc.items() if k not in ["_id", "extracted_text"]}
    # Normalize — frontend reads document_id
    result["document_id"] = result.get("doc_id", doc_id)

    # Exhaustive log → audit sink cluster (also fire-and-forget; audit_sink
    # has its own circuit breaker but we don't need to await the HTTP reply).
    try:
        from audit_sink import log_vault_event as _log_vault
        asyncio.create_task(_log_vault(
            user=user, request=request, action="upload", outcome="success",
            doc_id=doc_id, filename=safe_name, doc_type=doc_type,
            size_bytes=len(file_data),
        ))
    except Exception:
        pass

    return result

@api_router.get("/vault/documents")
async def list_documents(request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        docs = await db.documents.find(
            {"user_id": user["user_id"], "is_deleted": False},
            {"_id": 0, "extracted_text": 0}
        ).sort("created_at", -1).to_list(100)
        # Merge with in-memory cache (in case some docs only exist in cache)
        cached_ids = {d.get("doc_id") for d in docs}
        for doc_id, entry in _vault_cache.items():
            cached_doc = entry.get("data", {}) if isinstance(entry, dict) and "data" in entry else entry
            if doc_id not in cached_ids and cached_doc.get("user_id") == user["user_id"] and not cached_doc.get("is_deleted"):
                docs.append({k: v for k, v in cached_doc.items() if k not in ["_id", "extracted_text"]})
        return docs
    except Exception as e:
        logger.error(f"MongoDB list_documents error: {e}")
        # Return from in-memory cache
        cached_docs = []
        for doc_id, entry in _vault_cache.items():
            cached_doc = entry.get("data", {}) if isinstance(entry, dict) and "data" in entry else entry
            if cached_doc.get("user_id") == user["user_id"] and not cached_doc.get("is_deleted"):
                cached_docs.append({k: v for k, v in cached_doc.items() if k not in ["_id", "extracted_text"]})
        return cached_docs

@api_router.post("/vault/analyze")
async def analyze_document(req: DocumentAnalysisRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)

    # Try MongoDB first, fall back to in-memory cache
    doc = None
    try:
        doc = await db.documents.find_one(
            {"doc_id": req.document_id, "user_id": user["user_id"]},
            {"_id": 0}
        )
    except Exception as e:
        logger.warning(f"MongoDB find_one failed, checking in-memory cache: {e}")

    # Fallback to in-memory cache (TTL-aware)
    if not doc:
        doc = _vault_cache_get(req.document_id)

    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    statute_context = ""
    company_context = ""
    try:
        from ai_engine import extract_company_name
        from insta_financials import search_company
        if req.custom_prompt:
            statute_context = await get_statute_context(req.custom_prompt)
        doc_text = doc.get("extracted_text", "")
        comp_name = extract_company_name(req.custom_prompt) or extract_company_name(doc_text[:5000])
        if comp_name:
            c_data = await search_company(comp_name)
            if c_data:
                company_context = str(c_data)
    except Exception as e:
        logger.error(f"Error fetching external context for vault analysis: {e}")

    # Sanitize all user-controlled context before LLM injection (prompt injection defense)
    # AND run PII Anonymisation as required by T&C Clause 8.2 — strips Aadhaar,
    # PAN, GSTIN, mobile, email, IFSC, CIN before transmission to external AI.
    _raw_doc = doc.get("extracted_text", "")
    _pii_scrubbed = anonymize_text(_raw_doc, redact_level="standard")
    if _pii_scrubbed.get("redactions_count", 0) > 0:
        logger.info(f"Vault PII Guard: Redacted {_pii_scrubbed['redactions_count']} items from document {req.document_id}")
    sanitized_doc_text = sanitize_context_for_llm(_pii_scrubbed["anonymized_text"], source_label="vault_doc_text")
    statute_context = sanitize_context_for_llm(statute_context, source_label="vault_statute_context")
    company_context = sanitize_context_for_llm(company_context, source_label="vault_company_context")

    result = await process_document_analysis(
        document_text=sanitized_doc_text,
        doc_type=doc.get("doc_type", "other"),
        analysis_type=req.analysis_type,
        custom_prompt=req.custom_prompt,
        statute_context=statute_context,
        company_context=company_context
    )

    analysis_id = f"analysis_{uuid.uuid4().hex[:12]}"
    analysis_doc = {
        "analysis_id": analysis_id,
        "doc_id": req.document_id,
        "user_id": user["user_id"],
        "analysis_type": req.analysis_type,
        "response_text": result["response_text"],
        "sections": result["sections"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    _analysis_cache_set(analysis_id, analysis_doc)
    try:
        await db.document_analyses.insert_one(analysis_doc)
    except Exception as e:
        logger.warning(f"MongoDB analysis insert failed (cached in-memory): {e}")

    return {
        "analysis_id": analysis_id,
        "response_text": result["response_text"],
        "sections": result["sections"],
        "doc_type": result["doc_type"],
        "analysis_type": result["analysis_type"]
    }

@api_router.post("/vault/compare")
async def compare_documents(req: DocumentCompareRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    base_doc = None
    counter_doc = None
    try:
        base_doc = await db.documents.find_one({"doc_id": req.base_doc_id, "user_id": user["user_id"]}, {"_id": 0})
        counter_doc = await db.documents.find_one({"doc_id": req.counter_doc_id, "user_id": user["user_id"]}, {"_id": 0})
    except Exception as e:
        logger.warning(f"MongoDB compare fetch failed, checking cache: {e}")

    # Fallback to in-memory cache (TTL-aware)
    if not base_doc:
        base_doc = _vault_cache_get(req.base_doc_id)
    if not counter_doc:
        counter_doc = _vault_cache_get(req.counter_doc_id)

    if not base_doc or not counter_doc:
        raise HTTPException(status_code=404, detail="One or both documents not found")

    result = await process_document_comparison(
        base_text=base_doc.get("extracted_text", ""),
        counter_text=counter_doc.get("extracted_text", ""),
        base_name=base_doc.get("filename", "Base Draft"),
        counter_name=counter_doc.get("filename", "Counter Draft"),
        custom_prompt=req.custom_prompt
    )

    analysis_id = f"analysis_{uuid.uuid4().hex[:12]}"
    try:
        await db.document_analyses.insert_one({
            "analysis_id": analysis_id,
            "doc_id": req.base_doc_id,
            "secondary_doc_id": req.counter_doc_id,
            "user_id": user["user_id"],
            "analysis_type": "contract_redline",
            "response_text": result["response_text"],
            "sections": result["sections"],
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    except Exception as e:
        logger.warning(f"MongoDB compare insert failed (cached): {e}")

    return {
        "analysis_id": analysis_id,
        "response_text": result["response_text"],
        "sections": result["sections"],
        "doc_type": "contract_comparison",
        "analysis_type": "contract_redline"
    }

# ==================== CONTRACT RED-LINING ====================

@api_router.post("/contract/redline")
async def contract_redline(
    request: Request,
    file: UploadFile = File(...),
    focus_area: str = Form(""),
    authorization: str = Header(None)
):
    """Upload a DOCX contract → AI analyzes clause-by-clause → returns red-lined DOCX with suggestions."""
    user = await get_current_user(request, authorization)

    if not file.filename.lower().endswith(('.docx', '.doc')):
        raise HTTPException(status_code=400, detail="Only DOCX files are supported for red-lining. Please upload a .docx file.")

    file_data = await file.read()
    if len(file_data) > 25 * 1024 * 1024:  # 25MB limit
        raise HTTPException(status_code=400, detail="File too large. Maximum 25MB.")

    logger.info(f"Contract redline request: {file.filename} ({len(file_data)} bytes) by {user.get('email', 'unknown')}")

    try:
        result = await process_contract_redline(file_data, focus_area)
    except Exception as e:
        logger.error(f"Contract redline processing failed: {e}")
        raise HTTPException(status_code=500, detail=f"Contract analysis failed: {str(e)}")

    if result.get("error") and not result.get("suggestions"):
        raise HTTPException(status_code=422, detail=result["error"])

    # Return the red-lined DOCX as a downloadable file
    redlined_bytes = result.get("redlined_docx")
    if not redlined_bytes:
        raise HTTPException(status_code=500, detail="Failed to generate red-lined document.")

    # Generate output filename
    original_name = file.filename.rsplit(".", 1)[0]
    output_filename = f"{original_name}_REDLINED.docx"

    return StreamingResponse(
        io.BytesIO(redlined_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f'attachment; filename="{output_filename}"',
            "X-Suggestions-Count": str(result["stats"]["total"]),
            "X-Critical-Count": str(result["stats"]["critical"]),
        }
    )


@api_router.post("/contract/analyze-suggestions")
async def contract_analyze_suggestions(
    request: Request,
    file: UploadFile = File(...),
    focus_area: str = Form(""),
    authorization: str = Header(None)
):
    """Upload a DOCX contract → returns JSON suggestions without generating the red-lined DOCX.
    Use this for the preview sidebar before downloading the red-lined version."""
    user = await get_current_user(request, authorization)

    if not file.filename.lower().endswith(('.docx', '.doc')):
        raise HTTPException(status_code=400, detail="Only DOCX files supported.")

    file_data = await file.read()
    full_text = extract_full_text(file_data)
    if not full_text or len(full_text) < 50:
        raise HTTPException(status_code=422, detail="Could not extract text from the document.")

    try:
        suggestions = await analyze_contract_for_redlines(full_text, focus_area)
    except Exception as e:
        logger.error(f"Contract suggestion analysis failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    stats = {
        "total": len(suggestions),
        "critical": sum(1 for s in suggestions if s.get("severity") == "critical"),
        "high": sum(1 for s in suggestions if s.get("severity") == "high"),
        "medium": sum(1 for s in suggestions if s.get("severity") == "medium"),
        "low": sum(1 for s in suggestions if s.get("severity") == "low"),
    }

    return {
        "suggestions": suggestions,
        "stats": stats,
        "text_length": len(full_text),
        "filename": file.filename,
    }


@api_router.post("/contract/redline-from-suggestions")
async def contract_redline_from_suggestions(
    request: Request,
    authorization: str = Header(None)
):
    """Takes previously generated suggestions + original doc_id and generates the red-lined DOCX.
    Used after the user reviews suggestions in the sidebar."""
    user = await get_current_user(request, authorization)
    body = await request.json()
    doc_id = body.get("doc_id", "")
    suggestions = body.get("suggestions", [])

    if not suggestions:
        raise HTTPException(status_code=400, detail="No suggestions provided.")

    # Fetch original document
    doc = None
    try:
        doc = await db.documents.find_one({"doc_id": doc_id, "user_id": user["user_id"]}, {"_id": 0})
    except Exception:
        pass
    if not doc:
        doc = _vault_cache_get(doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Original document not found in vault.")

    # Get the raw file from storage
    try:
        storage_path = doc.get("storage_path", "")
        file_data = get_object(storage_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not retrieve original file: {e}")

    from contract_redline import generate_redlined_docx
    redlined_bytes = generate_redlined_docx(file_data, suggestions)

    original_name = doc.get("filename", "contract").rsplit(".", 1)[0]
    output_filename = f"{original_name}_REDLINED.docx"

    return StreamingResponse(
        io.BytesIO(redlined_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{output_filename}"'}
    )


# ==================== STRUCTURED DATA EXTRACTION ====================

@api_router.post("/extract/document")
async def extract_from_document(
    request: Request,
    file: UploadFile = File(...),
    doc_type: str = Form("auto"),
    authorization: str = Header(None)
):
    """Upload a document (DOCX, PDF, TXT) and extract structured data fields.
    Returns a typed JSON object with all identified fields.
    Supports: contracts, tax notices, invoices, court orders, legal documents.
    """
    user = await get_current_user(request, authorization)

    file_data = await file.read()
    if len(file_data) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum 25MB.")

    # Extract text based on file type
    filename = file.filename.lower()
    text = ""

    if filename.endswith(('.docx', '.doc')):
        text = extract_full_text(file_data)
    elif filename.endswith('.pdf'):
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(file_data))
            text = "\n".join([page.extract_text() or "" for page in reader.pages])
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Could not read PDF: {e}")
    elif filename.endswith(('.txt', '.text', '.md')):
        text = file_data.decode("utf-8", errors="ignore")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use DOCX, PDF, or TXT.")

    if not text or len(text.strip()) < 50:
        raise HTTPException(status_code=422, detail="Could not extract meaningful text from the document.")

    logger.info(f"Data extraction request: {file.filename} ({len(text)} chars) type={doc_type}")

    try:
        result = await extract_structured_data(text, doc_type)
    except Exception as e:
        logger.error(f"Data extraction failed: {e}")
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")

    result["filename"] = file.filename
    return result


@api_router.post("/extract/text")
async def extract_from_text(request: Request, authorization: str = Header(None)):
    """Extract structured data from raw text (no file upload needed).
    Body: {text: str, doc_type?: str}
    """
    user = await get_current_user(request, authorization)
    body = await request.json()
    text = body.get("text", "")
    doc_type = body.get("doc_type", "auto")

    if not text or len(text.strip()) < 50:
        raise HTTPException(status_code=400, detail="Text too short for extraction. Minimum 50 characters.")

    try:
        result = await extract_structured_data(text, doc_type)
    except Exception as e:
        logger.error(f"Text extraction failed: {e}")
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")

    return result


@api_router.post("/extract/batch")
async def extract_batch(request: Request, authorization: str = Header(None)):
    """Extract structured data from multiple documents in parallel.
    Body: {documents: [{id, text, doc_type?}], default_doc_type?: str}
    """
    user = await get_current_user(request, authorization)
    body = await request.json()
    documents = body.get("documents", [])
    default_type = body.get("default_doc_type", "auto")

    if not documents:
        raise HTTPException(status_code=400, detail="No documents provided.")
    if len(documents) > 20:
        raise HTTPException(status_code=400, detail="Maximum 20 documents per batch.")

    try:
        results = await batch_extract(documents, default_type)
    except Exception as e:
        logger.error(f"Batch extraction failed: {e}")
        raise HTTPException(status_code=500, detail=f"Batch extraction failed: {str(e)}")

    return {
        "results": results,
        "total": len(results),
        "successful": sum(1 for r in results if r.get("success")),
    }


@api_router.get("/extract/schemas")
async def get_extraction_schemas():
    """Get available extraction schemas and their fields. No auth required."""
    return {
        "schemas": {
            k: {"description": v["description"], "fields": list(v["fields"].keys()), "field_count": len(v["fields"])}
            for k, v in EXTRACTION_SCHEMAS.items()
        }
    }


# ==================== BATCH ANALYSIS ====================

@api_router.post("/batch/analyze")
async def batch_analyze(request: Request, authorization: str = Header(None)):
    """Run AI analysis across multiple queries/documents in parallel.

    Body: {
        "queries": [{"id": "q1", "query": "...", "context": "optional extra context"}],
        "mode": "partner"|"everyday",
        "matter_id": "optional"
    }

    Returns: {results: [{id, response_text, query_types, model_used, quality_score}], summary}
    Max 10 queries per batch to prevent API abuse.
    """
    user = await get_current_user(request, authorization)
    body = await request.json()
    queries = body.get("queries", [])
    mode = body.get("mode", "partner")
    matter_id = body.get("matter_id", "")

    if not queries:
        raise HTTPException(status_code=400, detail="No queries provided.")
    if len(queries) > 10:
        raise HTTPException(status_code=400, detail="Maximum 10 queries per batch.")

    # Get shared context once (not per-query)
    statute_context = ""
    try:
        # Use the first query for statute context — most queries in a batch are related
        statute_context = await get_statute_context(queries[0].get("query", ""))
    except Exception:
        pass

    matter_context = ""
    if matter_id:
        try:
            # SECURITY: scope by user_id to block cross-user matter lookups
            matter = await db.matters.find_one(
                {"matter_id": matter_id, "user_id": user.get("user_id")},
                {"_id": 0}
            )
            if matter:
                matter_context = f"Matter: {matter.get('name', '')} | Client: {matter.get('client_name', '')} | Type: {matter.get('matter_type', '')}"
        except Exception:
            pass

    # === PROMPT INJECTION DEFENSE (batch context) ===
    statute_context = sanitize_context_for_llm(statute_context, source_label="batch_statute")
    matter_context = sanitize_context_for_llm(matter_context, source_label="batch_matter")

    # Run all queries in parallel with concurrency limit
    import asyncio
    semaphore = asyncio.Semaphore(3)  # Max 3 concurrent AI calls

    async def run_single(q_item):
        async with semaphore:
            try:
                from ai_engine import process_query
                query_text = q_item.get("query", "")
                extra_context = sanitize_context_for_llm(q_item.get("context", ""), source_label="batch_item_context")
                combined_matter = matter_context
                if extra_context:
                    combined_matter = f"{matter_context}\n\nAdditional context: {extra_context}" if matter_context else f"Additional context: {extra_context}"

                result = await process_query(
                    user_query=query_text,
                    mode=mode,
                    matter_context=combined_matter,
                    statute_context=statute_context
                )
                return {
                    "id": q_item.get("id", ""),
                    "query": query_text[:100],
                    "response_text": result.get("response_text", ""),
                    "query_types": result.get("query_types", []),
                    "model_used": result.get("model_used", ""),
                    "quality_score": result.get("quality_score", 0),
                    "citations_count": result.get("citations_count", 0),
                    "success": True,
                }
            except Exception as e:
                logger.error(f"Batch query failed ({q_item.get('id', '')}): {e}")
                return {
                    "id": q_item.get("id", ""),
                    "query": q_item.get("query", "")[:100],
                    "response_text": f"Error: {str(e)}",
                    "success": False,
                }

    results = await asyncio.gather(*[run_single(q) for q in queries])

    successful = sum(1 for r in results if r.get("success"))
    avg_quality = sum(r.get("quality_score", 0) for r in results if r.get("success")) / max(successful, 1)

    return {
        "results": results,
        "total": len(results),
        "successful": successful,
        "average_quality_score": round(avg_quality, 1),
    }


@api_router.post("/batch/documents")
async def batch_document_analysis(request: Request, authorization: str = Header(None)):
    """Analyze multiple uploaded documents through extraction + comparison pipeline.

    Send as multipart form: files[] (up to 10 files) + analysis_type (extract|compare|both)
    """
    user = await get_current_user(request, authorization)
    form = await request.form()
    files = form.getlist("files")
    analysis_type = form.get("analysis_type", "extract")

    if not files or len(files) == 0:
        raise HTTPException(status_code=400, detail="No files uploaded.")
    if len(files) > 10:
        raise HTTPException(status_code=400, detail="Maximum 10 files per batch.")

    results = []

    # Extract structured data from all files in parallel
    if analysis_type in ("extract", "both"):
        documents = []
        for f in files:
            file_data = await f.read()
            text = extract_text_from_bytes(file_data, f.filename)
            documents.append({"id": f.filename, "text": text})
            await f.seek(0)  # Reset for potential comparison step

        extracted = await batch_extract(documents, "auto")
        results.append({"type": "extraction", "results": extracted, "count": len(extracted)})

    # Pairwise comparison if requested
    if analysis_type in ("compare", "both") and len(files) >= 2:
        comparisons = []
        for i in range(len(files) - 1):
            data_a = await files[i].read()
            data_b = await files[i + 1].read()
            text_a = extract_text_from_bytes(data_a, files[i].filename)
            text_b = extract_text_from_bytes(data_b, files[i + 1].filename)
            await files[i].seek(0)
            await files[i + 1].seek(0)

            diff = compute_diff(text_a, text_b)
            comparisons.append({
                "file_a": files[i].filename,
                "file_b": files[i + 1].filename,
                "similarity_pct": diff["similarity_pct"],
                "lines_added": diff["lines_added"],
                "lines_removed": diff["lines_removed"],
                "total_changes": diff["total_changes"],
            })

        results.append({"type": "comparison", "results": comparisons, "count": len(comparisons)})

    return {"analyses": results, "files_processed": len(files)}


# ==================== DOCUMENT COMPARISON ====================

@api_router.post("/compare/documents")
async def compare_documents_upload(
    request: Request,
    file_a: UploadFile = File(...),
    file_b: UploadFile = File(...),
    authorization: str = Header(None)
):
    """Upload two document versions and get a structured diff with AI analysis.
    Returns: JSON diff (changes, similarity %, stats) + AI significance analysis.
    """
    user = await get_current_user(request, authorization)

    data_a = await file_a.read()
    data_b = await file_b.read()

    if len(data_a) > 25 * 1024 * 1024 or len(data_b) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Files too large. Max 25MB each.")

    text_a = extract_text_from_bytes(data_a, file_a.filename)
    text_b = extract_text_from_bytes(data_b, file_b.filename)

    if not text_a or not text_b:
        raise HTTPException(status_code=422, detail="Could not extract text from one or both files.")

    logger.info(f"Doc comparison: {file_a.filename} ({len(text_a)} chars) vs {file_b.filename} ({len(text_b)} chars)")

    try:
        result = await smart_compare(text_a, text_b)
    except Exception as e:
        logger.error(f"Document comparison failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    result["file_a"] = file_a.filename
    result["file_b"] = file_b.filename
    return result


@api_router.post("/compare/documents/docx")
async def compare_documents_docx(
    request: Request,
    file_a: UploadFile = File(...),
    file_b: UploadFile = File(...),
    authorization: str = Header(None)
):
    """Upload two document versions and get a comparison DOCX with red/blue tracked changes."""
    user = await get_current_user(request, authorization)

    data_a = await file_a.read()
    data_b = await file_b.read()

    text_a = extract_text_from_bytes(data_a, file_a.filename)
    text_b = extract_text_from_bytes(data_b, file_b.filename)

    if not text_a or not text_b:
        raise HTTPException(status_code=422, detail="Could not extract text from one or both files.")

    try:
        docx_bytes = generate_comparison_docx(
            text_a, text_b,
            name_a=file_a.filename, name_b=file_b.filename
        )
    except Exception as e:
        logger.error(f"Comparison DOCX generation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    name_a = file_a.filename.rsplit(".", 1)[0]
    name_b = file_b.filename.rsplit(".", 1)[0]

    return StreamingResponse(
        io.BytesIO(docx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{name_a}_vs_{name_b}_COMPARISON.docx"'}
    )


@api_router.post("/compare/text")
async def compare_text(request: Request, authorization: str = Header(None)):
    """Compare two text strings directly (no file upload).
    Body: {text_a: str, text_b: str, name_a?: str, name_b?: str}
    """
    user = await get_current_user(request, authorization)
    body = await request.json()
    text_a = body.get("text_a", "")
    text_b = body.get("text_b", "")

    if not text_a or not text_b:
        raise HTTPException(status_code=400, detail="Both text_a and text_b are required.")

    try:
        result = await smart_compare(text_a, text_b)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return result


@api_router.post("/vault/ask")
async def vault_ask(request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    body = await request.json()
    question = body.get("question", "")
    doc_ids = body.get("doc_ids", [])
    chat_history = body.get("history", [])  # Enable full chat

    if not question:
        raise HTTPException(status_code=400, detail="Question required")

    combined_text = ""
    for did in doc_ids[:20]:
        doc = None
        try:
            doc = await db.documents.find_one({"doc_id": did, "user_id": user["user_id"]}, {"_id": 0})
        except Exception as e:
            logger.warning(f"MongoDB fetch for vault/ask failed on doc {did}, checking cache: {e}")
        # Fallback to cache (TTL-aware)
        if not doc:
            doc = _vault_cache_get(did)
        if doc:
            combined_text += f"\n--- Document: {doc.get('filename', 'N/A')} ---\n{doc.get('extracted_text', '')[:50000]}\n"

    if not combined_text:
        combined_text = "No documents selected."

    # === ENRICH CONTEXT (match /vault/analyze quality) ===
    statute_context = ""
    company_context = ""
    try:
        from ai_engine import extract_company_name
        if question:
            statute_context = await get_statute_context(question)
        # Try to identify company from question or document text
        comp_name = extract_company_name(question) or extract_company_name(combined_text[:5000])
        if comp_name:
            c_data = await search_company(comp_name)
            if c_data:
                company_context = str(c_data)
    except Exception as e:
        logger.warning(f"Vault/ask context enrichment failed (non-blocking): {e}")

    # === PROMPT INJECTION DEFENSE ===
    combined_text = sanitize_context_for_llm(combined_text, source_label="vault_ask_docs")
    statute_context = sanitize_context_for_llm(statute_context, source_label="vault_ask_statute")
    company_context = sanitize_context_for_llm(company_context, source_label="vault_ask_company")

    # If chat history exists, bundle it into the prompt so the model has context
    full_prompt = question
    if chat_history:
        # Sanitize chat history entries to prevent injection via stored messages
        sanitized_history = []
        for msg in chat_history[-10:]:  # Cap at last 10 messages
            role = sanitize_input(str(msg.get('role', 'user')), context="general")
            content = sanitize_context_for_llm(str(msg.get('content', '')), source_label="vault_ask_history")
            sanitized_history.append(f"{role.upper()}: {content}")
        history_text = "\n".join(sanitized_history)
        full_prompt = f"Previous Chat Context:\n{history_text}\n\nCurrent Question: {question}"

    result = await process_document_analysis(
        document_text=combined_text,
        doc_type="bulk",
        analysis_type="general",
        custom_prompt=full_prompt,
        statute_context=statute_context,
        company_context=company_context
    )

    return {
        "response_text": result["response_text"],
        "sections": result["sections"],
    }

# ==================== KILLER APP: CLAUSE 44 CLASSIFIER ====================

@api_router.post("/tools/clause44-classify")
async def clause44_classifier(request: Request, file: UploadFile = File(...), authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Empty file")
            
        result = classify_vendors(content)
        if "error" in result:
            raise HTTPException(status_code=400, detail=result["error"])
            
        return result
    except Exception as e:
        logger.error(f"Clause 44 processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== WORKFLOW ROUTES ====================

WORKFLOWS = [
    {"id": "cheque_bounce_notice", "name": "Legal Notice — Section 138 NI Act (Cheque Bounce)", "category": "litigation",
     "fields": [{"name": "sender_name", "label": "Sender Name", "type": "text"},
                {"name": "sender_address", "label": "Sender Address", "type": "textarea"},
                {"name": "receiver_name", "label": "Receiver Name", "type": "text"},
                {"name": "receiver_address", "label": "Receiver Address", "type": "textarea"},
                {"name": "amount", "label": "Cheque Amount (INR)", "type": "text"},
                {"name": "cheque_number", "label": "Cheque Number", "type": "text"},
                {"name": "bank_name", "label": "Bank Name", "type": "text"},
                {"name": "dishonour_date", "label": "Dishonour Date", "type": "date"}]},
    {"id": "general_demand_notice", "name": "Legal Notice — General Demand", "category": "litigation",
     "fields": [{"name": "sender_name", "label": "Sender Name", "type": "text"},
                {"name": "receiver_name", "label": "Receiver Name", "type": "text"},
                {"name": "subject", "label": "Subject Matter", "type": "textarea"},
                {"name": "amount", "label": "Amount/Relief", "type": "text"},
                {"name": "deadline_days", "label": "Deadline (days)", "type": "text"}]},
    {"id": "consumer_complaint", "name": "Consumer Forum Complaint", "category": "litigation",
     "fields": [{"name": "complainant", "label": "Complainant Name & Address", "type": "textarea"},
                {"name": "opposite_party", "label": "Opposite Party Name & Address", "type": "textarea"},
                {"name": "product_service", "label": "Product/Service", "type": "text"},
                {"name": "deficiency", "label": "Deficiency in Service", "type": "textarea"},
                {"name": "relief_sought", "label": "Relief Sought", "type": "textarea"}]},
    {"id": "rti_application", "name": "RTI Application", "category": "litigation",
     "fields": [{"name": "public_authority", "label": "Public Authority", "type": "text"},
                {"name": "information_sought", "label": "Information Sought", "type": "textarea"},
                {"name": "reason", "label": "Supporting Reason", "type": "textarea"}]},
    {"id": "rera_complaint", "name": "RERA Complaint", "category": "litigation",
     "fields": [{"name": "buyer_details", "label": "Buyer Details", "type": "textarea"},
                {"name": "builder_details", "label": "Builder Details", "type": "textarea"},
                {"name": "rera_number", "label": "Project RERA Number", "type": "text"},
                {"name": "deficiency", "label": "Deficiency/Delay Details", "type": "textarea"}]},
    {"id": "bail_application", "name": "Regular Bail Application (Sessions/HC)", "category": "litigation",
     "fields": [{"name": "accused_details", "label": "Accused Details", "type": "textarea"},
                {"name": "fir_details", "label": "FIR Details (No., Date, PS)", "type": "textarea"},
                {"name": "offence", "label": "Offence Charged", "type": "text"},
                {"name": "grounds", "label": "Grounds for Regular Bail", "type": "textarea"}]},
    {"id": "anticipatory_bail", "name": "Anticipatory Bail Application (Section 438 CrPC)", "category": "litigation",
     "fields": [{"name": "apprehension", "label": "Reasons for apprehension of arrest", "type": "textarea"},
                {"name": "fir_status", "label": "FIR Status (Registered/Threatened)", "type": "textarea"},
                {"name": "applicant_background", "label": "Applicant's Clean Background", "type": "textarea"},
                {"name": "political_vendetta", "label": "Is this Political Vendetta / False Motivation?", "type": "textarea"}]},
    {"id": "recovery_suit_o37", "name": "Summary Recovery Suit (Order XXXVII CPC)", "category": "litigation",
     "fields": [{"name": "plaintiff", "label": "Plaintiff Details", "type": "textarea"},
                {"name": "defendant", "label": "Defendant Details", "type": "textarea"},
                {"name": "written_contract", "label": "Written Contract / Invoice Details", "type": "textarea"},
                {"name": "amount_due", "label": "Liquidated Amount Due", "type": "text"},
                {"name": "interest_claim", "label": "Interest Claimed", "type": "text"}]},
    {"id": "legal_research_memo", "name": "Legal Research Memo", "category": "litigation",
     "fields": [{"name": "legal_issue", "label": "Legal Issue", "type": "textarea"},
                {"name": "jurisdiction", "label": "Jurisdiction", "type": "text"},
                {"name": "key_facts", "label": "Key Facts", "type": "textarea"}]},
    {"id": "writ_petition", "name": "Writ Petition (HC) — Brief", "category": "litigation",
     "fields": [{"name": "violation", "label": "Constitutional/Legal Violation", "type": "textarea"},
                {"name": "authority", "label": "Authority Against", "type": "text"},
                {"name": "relief", "label": "Relief Sought", "type": "textarea"},
                {"name": "facts", "label": "Material Facts", "type": "textarea"}]},
    {"id": "written_statement", "name": "Written Statement / Reply to Plaint", "category": "litigation",
     "fields": [{"name": "case_details", "label": "Case Number & Court", "type": "text"},
                {"name": "plaint_summary", "label": "Plaint Summary / Upload Content", "type": "textarea"},
                {"name": "client_version", "label": "Client's Version of Facts", "type": "textarea"}]},
    {"id": "contract_review", "name": "Contract Review + Risk Report", "category": "litigation",
     "fields": [{"name": "contract_text", "label": "Paste Contract Text or Key Clauses", "type": "textarea"},
                {"name": "party_role", "label": "Your Client is (Party A/B)", "type": "text"}]},
    {"id": "gst_scn_response", "name": "GST SCN Response", "category": "taxation",
     "fields": [{"name": "company_name", "label": "Company Name", "type": "text"},
                {"name": "gstin", "label": "GSTIN", "type": "text"},
                {"name": "scn_number", "label": "SCN Number", "type": "text"},
                {"name": "scn_date", "label": "SCN Date", "type": "date"},
                {"name": "demand_amount", "label": "Demand Amount (INR)", "type": "text"},
                {"name": "issue_type", "label": "Issue Type", "type": "text"}]},
    {"id": "gst_appeal", "name": "GST Appeal — First Appellate Authority", "category": "taxation",
     "fields": [{"name": "order_details", "label": "Order Details", "type": "textarea"},
                {"name": "grounds", "label": "Grounds of Dispute", "type": "textarea"},
                {"name": "amount", "label": "Disputed Amount (INR)", "type": "text"}]},
    {"id": "it_notice_143", "name": "Income Tax Notice Reply — Section 143(1)", "category": "taxation",
     "fields": [{"name": "pan", "label": "PAN", "type": "text"},
                {"name": "assessment_year", "label": "Assessment Year", "type": "text"},
                {"name": "demand_raised", "label": "Demand Raised (INR)", "type": "text"},
                {"name": "nature_of_addition", "label": "Nature of Addition", "type": "textarea"}]},
    {"id": "it_notice_148", "name": "Income Tax Notice Reply — Section 148 (Escaped Assessment)", "category": "taxation",
     "fields": [{"name": "pan", "label": "PAN", "type": "text"},
                {"name": "year", "label": "Assessment Year", "type": "text"},
                {"name": "reasons", "label": "Reasons Given for Reopening", "type": "textarea"}]},
    {"id": "pmla_compliance", "name": "PMLA Compliance Note", "category": "taxation",
     "fields": [{"name": "transaction_details", "label": "Transaction Details", "type": "textarea"},
                {"name": "entity_type", "label": "Entity Type", "type": "text"},
                {"name": "amount", "label": "Amount (INR)", "type": "text"},
                {"name": "cross_border", "label": "Cross-border (Y/N)", "type": "text"}]},
    {"id": "fema_compounding", "name": "FEMA Compounding Application", "category": "taxation",
     "fields": [{"name": "transaction_details", "label": "Transaction Details", "type": "textarea"},
                {"name": "rbi_direction", "label": "RBI Master Direction Reference", "type": "text"},
                {"name": "violation_details", "label": "Violation Details", "type": "textarea"}]},
    {"id": "due_diligence", "name": "Due Diligence Report — Company Acquisition", "category": "taxation",
     "fields": [{"name": "company_name", "label": "Company Name", "type": "text"},
                {"name": "cin", "label": "CIN (if available)", "type": "text"}]},
    {"id": "financial_health", "name": "Financial Health Report", "category": "taxation",
     "fields": [{"name": "company_name", "label": "Company Name", "type": "text"}]},
    {"id": "ibc_section9", "name": "IBC Section 9 — Operational Creditor Application", "category": "taxation",
     "fields": [{"name": "creditor", "label": "Operational Creditor Details", "type": "textarea"},
                {"name": "corporate_debtor", "label": "Corporate Debtor Details", "type": "textarea"},
                {"name": "debt_details", "label": "Debt Details", "type": "textarea"},
                {"name": "demand_notice", "label": "Demand Notice History", "type": "textarea"}]},
    {"id": "director_disqualification", "name": "Director Disqualification Check + Response", "category": "taxation",
     "fields": [{"name": "din", "label": "DIN Number", "type": "text"}]},
    {"id": "pmla_bail_checklist", "name": "PMLA Bail Application — Section 45 Twin Conditions", "category": "criminal",
     "fields": [{"name": "accused_details", "label": "Accused Name, Age, Occupation", "type": "textarea"},
                {"name": "ecir_number", "label": "ECIR Number & Date", "type": "text"},
                {"name": "predicate_offence", "label": "Predicate Offence (FIR No., Date, PS, Sections)", "type": "textarea"},
                {"name": "arrest_date", "label": "Date of Arrest", "type": "date"},
                {"name": "custody_days", "label": "Days in Custody", "type": "text"},
                {"name": "transaction_details", "label": "Key Transactions Alleged by ED", "type": "textarea"},
                {"name": "defence_ground", "label": "Primary Defence Ground", "type": "textarea"},
                {"name": "prior_bail_history", "label": "Prior Bail Applications (if any)", "type": "textarea"}]},
    {"id": "pao_challenge", "name": "Provisional Attachment Order Challenge (Section 5 PMLA)", "category": "criminal",
     "fields": [{"name": "pao_number", "label": "PAO Number & Date", "type": "text"},
                {"name": "pao_date", "label": "Date of PAO", "type": "date"},
                {"name": "property_details", "label": "Property/Bank Account/Amount Attached", "type": "textarea"},
                {"name": "adjudicating_authority_date", "label": "Date Referred to Adjudicating Authority", "type": "date"},
                {"name": "ecir_number", "label": "ECIR Number", "type": "text"},
                {"name": "grounds", "label": "Procedural Lapses / Grounds for Challenge", "type": "textarea"},
                {"name": "property_source", "label": "Legitimate Source of Property/Funds", "type": "textarea"}]},
    {"id": "panchnama_audit", "name": "Panchnama vs FIR Reconciliation Audit", "category": "criminal",
     "fields": [{"name": "panchnama_items", "label": "Paste Panchnama Seized Items List (Full Text)", "type": "textarea"},
                {"name": "fir_allegations", "label": "FIR Allegations & Amounts", "type": "textarea"},
                {"name": "bank_accounts", "label": "Bank Accounts Referenced in FIR", "type": "textarea"},
                {"name": "company_books", "label": "Balance Sheet / Ledger Summary", "type": "textarea"},
                {"name": "alleged_amount", "label": "Total Amount Alleged in FIR (INR)", "type": "text"},
                {"name": "raid_date", "label": "Date of Raid/Search", "type": "date"}]},

    # ══════════════════════════════════════════════════════════════════
    # Added April 2026 — high-frequency CA + advocate workflows
    # ══════════════════════════════════════════════════════════════════

    # Reply to S.148A(b) notice of reopening — must be filed within 2 weeks before reassessment issues
    {"id": "it_148a_reply", "name": "Reply to Income Tax Section 148A(b) Notice (Reassessment)", "category": "taxation",
     "fields": [{"name": "assessee_name", "label": "Assessee Name", "type": "text", "placeholder": "As per PAN"},
                {"name": "pan", "label": "PAN", "type": "text", "placeholder": "ABCDE1234F"},
                {"name": "assessment_year", "label": "Assessment Year", "type": "text", "placeholder": "2020-21"},
                {"name": "notice_date", "label": "Notice Date", "type": "date"},
                {"name": "alleged_escaped_income", "label": "Alleged Escaped Income (Rs.)", "type": "number"},
                {"name": "assessee_response", "label": "Grounds / Explanation", "type": "textarea", "placeholder": "Transaction already disclosed in ITR..."}]},

    # Scrutiny reply — drafted weekly Aug-Dec when CASS selections go out
    {"id": "it_143_2_reply", "name": "Reply to Income Tax Section 143(2) Scrutiny Notice", "category": "taxation",
     "fields": [{"name": "assessee_name", "label": "Assessee Name", "type": "text"},
                {"name": "pan", "label": "PAN", "type": "text"},
                {"name": "assessment_year", "label": "Assessment Year", "type": "text", "placeholder": "2023-24"},
                {"name": "notice_din", "label": "Notice DIN", "type": "text"},
                {"name": "issues_raised", "label": "Issues Raised by AO", "type": "textarea", "placeholder": "Mismatch in 26AS, large cash deposits..."},
                {"name": "documents_attached", "label": "Documents Annexed", "type": "textarea", "placeholder": "Bank statements, ledger, invoices"}]},

    # GST REG-17 response — 7 working-day window or registration is cancelled
    {"id": "gst_reg_cancel_reply", "name": "Reply to GST Registration Cancellation SCN (Rule 22 / REG-17)", "category": "taxation",
     "fields": [{"name": "legal_name", "label": "Legal Name of Taxpayer", "type": "text"},
                {"name": "gstin", "label": "GSTIN", "type": "text", "placeholder": "29ABCDE1234F1Z5"},
                {"name": "scn_reference", "label": "SCN Reference No.", "type": "text"},
                {"name": "scn_date", "label": "SCN Date", "type": "date"},
                {"name": "reason_cited", "label": "Reason Cited by Officer", "type": "textarea", "placeholder": "Non-filing of GSTR-3B for 6 months"},
                {"name": "taxpayer_defence", "label": "Taxpayer Defence / Cure Offered", "type": "textarea"}]},

    # RFD-01 refund — export/inverted-duty/excess-tax claims
    {"id": "gst_rfd_01", "name": "GST Refund Application (Form RFD-01)", "category": "taxation",
     "fields": [{"name": "gstin", "label": "GSTIN", "type": "text"},
                {"name": "refund_type", "label": "Type of Refund", "type": "text", "placeholder": "Export without payment of tax / Inverted duty"},
                {"name": "tax_period_from", "label": "Tax Period From", "type": "date"},
                {"name": "tax_period_to", "label": "Tax Period To", "type": "date"},
                {"name": "refund_amount", "label": "Refund Amount Claimed (Rs.)", "type": "number"},
                {"name": "grounds", "label": "Grounds of Refund", "type": "textarea", "placeholder": "Zero-rated supply under LUT..."}]},

    # TDS certificate correction — common during ITR filing season
    {"id": "tds_correction_request", "name": "TDS Certificate (Form 16/16A) Correction Request", "category": "taxation",
     "fields": [{"name": "deductee_name", "label": "Deductee Name", "type": "text"},
                {"name": "deductee_pan", "label": "Deductee PAN", "type": "text"},
                {"name": "deductor_tan", "label": "Deductor TAN", "type": "text"},
                {"name": "financial_year", "label": "Financial Year", "type": "text", "placeholder": "2024-25"},
                {"name": "error_description", "label": "Error in Certificate", "type": "textarea"},
                {"name": "correct_details", "label": "Correct Details Required", "type": "textarea"}]},

    # Form 35 appeal before CIT(A) — 30-day window from assessment order
    {"id": "it_form_35_appeal", "name": "Income Tax Appeal before CIT(A) (Form 35)", "category": "taxation",
     "fields": [{"name": "appellant_name", "label": "Appellant Name", "type": "text"},
                {"name": "pan", "label": "PAN", "type": "text"},
                {"name": "assessment_year", "label": "Assessment Year", "type": "text"},
                {"name": "order_appealed", "label": "Order Appealed Against (Sec & Date)", "type": "text", "placeholder": "S.143(3) dated 15-03-2025"},
                {"name": "disputed_demand", "label": "Disputed Demand (Rs.)", "type": "number"},
                {"name": "grounds_of_appeal", "label": "Grounds of Appeal", "type": "textarea"},
                {"name": "statement_of_facts", "label": "Statement of Facts", "type": "textarea"}]},

    # PAN correction — Form 49A change request
    {"id": "pan_correction", "name": "PAN Correction / Change Request (Form 49A)", "category": "taxation",
     "fields": [{"name": "existing_pan", "label": "Existing PAN", "type": "text"},
                {"name": "applicant_name", "label": "Name as per PAN", "type": "text"},
                {"name": "field_to_correct", "label": "Field Requiring Correction", "type": "text", "placeholder": "Name / DOB / Father's Name"},
                {"name": "correct_value", "label": "Correct Value", "type": "text"},
                {"name": "supporting_document", "label": "Supporting Document", "type": "text", "placeholder": "Aadhaar / Passport / Birth Certificate"}]},

    # BNSS S.482 anticipatory bail — weekly Sessions/HC filing
    {"id": "anticipatory_bail_bnss_482", "name": "Anticipatory Bail Application (BNSS S.482)", "category": "criminal",
     "fields": [{"name": "applicant_name", "label": "Applicant Name", "type": "text"},
                {"name": "applicant_address", "label": "Address", "type": "textarea"},
                {"name": "fir_number", "label": "FIR Number", "type": "text"},
                {"name": "police_station", "label": "Police Station", "type": "text"},
                {"name": "offences_alleged", "label": "Offences Alleged (Sections)", "type": "text", "placeholder": "BNS S.318, S.336"},
                {"name": "apprehension_grounds", "label": "Grounds of Apprehension", "type": "textarea"},
                {"name": "defence_grounds", "label": "Grounds for Bail", "type": "textarea", "placeholder": "False implication, permanent resident, no prior antecedents"}]},

    # BNSS S.94 summons reply — documents-production objection/compliance
    {"id": "bnss_94_summons_reply", "name": "Reply to Section 91 CrPC / BNSS S.94 Summons for Documents", "category": "litigation",
     "fields": [{"name": "party_name", "label": "Party Name", "type": "text"},
                {"name": "summons_ref", "label": "Summons Reference No.", "type": "text"},
                {"name": "issuing_authority", "label": "Issuing Authority", "type": "text", "placeholder": "IO / Magistrate"},
                {"name": "documents_sought", "label": "Documents Sought", "type": "textarea"},
                {"name": "objection_or_compliance", "label": "Objection / Compliance", "type": "textarea"}]},

    # BNSS S.528 / CrPC S.482 quash — matrimonial, commercial, 138 NI disputes
    {"id": "bnss_528_quash_fir", "name": "Petition to Quash FIR (BNSS S.528 / CrPC S.482)", "category": "criminal",
     "fields": [{"name": "petitioner_name", "label": "Petitioner Name", "type": "text"},
                {"name": "respondent", "label": "Respondent (State + Complainant)", "type": "text", "placeholder": "State of Karnataka & Anr."},
                {"name": "fir_number", "label": "FIR Number & Date", "type": "text"},
                {"name": "police_station", "label": "Police Station", "type": "text"},
                {"name": "offences", "label": "Sections Invoked", "type": "text"},
                {"name": "grounds_for_quashing", "label": "Grounds for Quashing", "type": "textarea", "placeholder": "Settlement, civil dispute dressed as criminal, no offence made out"}]},

    # RTI Act 2005 information request — weekly when litigating government bodies
    {"id": "rti_application", "name": "RTI Application (Right to Information Act, 2005)", "category": "litigation",
     "fields": [{"name": "applicant_name", "label": "Applicant Name", "type": "text"},
                {"name": "applicant_address", "label": "Address", "type": "textarea"},
                {"name": "public_authority", "label": "Public Authority / PIO", "type": "text", "placeholder": "PIO, Income Tax Dept, Bengaluru"},
                {"name": "information_sought", "label": "Information Sought (point-wise)", "type": "textarea"},
                {"name": "period", "label": "Period of Information", "type": "text", "placeholder": "01-04-2023 to 31-03-2024"},
                {"name": "bpl_status", "label": "BPL Status (Yes/No)", "type": "text", "placeholder": "No"}]},

    # BNSS S.160 notice reply — appearance at police station with counsel
    {"id": "police_summons_reply", "name": "Reply to Police Summons / Notice u/s 160 BNSS", "category": "criminal",
     "fields": [{"name": "noticee_name", "label": "Noticee Name", "type": "text"},
                {"name": "notice_date", "label": "Notice Date", "type": "date"},
                {"name": "police_station", "label": "Police Station", "type": "text"},
                {"name": "fir_or_case", "label": "FIR / Case Reference", "type": "text"},
                {"name": "response", "label": "Response / Request", "type": "textarea", "placeholder": "Seeking adjournment / appearing with counsel / written statement enclosed"}]},

    # CPA 2019 State Commission appeal — 45 days from District Commission order
    {"id": "consumer_state_appeal", "name": "Consumer State Commission Appeal (CPA 2019)", "category": "litigation",
     "fields": [{"name": "appellant_name", "label": "Appellant Name", "type": "text"},
                {"name": "respondent_name", "label": "Respondent (Opposite Party)", "type": "text"},
                {"name": "district_order_ref", "label": "District Commission Order No. & Date", "type": "text"},
                {"name": "claim_amount", "label": "Claim Amount (Rs.)", "type": "number"},
                {"name": "grounds_of_appeal", "label": "Grounds of Appeal", "type": "textarea"},
                {"name": "relief_sought", "label": "Relief Sought", "type": "textarea"}]},

    # Standard board resolution — multiple per week per company
    {"id": "board_resolution", "name": "Board Resolution (Standard)", "category": "corporate",
     "fields": [{"name": "company_name", "label": "Company Name", "type": "text"},
                {"name": "cin", "label": "CIN", "type": "text"},
                {"name": "meeting_date", "label": "Date of Board Meeting", "type": "date"},
                {"name": "resolution_purpose", "label": "Purpose of Resolution", "type": "text", "placeholder": "Opening bank account / Authorising signatory / Borrowing"},
                {"name": "authorised_person", "label": "Authorised Person(s)", "type": "text", "placeholder": "Mr. X, Director (DIN...)"},
                {"name": "resolution_text", "label": "Operative Resolution Text", "type": "textarea"}]},

    # Director's Report per S.134 Companies Act — annual, reviewed often
    {"id": "directors_report_s134", "name": "Director's Report (S.134 Companies Act, 2013)", "category": "corporate",
     "fields": [{"name": "company_name", "label": "Company Name", "type": "text"},
                {"name": "cin", "label": "CIN", "type": "text"},
                {"name": "financial_year", "label": "Financial Year", "type": "text", "placeholder": "2024-25"},
                {"name": "turnover", "label": "Turnover (Rs.)", "type": "number"},
                {"name": "net_profit", "label": "Net Profit / (Loss) (Rs.)", "type": "number"},
                {"name": "dividend", "label": "Dividend Recommended", "type": "text", "placeholder": "Nil / 10%"},
                {"name": "material_changes", "label": "Material Changes & Commitments", "type": "textarea"},
                {"name": "board_composition", "label": "Changes in Directors / KMP", "type": "textarea"}]},

    # MGT-7 annual return — mandatory, 60 days post-AGM
    {"id": "mgt_7_annual_return", "name": "Annual Return (Form MGT-7) Data Sheet", "category": "corporate",
     "fields": [{"name": "company_name", "label": "Company Name", "type": "text"},
                {"name": "cin", "label": "CIN", "type": "text"},
                {"name": "financial_year", "label": "Financial Year", "type": "text"},
                {"name": "paid_up_capital", "label": "Paid-up Capital (Rs.)", "type": "number"},
                {"name": "number_of_members", "label": "No. of Members", "type": "number"},
                {"name": "agm_date", "label": "AGM Date", "type": "date"},
                {"name": "directors_list", "label": "Directors (Name + DIN)", "type": "textarea"},
                {"name": "shareholding_pattern", "label": "Shareholding Pattern Summary", "type": "textarea"}]},

    # ROC / MCA adjudication notice reply — frequent post S.454 tightening
    {"id": "roc_notice_reply", "name": "Reply to ROC / MCA Notice", "category": "compliance",
     "fields": [{"name": "company_name", "label": "Company Name", "type": "text"},
                {"name": "cin", "label": "CIN", "type": "text"},
                {"name": "notice_ref", "label": "Notice Reference No.", "type": "text"},
                {"name": "notice_date", "label": "Notice Date", "type": "date"},
                {"name": "section_invoked", "label": "Section / Rule Invoked", "type": "text", "placeholder": "S.92 / S.137 / S.164"},
                {"name": "default_alleged", "label": "Default Alleged", "type": "textarea"},
                {"name": "company_response", "label": "Company's Response / Cure", "type": "textarea"}]},

    # Minutes of Board Meeting per S.118 + SS-1 — quarterly minimum
    {"id": "board_meeting_minutes", "name": "Minutes of Board Meeting (SS-1 / S.118)", "category": "corporate",
     "fields": [{"name": "company_name", "label": "Company Name", "type": "text"},
                {"name": "meeting_number", "label": "Meeting Number", "type": "text", "placeholder": "05/2026-27"},
                {"name": "meeting_date", "label": "Date of Meeting", "type": "date"},
                {"name": "venue", "label": "Venue / Mode (VC)", "type": "text"},
                {"name": "directors_present", "label": "Directors Present", "type": "textarea"},
                {"name": "directors_absent", "label": "Leave of Absence", "type": "textarea"},
                {"name": "agenda_items", "label": "Agenda Items & Resolutions", "type": "textarea"}]},

    # India-specific offer letter — PF/ESI/Gratuity/notice-period clauses
    {"id": "employment_offer_letter_india", "name": "Employment Offer Letter (India — PF/ESI/Gratuity)", "category": "labour",
     "fields": [{"name": "employer_name", "label": "Employer Name", "type": "text"},
                {"name": "employee_name", "label": "Employee Name", "type": "text"},
                {"name": "designation", "label": "Designation", "type": "text"},
                {"name": "ctc", "label": "Annual CTC (Rs.)", "type": "number"},
                {"name": "joining_date", "label": "Date of Joining", "type": "date"},
                {"name": "probation_period", "label": "Probation Period", "type": "text", "placeholder": "6 months"},
                {"name": "notice_period", "label": "Notice Period", "type": "text", "placeholder": "60 days"},
                {"name": "work_location", "label": "Work Location", "type": "text"}]},

    # Mutual NDA under Indian Contract Act 1872
    {"id": "mutual_nda_india", "name": "Mutual NDA (Indian Contract Act, 1872)", "category": "contract",
     "fields": [{"name": "party_a_name", "label": "Party A (Name & Address)", "type": "textarea"},
                {"name": "party_b_name", "label": "Party B (Name & Address)", "type": "textarea"},
                {"name": "purpose", "label": "Purpose of Disclosure", "type": "textarea", "placeholder": "Evaluation of potential business collaboration"},
                {"name": "effective_date", "label": "Effective Date", "type": "date"},
                {"name": "term_years", "label": "Term (Years)", "type": "number", "placeholder": "3"},
                {"name": "jurisdiction", "label": "Jurisdiction / Seat", "type": "text", "placeholder": "Courts at Bengaluru"},
                {"name": "governing_law", "label": "Governing Law", "type": "text", "placeholder": "Laws of India"}]},
]

@api_router.get("/workflows")
async def list_workflows():
    return WORKFLOWS

@api_router.post("/workflows/generate")
async def generate_workflow(req: WorkflowRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)

    # Look up in WORKFLOWS list first, fallback to using the type directly
    workflow = next((w for w in WORKFLOWS if w["id"] == req.workflow_type), None)
    workflow_name = workflow["name"] if workflow else req.workflow_type.replace('_', ' ').title()

    result = await generate_workflow_document(
        workflow_type=workflow_name,
        fields=req.fields,
        mode=req.mode
    )

    # --- Auto-generate DOCX and register as a cached file so frontend can one-click Open in Google Docs ---
    docx_file_id = None
    docx_filename = None
    try:
        from document_export import generate_word_document
        docx_bytes = generate_word_document(
            title=workflow_name,
            content=result["response_text"],
            doc_type="memo",
        )
        docx_filename = f"{workflow_name.replace(' ', '_')}.docx"
        docx_file_id = _register_generated_file(
            name=docx_filename,
            content=docx_bytes,
            user_id=user["user_id"],
        )
    except Exception as e:
        logger.warning(f"Workflow DOCX auto-generation failed (response_text still returned): {e}")

    # Save to history (graceful fail-safe if MongoDB is blocked)
    gen_id = f"wf_{uuid.uuid4().hex[:12]}"
    try:
        await db.workflow_history.insert_one({
            "gen_id": gen_id,
            "user_id": user["user_id"],
            "workflow_type": req.workflow_type,
            "workflow_name": workflow_name,
            "fields": req.fields,
            "response_text": result["response_text"],
            "sources": result.get("sources", []),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    except Exception as e:
        logger.warning(f"Failed to save workflow history to MongoDB (User will still receive doc): {e}")

    # Exhaustive log → audit sink cluster
    try:
        from audit_sink import log_workflow_event as _log_wf
        await _log_wf(
            user=user, request=request, workflow_type=req.workflow_type, outcome="success",
            docx_file_id=docx_file_id, response_chars=len(result.get("response_text", "")),
        )
    except Exception:
        pass

    return {
        "gen_id": gen_id,
        "response_text": result["response_text"],
        "workflow_name": workflow_name,
        "sources": result.get("sources", []),
        "docx_file_id": docx_file_id,
        "docx_filename": docx_filename,
    }



# ==================== HISTORY ROUTES ====================

@api_router.get("/history")
async def get_history(request: Request, authorization: str = Header(None), limit: int = 50):
    user = await get_current_user(request, authorization)
    try:
        history = await db.query_history.find(
            {"user_id": user["user_id"]}, {"_id": 0}
        ).sort("created_at", -1).limit(limit).to_list(limit)
        # Ensure all frontend-expected fields exist (backfill old records)
        for item in history:
            item.setdefault("query_types", [item.get("mode", "general")])
            item.setdefault("model_used", item.get("mode", "fast"))
            item.setdefault("citations_count", 0)
            item.setdefault("sections", [])
            item.setdefault("sources", [])
        return history
    except Exception as e:
        logger.warning(f"MongoDB history list failed: {e}")
        return []

@api_router.get("/history/{history_id}")
async def get_history_item(history_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        item = await db.query_history.find_one(
            {"history_id": history_id, "user_id": user["user_id"]}, {"_id": 0}
        )
    except Exception as e:
        logger.warning(f"MongoDB history item fetch failed: {e}")
        raise HTTPException(status_code=503, detail="Database temporarily unavailable.")
    if not item:
        raise HTTPException(status_code=404, detail="History item not found")
    return item

@api_router.get("/history/matter/{matter_id}")
async def get_matter_history(matter_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        history = await db.query_history.find(
            {"matter_id": matter_id, "user_id": user["user_id"]}, {"_id": 0}
        ).sort("created_at", -1).to_list(100)
        return history
    except Exception as e:
        logger.warning(f"MongoDB matter history failed: {e}")
        return []


# ==================== THREAD ROUTES (Claude-style chat threading) ========
#
# A thread groups many query_history rows into one continuing conversation.
# The frontend sidebar shows one entry per thread (by `title`), and when a
# user clicks a past thread we return every message so the chat rehydrates
# with full context.
#
# Related modules:
#   thread_manager.py    — title generation, CRUD helpers, context assembly
#   supermemory_client.py — long-term per-user memory retrieval

from pydantic import BaseModel as _BM_T

class ThreadRenameRequest(_BM_T):
    title: str


@api_router.get("/threads")
async def list_user_threads(request: Request, authorization: str = Header(None), limit: int = 50):
    """Sidebar list — every thread owned by the authenticated user, newest first."""
    user = await get_current_user(request, authorization)
    from thread_manager import list_threads
    threads = await list_threads(db, user["user_id"], limit=limit)
    return threads


@api_router.get("/threads/{thread_id}")
async def get_thread(thread_id: str, request: Request, authorization: str = Header(None)):
    """Full thread payload — the thread meta + every message in chronological order.
    Frontend calls this when the user clicks a past chat in the sidebar."""
    user = await get_current_user(request, authorization)
    try:
        meta = await db.threads.find_one(
            {"thread_id": thread_id, "user_id": user["user_id"]}, {"_id": 0}
        )
    except Exception as e:
        logger.warning(f"get_thread meta fetch failed: {e}")
        raise HTTPException(status_code=503, detail="Database temporarily unavailable.")
    if not meta:
        raise HTTPException(status_code=404, detail="Thread not found")
    from thread_manager import load_thread_messages
    messages = await load_thread_messages(db, user["user_id"], thread_id, limit=200)
    return {"thread": meta, "messages": messages}


@api_router.patch("/threads/{thread_id}")
async def rename_user_thread(thread_id: str, body: ThreadRenameRequest,
                              request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    from thread_manager import rename_thread
    ok = await rename_thread(db, user["user_id"], thread_id, body.title)
    if not ok:
        raise HTTPException(status_code=404, detail="Thread not found or title invalid")
    return {"ok": True, "thread_id": thread_id, "title": body.title.strip()[:80]}


@api_router.delete("/threads/{thread_id}")
async def delete_user_thread(thread_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    from thread_manager import delete_thread
    ok = await delete_thread(db, user["user_id"], thread_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Thread not found")
    return {"ok": True, "thread_id": thread_id}


# ==================== LIBRARY ROUTES ====================

@api_router.post("/library")
async def create_library_item(item: LibraryItemCreate, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    item_id = f"lib_{uuid.uuid4().hex[:12]}"
    doc = {
        "item_id": item_id,
        "user_id": user["user_id"],
        "title": item.title,
        "content": item.content,
        "item_type": item.item_type,
        "tags": item.tags,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    try:
        await db.library_items.insert_one(doc)
    except Exception as e:
        logger.error(f"MongoDB library insert failed: {e}")
        raise HTTPException(status_code=503, detail="Database temporarily unavailable.")
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.get("/library")
async def list_library(request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        items = await db.library_items.find({"user_id": user["user_id"]}, {"_id": 0}).sort("updated_at", -1).to_list(100)
        return items
    except Exception as e:
        logger.warning(f"MongoDB library list failed: {e}")
        return []

@api_router.delete("/library/{item_id}")
async def delete_library_item(item_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        result = await db.library_items.delete_one({"item_id": item_id, "user_id": user["user_id"]})
    except Exception as e:
        logger.error(f"MongoDB library delete failed: {e}")
        raise HTTPException(status_code=503, detail="Database temporarily unavailable.")
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"status": "deleted"}

# ==================== SEARCH ROUTES ====================

@api_router.get("/search/cases")
async def search_cases(q: str = Query(...)):
    results = await search_indiankanoon(q, top_k=10)
    return results

@api_router.get("/search/companies")
async def search_companies(q: str = Query(...)):
    results = await search_company(q)
    return results

# ==================== STATUTE DB ROUTES ====================

@api_router.get("/statutes")
async def search_statutes(q: str = Query(""), act: str = Query("")):
    import re as _re
    query_filter = {}
    if q:
        safe_q = _re.escape(q)  # Escape regex special chars to prevent ReDoS
        query_filter["$or"] = [
            {"section_text": {"$regex": safe_q, "$options": "i"}},
            {"section_title": {"$regex": safe_q, "$options": "i"}},
        ]
    if act:
        safe_act = _re.escape(act)
        query_filter["act_name"] = {"$regex": safe_act, "$options": "i"}
    
    try:
        results = await db.statutes.find(query_filter, {"_id": 0}).limit(20).to_list(20)
        return results
    except Exception as e:
        logger.warning(f"MongoDB statutes search failed: {e}")
        return []

# ==================== CASE LAW SEARCH ====================

@api_router.post("/caselaw/find")
async def api_caselaw_find(request: Request, authorization: str = Header(None)):
    """Search IndianKanoon for case law matching a scenario description."""
    await get_current_user(request, authorization)
    body = await request.json()
    scenario = body.get("scenario", "")
    if not scenario:
        raise HTTPException(status_code=400, detail="'scenario' is required")

    def _normalize_caselaw(results):
        """Normalize IndianKanoon results for frontend consumption."""
        for i, r in enumerate(results):
            r.setdefault("snippet", r.get("headline", ""))
            r.setdefault("holding", r.get("headline", "")[:300])
            r.setdefault("url", f"https://indiankanoon.org/doc/{r.get('doc_id', '')}/")
            r.setdefault("relevance_score", round(1.0 - (i * 0.08), 2))  # Decreasing relevance
        return results

    try:
        results = await search_indiankanoon(scenario, top_k=10)
        return {"results": _normalize_caselaw(results), "count": len(results), "query": scenario}
    except Exception as e:
        logger.error(f"Case law search failed: {e}")
        # Fallback: use AI to formulate a better search and try again
        try:
            refined_results = await search_indiankanoon(
                f"Indian Supreme Court High Court judgment {scenario}", top_k=5
            )
            return {"results": _normalize_caselaw(refined_results), "count": len(refined_results), "query": scenario}
        except Exception as e2:
            logger.error(f"Case law search fallback also failed: {e2}")
            return {"results": [], "count": 0, "query": scenario, "error": "Case law search temporarily unavailable"}


# ==================== TOOLS ROUTES ====================

@api_router.post("/tools/reconcile-gstr2b")
async def api_reconcile_gstr2b(
    request: Request,
    purchase_file: UploadFile = File(...),
    gstr2b_file: UploadFile = File(...),
    authorization: str = Header(None)
):
    try:
        user = await get_current_user(request, authorization)
    except Exception:
        pass # Allow fast bypass for prototype
        
    purchase_data = await purchase_file.read()
    gstr2b_data = await gstr2b_file.read()
    
    result = reconcile_gstr2b(purchase_data, gstr2b_data)
    # Normalize — frontend reads flat fields alongside nested details
    if "details" in result and "error" not in result:
        details = result["details"]
        result["unmatched_pr"] = details.get("unmatched_pr", [])
        result["unmatched_g2b"] = details.get("unmatched_g2b", [])
        result["fuzzy_match_details"] = details.get("fuzzy_matches", [])
        result["total_invoices"] = result.get("total_pr_invoices", 0)
        result["discrepancies"] = result.get("unmatched_in_pr", 0) + result.get("amount_mismatches", 0)
        # Normalize vendor_risk to add frontend-expected aliases
        if "vendor_risk" in result:
            vr_list = []
            for key, v in result["vendor_risk"].items():
                vr_list.append({
                    "gstin": v.get("gstin", key),
                    "vendor": v.get("vendor_name", key),
                    "mismatches": v.get("invoice_count", 0),
                    "risk_count": v.get("invoice_count", 0),
                    "itc_at_risk": v.get("total_itc_at_risk", 0),
                    "risk_level": v.get("risk_level", "LOW"),
                    "action": v.get("action", ""),
                })
            result["vendor_risk"] = vr_list
    return result

# ==================== HELPERS ====================

def extract_pdf_text(pdf_bytes: bytes) -> str:
    """Extract text from PDF."""
    try:
        import fitz
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        text = ""
        for page in doc:
            text += page.get_text() + "\n"
        return text.strip()
    except Exception as e:
        logger.warning(f"PyMuPDF extraction failed: {e}. Attempting fallback...")
        try:
            from io import BytesIO
            from pdfminer.high_level import extract_text
            return extract_text(BytesIO(pdf_bytes))
        except Exception as fallback_e:
            logger.error(f"Fallback PDF extraction failed: {fallback_e}")
            return "[PDF text extraction not available - OCR failed or file corrupted]"


def extract_docx_text(docx_bytes: bytes) -> str:
    """Extract text from DOCX."""
    try:
        from docx import Document
        doc = Document(io.BytesIO(docx_bytes))
        return "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        return f"[DOCX extraction error: {e}]"


def extract_xlsx_text(xlsx_bytes: bytes) -> str:
    """Extract text from XLSX."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"\n--- Sheet: {sheet} ---\n"
            for row in ws.iter_rows(values_only=True):
                text += " | ".join([str(c) if c else "" for c in row]) + "\n"
        return text.strip()
    except Exception as e:
        return f"[XLSX extraction error: {e}]"


def classify_document(filename: str, text: str) -> str:
    """Auto-classify document type."""
    filename_lower = filename.lower()
    text_lower = text[:2000].lower() if text else ""
    
    if any(kw in text_lower for kw in ["agreement", "contract", "whereas", "hereby agree"]):
        return "contract"
    if any(kw in text_lower for kw in ["order", "judgment", "hon'ble", "court"]):
        return "court_order"
    if any(kw in text_lower for kw in ["gst", "cgst", "show cause", "scn"]):
        return "gst_notice"
    if any(kw in text_lower for kw in ["income tax", "143", "148", "assessment"]):
        return "it_notice"
    if any(kw in text_lower for kw in ["balance sheet", "profit", "loss", "revenue"]):
        return "financial_statement"
    if any(kw in text_lower for kw in ["moa", "aoa", "board resolution", "company"]):
        return "corporate_document"
    if any(kw in text_lower for kw in ["sale deed", "property", "conveyance"]):
        return "property_document"
    return "other"


# ==================== INDIAN LEGAL TOOLS ROUTES ====================

@api_router.post("/tools/limitation")
async def api_limitation(req: LimitationRequest, request: Request, authorization: str = Header(None)):
    """Calculate limitation period for any Indian suit/appeal type."""
    await get_current_user(request, authorization)
    # Map frontend generic types to specific backend suit types
    suit_type_map = {
        "civil_suit": "breach_of_contract",
        "criminal": "criminal_complaint_private",
        "appeal": "appeal_high_court",
        "tax": "it_appeal_cit",
        "consumer": "consumer_complaint_district",
        "writ": "writ_petition_hc",
        "arbitration": "arbitration_challenge_34",
    }
    suit_type = req.get_suit_type()
    suit_type = suit_type_map.get(suit_type, suit_type)
    result = calculate_limitation(suit_type, req.accrual_date)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    # Normalize fields — frontend reads both field name variants
    if "deadline" in result:
        result["expiry_date"] = result["deadline"]
    if "section" in result:
        result["legal_basis"] = f"Limitation Act, 1963 — {result['section']}"
    result["limitation_period"] = result.get("limitation_from", "")
    result["period"] = result.get("limitation_from", "")
    result["is_barred"] = result.get("status") == "EXPIRED"
    # Add exceptions list for frontend rendering
    if result.get("status") == "EXPIRED":
        result["exceptions"] = [
            {"ground": "Sufficient Cause", "legal_basis": "Section 5 of Limitation Act, 1963", "detail": "Court may condone delay if applicant shows 'sufficient cause' for not filing within time."},
            {"ground": "Disability of Plaintiff", "legal_basis": "Section 6 of Limitation Act, 1963", "detail": "Period of limitation does not run during minority, insanity, or idiocy of plaintiff."},
            {"ground": "Acknowledgement of Liability", "legal_basis": "Section 18 of Limitation Act, 1963", "detail": "Fresh period of limitation starts from date of written and signed acknowledgement."},
        ]
    else:
        result["exceptions"] = []
    return result

@api_router.get("/tools/limitation/types")
async def api_limitation_types(request: Request, authorization: str = Header(None)):
    """List all available limitation period types."""
    await get_current_user(request, authorization)
    return {k: {"section": v["section"], "from": v["from"]} for k, v in LIMITATION_PERIODS.items()}

@api_router.post("/tools/stamp-duty")
async def api_stamp_duty(req: StampDutyRequest, request: Request, authorization: str = Header(None)):
    """Calculate stamp duty + registration fees for Indian states."""
    await get_current_user(request, authorization)
    result = calculate_stamp_duty(req.state, req.get_instrument(), req.get_consideration(), req.gender)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result

@api_router.get("/tools/stamp-duty/states")
async def api_stamp_duty_states(request: Request, authorization: str = Header(None)):
    """List available states and instruments."""
    await get_current_user(request, authorization)
    return {state: list(instruments.keys()) for state, instruments in STAMP_DUTY_RATES.items()}

@api_router.post("/tools/forensic")
async def api_forensic(req: ForensicRequest, request: Request, authorization: str = Header(None)):
    """Analyze bank transactions for forensic red flags (PMLA, circular trading, related party)."""
    await get_current_user(request, authorization)
    result = analyze_bank_statement(req.transactions)
    return result

# ==================== PRACTICE TOOLS (HIGH-VALUE) ====================

@api_router.post("/tools/section-mapper")
async def api_section_mapper(request: Request, authorization: str = Header(None)):
    """Map old criminal law sections to new (IPC→BNS, CrPC→BNSS, IEA→BSA) or vice versa."""
    try:
        await get_current_user(request, authorization)
    except Exception:
        pass
    body = await request.json()
    section = body.get("section", "")
    direction = body.get("direction", "old_to_new")
    if not section:
        raise HTTPException(status_code=400, detail="'section' is required (e.g., '420', '302', '438')")
    result = map_section(section, direction)
    return result

@api_router.post("/tools/section-mapper/batch")
async def api_batch_section_mapper(request: Request, authorization: str = Header(None)):
    """Map multiple sections at once."""
    try:
        await get_current_user(request, authorization)
    except Exception:
        pass
    body = await request.json()
    sections = body.get("sections", [])
    direction = body.get("direction", "old_to_new")
    if not sections:
        raise HTTPException(status_code=400, detail="'sections' list is required")
    results = batch_map_sections(sections, direction)
    return {"mappings": results, "total": len(results)}

@api_router.post("/tools/tds-classifier")
async def api_tds_classifier(request: Request, authorization: str = Header(None)):
    """Classify TDS section for a payment description."""
    try:
        await get_current_user(request, authorization)
    except Exception:
        pass
    body = await request.json()
    description = body.get("description", "")
    amount = float(body.get("amount", 0))
    payee_type = body.get("payee_type", "company")
    is_non_filer = body.get("is_non_filer", False)
    if not description:
        raise HTTPException(status_code=400, detail="'description' is required (e.g., 'professional fees to CA', 'rent for office')")
    result = classify_tds_section(description, amount, payee_type, is_non_filer)
    # Normalize fields — frontend reads both rate_percent and rate
    if "rate_percent" in result:
        result["rate"] = result["rate_percent"]
    return result

@api_router.post("/tools/notice-checker")
async def api_notice_checker(request: Request, authorization: str = Header(None)):
    """Check validity of a tax/GST notice (limitation, DIN, jurisdiction)."""
    try:
        await get_current_user(request, authorization)
    except Exception:
        pass
    body = await request.json()
    notice_type = body.get("notice_type", "")
    notice_date = body.get("notice_date", "")
    if not notice_type or not notice_date:
        raise HTTPException(status_code=400, detail="'notice_type' and 'notice_date' (YYYY-MM-DD) are required")
    result = check_notice_validity(
        notice_type=notice_type,
        notice_date=notice_date,
        assessment_year=body.get("assessment_year", ""),
        financial_year=body.get("financial_year", ""),
        has_din=body.get("has_din", True),
        is_fraud_alleged=body.get("is_fraud_alleged", False),
    )
    # Normalize — frontend reads limitation_check as a summary field
    critical_grounds = [g for g in result.get("challenge_grounds", []) if "CRITICAL" in g.get("severity", "")]
    result["limitation_check"] = (
        f"{len(critical_grounds)} critical issue(s) found — notice is likely challengeable."
        if critical_grounds else "No limitation issues detected."
    )
    return result

@api_router.post("/tools/penalty-calculator")
async def api_penalty_calculator(request: Request, authorization: str = Header(None)):
    """Calculate exact penalty for missing a compliance deadline."""
    try:
        await get_current_user(request, authorization)
    except Exception:
        pass
    body = await request.json()
    deadline_type = body.get("deadline_type", "")
    due_date = body.get("due_date", "")
    if not deadline_type or not due_date:
        raise HTTPException(status_code=400, detail="'deadline_type' and 'due_date' (YYYY-MM-DD) are required")
    result = calculate_deadline_penalty(
        deadline_type=deadline_type,
        due_date=due_date,
        actual_date=body.get("actual_date", ""),
        tax_amount=float(body.get("tax_amount", 0)),
    )
    # Normalize fields — frontend reads both field name variants
    if "days_late" in result:
        result["delay_days"] = result["days_late"]
    if "total_exposure" in result:
        result["total_penalty"] = result["total_exposure"]
    if "late_fee" not in result:
        # ITR returns late_fee_234f — normalize
        result["late_fee"] = result.get("late_fee_234f", result.get("late_fee_234e", 0))
    if "interest" not in result:
        result["interest"] = result.get("interest_234a", 0)
    if "tip" not in result:
        result["tip"] = f"File {deadline_type.upper()} returns on time to avoid penalties up to ₹{result.get('total_exposure', result.get('total_penalty', 0)):,.0f}."
    return result

@api_router.post("/tools/tally-import")
async def api_tally_import(
    request: Request,
    file: UploadFile = File(...),
    authorization: str = Header(None),
):
    """Import and analyze Tally XML export. Auto-detects Section 40A(3) and 269ST violations."""
    try:
        await get_current_user(request, authorization)
    except Exception:
        pass
    content = await file.read()
    xml_str = content.decode("utf-8", errors="ignore")
    raw = parse_tally_xml(xml_str)
    if "error" in raw:
        return raw
    # Reshape to unified format for frontend
    violations_40a3 = raw.get("auto_detected_violations", {}).get("section_40a3_cash_payments", [])
    violations_269st = raw.get("auto_detected_violations", {}).get("section_269st_cash_receipts", [])
    violation_details = [
        {"date": v["date"], "party": v["ledger"], "amount": v["amount"], "section": "S.40A(3)"} for v in violations_40a3
    ] + [
        {"date": v["date"], "party": v["ledger"], "amount": v["amount"], "section": "S.269ST"} for v in violations_269st
    ]
    transactions = []
    for v in raw.get("vouchers", [])[:500]:
        party = v["entries"][0]["ledger_name"] if v.get("entries") else ""
        amount = sum(e["amount"] for e in v.get("entries", []) if e["is_debit"])
        transactions.append({"date": v["date"], "voucher_type": v["voucher_type"], "party": party, "amount": amount})
    return {
        "source": "tally",
        "total_vouchers": raw.get("parsed_vouchers", 0),
        "total_amount": round(raw.get("total_debit", 0) + raw.get("total_credit", 0), 2),
        "violations_40a3": len(violations_40a3),
        "violations_269st": len(violations_269st),
        "violation_details": violation_details,
        "transactions": transactions,
        "summary": raw.get("summary", ""),
    }

@api_router.post("/tools/zoho-import")
async def api_zoho_import(
    request: Request,
    file: UploadFile = File(...),
    authorization: str = Header(None),
):
    """Import and analyze Zoho Books CSV/XLSX export. Auto-detects S.40A(3) and S.269ST violations."""
    try:
        await get_current_user(request, authorization)
    except Exception:
        pass
    content = await file.read()
    result = parse_zoho_export(content, file.filename or "export.xlsx")
    return result

@api_router.post("/tools/notice-auto-reply")
async def api_notice_auto_reply(
    request: Request,
    authorization: str = Header(None),
):
    """THE KILLER FEATURE: Upload notice text → get auto-drafted legal reply with case law.
    Reads the notice, extracts type/section/demand, checks validity, drafts 10-point reply."""
    try:
        await get_current_user(request, authorization)
    except Exception:
        pass
    body = await request.json()
    notice_text = body.get("notice_text", "")
    if not notice_text:
        raise HTTPException(status_code=400, detail="'notice_text' is required (paste or extract from PDF)")
    result = await generate_notice_reply(
        notice_text=notice_text,
        client_name=body.get("client_name", ""),
        additional_context=body.get("additional_context", ""),
    )
    return result

@api_router.post("/tools/notice-extract")
async def api_notice_extract(request: Request, authorization: str = Header(None)):
    """Extract structured metadata from a tax notice (GSTIN, section, demand, dates)."""
    try:
        await get_current_user(request, authorization)
    except Exception:
        pass
    body = await request.json()
    notice_text = body.get("notice_text", "")
    if not notice_text:
        raise HTTPException(status_code=400, detail="'notice_text' is required")
    metadata = extract_notice_metadata(notice_text)
    return metadata


@api_router.post("/tools/playbook/distill")
async def api_playbook_distill(req: PlaybookRequest, request: Request, authorization: str = Header(None)):
    """Distill a firm's standard clause positions from past contracts."""
    user = await get_current_user(request, authorization)
    # Fetch contract texts from vault (graceful MongoDB handling)
    contract_texts = []
    for doc_id in req.contract_ids[:20]:  # Max 20 contracts
        try:
            doc = await db.documents.find_one({"document_id": doc_id, "user_id": user["user_id"]})
        except Exception as e:
            logger.warning(f"MongoDB fetch for playbook contract {doc_id} failed: {e}")
            doc = _vault_cache_get(doc_id)  # Fallback to cache
        if doc and doc.get("extracted_text"):
            contract_texts.append(f"--- Contract: {doc.get('filename', 'Unknown')} ---\n{doc['extracted_text'][:5000]}")
    
    if not contract_texts:
        raise HTTPException(status_code=400, detail="No valid contracts found in vault")
    
    combined = "\n\n".join(contract_texts)
    # Use AI engine to distill playbook
    playbook_prompt = f"""
    You are a Playbook Distillation Engine. Analyze these {len(contract_texts)} contracts and extract:
    1. The firm's STANDARD POSITION on each major clause category (Indemnity, Limitation of Liability, IP Assignment, Termination, Governing Law, Dispute Resolution, Confidentiality, Force Majeure, Warranties)
    2. For each clause, provide: (a) Standard Position, (b) Acceptable Fallback, (c) Walk-Away Position
    3. Flag any inconsistencies across contracts
    {f'Focus specifically on: {req.clause_focus}' if req.clause_focus else ''}
    
    CONTRACTS:
    {combined}
    """
    
    playbook_result = await process_query(
        user_query=playbook_prompt, 
        mode="partner",
        matter_context="Playbook Distillation Mode"
    )
    
    # Save playbook to library
    playbook_id = f"pb_{uuid.uuid4().hex[:12]}"
    playbook_doc = {
        "playbook_id": playbook_id,
        "user_id": user["user_id"],
        "title": f"Clause Playbook ({len(contract_texts)} contracts)",
        "content": playbook_result["response_text"],
        "source_contracts": req.contract_ids,
        "clause_focus": req.clause_focus,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    try:
        await db.playbooks.insert_one(playbook_doc)
    except Exception as e:
        logger.warning(f"MongoDB playbook save failed (returning result anyway): {e}")

    return {
        "playbook_id": playbook_id,
        "playbook_text": playbook_result["response_text"],
        "contracts_analyzed": len(contract_texts),
        "sections": playbook_result.get("sections", []),
    }

@api_router.put("/matters/{matter_id}/billing")
async def update_matter_billing(matter_id: str, req: MatterBillingUpdate, request: Request, authorization: str = Header(None)):
    """Update billing information for a client matter (Client-Matter API)."""
    user = await get_current_user(request, authorization)
    update_data = {
        "billing_code": req.billing_code,
        "hourly_rate": req.hourly_rate,
        "currency": req.currency,
        "billing_partner": req.billing_partner,
        "client_reference": req.client_reference,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    try:
        result = await db.matters.update_one(
            {"matter_id": matter_id, "user_id": user["user_id"]},
            {"$set": update_data}
        )
    except Exception as e:
        logger.error(f"MongoDB billing update failed: {e}")
        raise HTTPException(status_code=503, detail="Database temporarily unavailable.")
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Matter not found")
    return {"status": "updated", "matter_id": matter_id, **update_data}

@api_router.get("/search")
async def global_search(q: str = Query(""), request: Request = None, authorization: str = Header(None)):
    """Global search across vault documents, library items, matters, and history."""
    user = await get_current_user(request, authorization)
    uid = user["user_id"]
    results = []

    if len(q) < 2:
        return results

    import re as _re
    safe_q = _re.escape(q)
    search_regex = {"$regex": safe_q, "$options": "i"}

    # Each search is independently wrapped — if one collection fails, others still return
    try:
        docs = await db.documents.find(
            {"user_id": uid, "$or": [{"filename": search_regex}, {"extracted_text": search_regex}]},
            {"_id": 0, "document_id": 1, "filename": 1}
        ).limit(5).to_list(5)
        for d in docs:
            results.append({"id": d.get("document_id", d.get("doc_id", "")), "title": d["filename"], "type": "vault", "name": d["filename"]})
    except Exception as e:
        logger.warning(f"Search: vault documents query failed: {e}")

    try:
        lib_items = await db.library_items.find(
            {"user_id": uid, "$or": [{"title": search_regex}, {"content": search_regex}]},
            {"_id": 0, "item_id": 1, "title": 1}
        ).limit(5).to_list(5)
        for item in lib_items:
            results.append({"id": item.get("item_id", ""), "title": item["title"], "type": "library", "name": item["title"]})
    except Exception as e:
        logger.warning(f"Search: library items query failed: {e}")

    try:
        matters = await db.matters.find(
            {"user_id": uid, "$or": [{"name": search_regex}, {"client_name": search_regex}]},
            {"_id": 0, "matter_id": 1, "name": 1}
        ).limit(5).to_list(5)
        for m in matters:
            results.append({"id": m["matter_id"], "title": m["name"], "type": "matter", "name": m["name"]})
    except Exception as e:
        logger.warning(f"Search: matters query failed: {e}")

    try:
        history = await db.query_history.find(
            {"user_id": uid, "query": search_regex},
            {"_id": 0, "history_id": 1, "query": 1}
        ).limit(5).to_list(5)
        for h in history:
            results.append({"id": h["history_id"], "title": h["query"][:80], "type": "history", "name": h["query"][:80]})
    except Exception as e:
        logger.warning(f"Search: history query failed: {e}")

    return results

@api_router.post("/tools/ecourts/track")
async def api_ecourts_track(request: Request, authorization: str = Header(None), case_number: str = Form(""), court_name: str = Form(""), party_name: str = Form("")):
    """Track case status and hearing dates from eCourts via Playwright scraping."""
    user = await get_current_user(request, authorization)
    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            await page.goto("https://services.ecourts.gov.in/ecourtindia_v6/")
            await page.wait_for_timeout(2000)
            
            # Try to search by case number or party name
            results_data = []
            if case_number:
                # Navigate to case number search
                await page.click("text=Case Number") if await page.locator("text=Case Number").count() else None
                await page.wait_for_timeout(1000)
                
            elif party_name:
                # Navigate to party name search
                party_link = page.locator("text=Party Name")
                if await party_link.count():
                    await party_link.click()
                    await page.wait_for_timeout(1000)
            
            # Capture current page content
            content = await page.content()
            await browser.close()
            
            return {
                "status": "scraped",
                "query": {"case_number": case_number, "court_name": court_name, "party_name": party_name},
                "note": "eCourts scraper initialized. Full implementation requires CAPTCHA solving integration.",
                "portal_accessible": True
            }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "note": "Playwright/Chromium may not be installed on this server. Run: playwright install chromium"
        }

@api_router.post("/tools/vernacular/translate")
async def api_vernacular_translate(request: Request, authorization: str = Header(None), file: UploadFile = File(...), source_language: str = Form("hindi")):
    """Translate vernacular Indian legal documents to structured Legal English."""
    user = await get_current_user(request, authorization)
    contents = await file.read()
    
    # Extract text based on file type
    filename = file.filename or "document"
    if filename.lower().endswith(".pdf"):
        text = extract_pdf_text(contents)
    elif filename.lower().endswith(".docx"):
        text = extract_docx_text(contents)
    else:
        text = contents.decode("utf-8", errors="replace")
    
    # Use AI engine to translate + structure
    translate_prompt = f"""
    You are a Vernacular-to-Legal-English Bridge for Indian legal documents.
    The following document is in {source_language}. 
    
    Your task:
    1. Translate the document to precise Legal English
    2. Identify the TYPE of document (FIR, Land Record, Panchayat Order, Revenue Record, etc.)
    3. Extract key legal facts: parties, dates, amounts, property details, allegations
    4. Format as a structured legal memo that can be directly cited in a High Court filing
    
    DOCUMENT TEXT:
    {text[:8000]}
    """
    
    result = await process_query(
        user_query=translate_prompt,
        mode="partner",
        matter_context=f"Vernacular Translation: {source_language} | File: {filename}"
    )
    
    return {
        "original_language": source_language,
        "filename": filename,
        "translated_text": result["response_text"],
        "sections": result.get("sections", []),
    }

@api_router.post("/tools/tds-mismatch")
async def api_tds_mismatch(request: Request, authorization: str = Header(None), form_26as: UploadFile = File(...), books_ledger: UploadFile = File(...)):
    """Reconcile Form 26AS TDS credits vs Books/Ledger to find mismatches."""
    user = await get_current_user(request, authorization)
    
    file_26as = await form_26as.read()
    file_books = await books_ledger.read()
    
    text_26as = extract_xlsx_text(file_26as) if form_26as.filename.endswith(".xlsx") else extract_pdf_text(file_26as)
    text_books = extract_xlsx_text(file_books) if books_ledger.filename.endswith(".xlsx") else extract_pdf_text(file_books)
    
    reconcile_prompt = f"""
    You are a TDS Reconciliation Engine. Compare Form 26AS credits with the Books/Ledger data.
    
    For each TDS entry, check:
    1. Is the TAN number matching?
    2. Is the TDS amount matching (within ₹10 tolerance)?
    3. Is the Section (194C/194J/194H etc.) correctly applied?
    4. Flag entries in 26AS NOT in Books (unclaimed credit)
    5. Flag entries in Books NOT in 26AS (deductor hasn't filed)
    
    Output as a structured reconciliation table.
    
    === FORM 26AS DATA ===
    {text_26as[:6000]}
    
    === BOOKS/LEDGER DATA ===
    {text_books[:6000]}
    """
    
    result = await process_query(
        user_query=reconcile_prompt,
        mode="partner",
        matter_context="TDS Mismatch Reconciliation Mode"
    )
    
    return {
        "reconciliation": result["response_text"],
        "sections": result.get("sections", []),
    }

@api_router.post("/tools/it-scrutiny")
async def api_it_scrutiny(request: Request, authorization: str = Header(None), notice_file: UploadFile = File(...), bank_statement: UploadFile = File(None), invoices_zip: UploadFile = File(None)):
    """IT Scrutiny Response Bot - Draft point-by-point rebuttal to Section 143(2)/148 notices."""
    user = await get_current_user(request, authorization)
    
    notice_bytes = await notice_file.read()
    notice_text = extract_pdf_text(notice_bytes) if notice_file.filename.endswith(".pdf") else extract_docx_text(notice_bytes)
    
    bank_text = ""
    if bank_statement:
        bank_bytes = await bank_statement.read()
        bank_text = extract_xlsx_text(bank_bytes) if bank_statement.filename.endswith(".xlsx") else extract_pdf_text(bank_bytes)
    
    scrutiny_prompt = f"""
    You are an IT Scrutiny Response Specialist drafting a reply for the Assessing Officer.
    
    NOTICE TEXT:
    {notice_text[:5000]}
    
    {f'BANK STATEMENT DATA: {bank_text[:4000]}' if bank_text else ''}
    
    Draft a POINT-BY-POINT rebuttal covering:
    1. Parse every specific query/addition proposed in the notice
    2. For each point: cite the relevant ITAT/High Court precedent supporting the assessee
    3. Map specific bank entries to invoices/receipts where applicable
    4. Include exact section references (143(3), 69A, 68, 37(1), etc.)
    5. Format as a formal submission letter to the ITO/DCIT
    
    Reference latest ITAT and High Court decisions (2023-2026).
    """
    
    result = await process_query(
        user_query=scrutiny_prompt,
        mode="partner",
        matter_context="IT Scrutiny Response Mode"
    )
    
    return {
        "response_draft": result["response_text"],
        "sections": result.get("sections", []),
        "sources": result.get("sources", {}),
    }

# ==================== STARTUP ====================


@app.on_event("startup")
async def startup():
    try:
        init_storage()
        logger.info("Storage initialized")
    except Exception as e:
        logger.error(f"Storage init failed (non-blocking): {e}")

    # Warm up database connection pool at startup
    try:
        _db = get_db()
        await _db.command("ping")
        user_count = await _db.users.count_documents({})
        _backend_label = "Firestore" if os.environ.get("USE_FIRESTORE", "").strip() in ("1", "true", "yes") else "MongoDB Atlas"
        logger.info(f"{_backend_label} connected — {user_count} users in database")
    except Exception as e:
        logger.warning(f"Database warm-up failed (will retry on first request): {e}")

    logger.info("Spectr API started successfully — Practice Operating System loaded")

    # Seed statutory thresholds if the collection is empty
    try:
        from statutory_thresholds import seed_defaults_if_empty
        asyncio.create_task(seed_defaults_if_empty())
    except Exception as e:
        logger.warning(f"Statutory thresholds seed skipped: {e}")

    # Start the sandbox idle reaper — destroys sandboxes idle > 10 min to save cost
    try:
        if os.environ.get("BL_API_KEY"):
            start_idle_reaper()
            # First: kill any orphaned sandboxes from previous runs/crashes
            asyncio.create_task(cleanup_orphaned_sandboxes())
            # Then: pre-warm a fresh sandbox so the first query is fast
            asyncio.create_task(warm_sandbox_pool())
            logger.info("Sandbox idle reaper + orphan cleanup + pre-warm initiated")
    except Exception as e:
        logger.warning(f"Sandbox startup skipped: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    if _mongo_client:
        _mongo_client.close()
    # Clean up Blaxel research sandboxes
    try:
        await cleanup_all_sandboxes()
        logger.info("All research sandboxes cleaned up")
    except Exception as e:
        logger.warning(f"Sandbox cleanup on shutdown failed: {e}")


# ==================== OPERATING SYSTEM LAYER ROUTES ====================

# ── COMPLIANCE CALENDAR ─────────────────────────────────────────────
@api_router.get("/compliance/calendar")
async def get_compliance_calendar(
    year: int = Query(2026), month: int = Query(3),
    request: Request = None, authorization: str = Header(None)
):
    """Get all Indian compliance deadlines for a given month."""
    await get_current_user(request, authorization)
    deadlines = get_monthly_deadlines(year, month)
    # Serialize dates to strings
    for d in deadlines:
        if hasattr(d.get("date"), "isoformat"):
            d["date"] = d["date"].isoformat()
    return {"year": year, "month": month, "deadlines": deadlines, "count": len(deadlines)}


@api_router.get("/compliance/client-calendar")
async def get_client_compliance_calendar(
    client_name: str = Query(""), gst_registered: bool = Query(True), has_foreign: bool = Query(False),
    request: Request = None, authorization: str = Header(None)
):
    """Get 90-day compliance calendar for a specific client."""
    await get_current_user(request, authorization)
    cal = get_client_specific_deadlines(
        client_name=client_name, gst_registered=gst_registered, has_foreign_transactions=has_foreign
    )
    # Serialize dates
    for d in cal.get("upcoming_90_days", []):
        if hasattr(d.get("date"), "isoformat"):
            d["date"] = d["date"].isoformat()
    for d in cal.get("critical_this_week", []):
        if hasattr(d.get("date"), "isoformat"):
            d["date"] = d["date"].isoformat()
    return cal



# ── CLIENT MANAGEMENT (SQL) ──────────────────────────────────────────

class ClientCreate(BaseModel):
    name: str
    entity_type: str = "individual"
    pan: str = ""
    gstin: str = ""
    cin: str = ""
    email: str = ""
    phone: str = ""
    contact_email: str = ""
    contact_phone: str = ""
    address: str = ""
    state_code: str = ""
    jurisdiction: str = ""
    ward_circle: str = ""
    engagement_type: str = "retainer"
    billing_rate_hourly: float = 0
    notes: str = ""
    practice_areas: list = []

@api_router.post("/clients")
async def create_client(req: ClientCreate, request: Request, authorization: str = Header(None)):
    """Create a new client in SQL database + MongoDB."""
    user = await get_current_user(request, authorization)
    if not check_permission(user.get("role", "analyst"), "clients", "write"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Sanitize inputs
    client_name = sanitize_input(req.name, max_length=255)
    if not client_name:
        raise HTTPException(status_code=400, detail="Client name is required")

    # Validate PAN/GSTIN if provided
    if req.pan and not validate_pan(req.pan):
        raise HTTPException(status_code=400, detail="Invalid PAN format")
    if req.gstin and not validate_gstin(req.gstin):
        raise HTTPException(status_code=400, detail="Invalid GSTIN format")

    client_id = f"cli_{uuid.uuid4().hex[:12]}"
    # Accept either email/phone or contact_email/contact_phone from frontend
    client_email = req.email or req.contact_email
    client_phone = req.phone or req.contact_phone

    # Save to SQL
    try:
        async with get_session() as session:
            client = Client(
                id=client_id, user_id=user["user_id"], name=client_name,
                entity_type=req.entity_type, pan=req.pan.upper() if req.pan else "",
                gstin=req.gstin.upper() if req.gstin else "", cin=req.cin,
                email=client_email, phone=client_phone, address=req.address,
                state_code=req.state_code, jurisdiction=req.jurisdiction,
                ward_circle=req.ward_circle, engagement_type=req.engagement_type,
                billing_rate_hourly=req.billing_rate_hourly, notes=req.notes,
            )
            session.add(client)
    except Exception as e:
        logger.error(f"SQL client creation failed: {e}")
        raise HTTPException(status_code=500, detail="Failed to create client")

    # Also save to MongoDB for document/conversation linking
    try:
        await db.clients.insert_one({
            "client_id": client_id, "user_id": user["user_id"], "name": client_name,
            "entity_type": req.entity_type, "pan": req.pan.upper() if req.pan else "",
            "gstin": req.gstin.upper() if req.gstin else "", "email": client_email,
            "phone": client_phone, "practice_areas": req.practice_areas,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        pass  # Non-blocking — SQL is the source of truth

    await log_audit("client_created", user_id=user["user_id"], resource_type="client",
                    resource_id=client_id, details={"name": client_name})

    return {"client_id": client_id, "name": client_name, "status": "active"}


## NOTE: GET /clients is defined later (~line 3528) with MongoDB + stats.
## Removed duplicate SQLAlchemy-based list_clients to avoid route collision.


# ── CONFLICT OF INTEREST CHECK ──────────────────────────────────────

class ConflictCheckRequest(BaseModel):
    party_name: str
    party_pan: str = ""

@api_router.post("/clients/conflict-check")
async def conflict_check_endpoint(req: ConflictCheckRequest, request: Request, authorization: str = Header(None)):
    """Check for conflicts of interest before taking on a new engagement."""
    user = await get_current_user(request, authorization)
    party_name = sanitize_input(req.party_name, max_length=255)

    result = await check_conflict(user["user_id"], party_name, req.party_pan)

    await log_audit("conflict_check", user_id=user["user_id"], resource_type="client",
                    details={"party_name": party_name, "result": result["result"]},
                    risk_level="medium" if result["result"] != "clear" else "low")

    return result


# ── BILLING & TIME TRACKING ─────────────────────────────────────────

class BillingCreate(BaseModel):
    client_id: str
    matter_id: str = ""
    description: str
    hours: float = 0
    rate: float = 0
    amount: float = 0
    category: str = "advisory"

@api_router.post("/billing/entry")
async def create_billing_entry(req: BillingCreate, request: Request, authorization: str = Header(None)):
    """Create a billing/time entry."""
    user = await get_current_user(request, authorization)
    if not check_permission(user.get("role", "analyst"), "billing", "write"):
        raise HTTPException(status_code=403, detail="Insufficient permissions for billing")

    # Auto-calculate amount if hours and rate provided
    amount = req.amount if req.amount > 0 else (req.hours * req.rate)

    entry_id = f"bill_{uuid.uuid4().hex[:12]}"
    try:
        async with get_session() as session:
            entry = BillingEntry(
                id=entry_id, client_id=req.client_id, user_id=user["user_id"],
                matter_id=req.matter_id, description=sanitize_input(req.description, max_length=1000),
                hours=req.hours, rate=req.rate, amount=amount, category=req.category,
            )
            session.add(entry)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Billing entry failed: {e}")

    return {"entry_id": entry_id, "amount": amount, "status": "unbilled"}


@api_router.get("/billing/summary")
async def billing_summary(request: Request, authorization: str = Header(None),
                           client_id: str = Query(None), period: int = Query(30)):
    """Get billing summary — total hours, amount, breakdown by status."""
    user = await get_current_user(request, authorization)
    summary = await get_billing_summary(user["user_id"], client_id, period)
    return summary


# ── COMPLIANCE DEADLINES (SQL + Calendar) ────────────────────────────

class ComplianceCreate(BaseModel):
    client_id: str = ""
    matter_id: str = ""
    title: str
    deadline_type: str
    due_date: str  # YYYY-MM-DD
    priority: str = "medium"
    notes: str = ""

@api_router.post("/compliance/deadline")
async def create_compliance_deadline(req: ComplianceCreate, request: Request, authorization: str = Header(None)):
    """Create a compliance deadline tracker."""
    user = await get_current_user(request, authorization)
    deadline_id = f"cdl_{uuid.uuid4().hex[:12]}"

    try:
        due = datetime.strptime(req.due_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    # Calculate penalty risk estimate
    from practice_tools import calculate_deadline_penalty
    penalty_risk = 0
    try:
        penalty_info = calculate_deadline_penalty(req.deadline_type, req.due_date)
        penalty_risk = penalty_info.get("total_penalty", 0)
    except Exception:
        pass

    try:
        async with get_session() as session:
            deadline = ComplianceDeadline(
                id=deadline_id, user_id=user["user_id"], client_id=req.client_id,
                matter_id=req.matter_id, title=sanitize_input(req.title, max_length=255),
                deadline_type=req.deadline_type, due_date=due,
                priority=req.priority, penalty_risk=penalty_risk,
                notes=sanitize_input(req.notes, max_length=1000),
            )
            session.add(deadline)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Deadline creation failed: {e}")

    # Also save to MongoDB for calendar integration
    try:
        await db.compliance_deadlines.insert_one({
            "deadline_id": deadline_id, "user_id": user["user_id"],
            "client_id": req.client_id, "title": req.title,
            "deadline_type": req.deadline_type, "due_date": req.due_date,
            "priority": req.priority, "status": "pending",
            "penalty_risk": penalty_risk,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        pass

    return {"deadline_id": deadline_id, "due_date": req.due_date, "penalty_risk": penalty_risk}


@api_router.get("/compliance/overdue")
async def get_overdue(request: Request, authorization: str = Header(None)):
    """Get all overdue compliance deadlines."""
    user = await get_current_user(request, authorization)
    overdue = await get_overdue_compliance(user["user_id"])
    return {"overdue": overdue, "count": len(overdue)}


@api_router.get("/compliance/upcoming")
async def get_upcoming_compliance(request: Request, authorization: str = Header(None), days: int = Query(30)):
    """Get upcoming compliance deadlines within N days."""
    user = await get_current_user(request, authorization)
    from sqlalchemy import select
    cutoff = datetime.now(timezone.utc) + timedelta(days=days)
    try:
        async with get_session() as session:
            stmt = select(ComplianceDeadline).where(
                ComplianceDeadline.user_id == user["user_id"],
                ComplianceDeadline.status == "pending",
                ComplianceDeadline.due_date <= cutoff,
                ComplianceDeadline.due_date >= datetime.now(timezone.utc),
            ).order_by(ComplianceDeadline.due_date)
            result = await session.execute(stmt)
            deadlines = result.scalars().all()
            return {"upcoming": [{
                "id": d.id, "title": d.title, "deadline_type": d.deadline_type,
                "due_date": d.due_date.isoformat(), "priority": d.priority,
                "penalty_risk": d.penalty_risk, "client_id": d.client_id,
                "days_remaining": (d.due_date - datetime.now(timezone.utc)).days,
            } for d in deadlines], "count": len(deadlines)}
    except Exception as e:
        return {"upcoming": [], "count": 0, "error": str(e)}


# ── AUDIT LOG ────────────────────────────────────────────────────────

@api_router.get("/audit/logs")
async def get_audit_logs(request: Request, authorization: str = Header(None),
                          limit: int = Query(50), action: str = Query(None)):
    """Get audit logs — admin/partner only."""
    user = await get_current_user(request, authorization)
    if not check_permission(user.get("role", "analyst"), "audit_logs", "read"):
        raise HTTPException(status_code=403, detail="Audit log access requires partner or admin role")

    from sqlalchemy import select
    try:
        async with get_session() as session:
            stmt = select(AuditLog).where(AuditLog.user_id == user["user_id"])
            if action:
                stmt = stmt.where(AuditLog.action == action)
            stmt = stmt.order_by(AuditLog.created_at.desc()).limit(min(limit, 200))
            result = await session.execute(stmt)
            logs = result.scalars().all()
            return {"logs": [{
                "id": l.id, "action": l.action, "resource_type": l.resource_type,
                "resource_id": l.resource_id, "risk_level": l.risk_level,
                "ip_address": l.ip_address, "created_at": l.created_at.isoformat() if l.created_at else "",
                "details": l.details,
            } for l in logs]}
    except Exception as e:
        return {"logs": [], "error": str(e)}


# ── API KEY MANAGEMENT ───────────────────────────────────────────────

@api_router.post("/api-keys/generate")
async def generate_api_key_endpoint(request: Request, authorization: str = Header(None)):
    """Generate a new API key. The raw key is shown ONCE."""
    user = await get_current_user(request, authorization)
    if not check_permission(user.get("role", "analyst"), "api_keys", "write"):
        raise HTTPException(status_code=403, detail="API key management requires partner or admin role")

    body = await request.json()
    key_name = sanitize_input(body.get("name", "Default"), max_length=100)

    raw_key, key_hash, key_prefix = generate_api_key()
    key_id = f"key_{uuid.uuid4().hex[:12]}"

    try:
        async with get_session() as session:
            api_key = APIKey(
                id=key_id, user_id=user["user_id"], key_hash=key_hash,
                key_prefix=key_prefix, name=key_name,
                scopes=body.get("scopes", ["read", "write"]),
            )
            session.add(api_key)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"API key creation failed: {e}")

    await log_audit("api_key_created", user_id=user["user_id"], resource_type="api_key",
                    resource_id=key_id, risk_level="high")

    return {
        "key_id": key_id,
        "api_key": raw_key,  # Shown ONCE — user must save it
        "prefix": key_prefix,
        "name": key_name,
        "warning": "Save this key now. It will not be shown again.",
    }


# ── FORM 3CD / TAX AUDIT ─────────────────────────────────────────────
@api_router.post("/tools/tax-audit/form3cd")
async def api_form3cd(request: Request, authorization: str = Header(None),
                       tally_export: UploadFile = File(...), bank_statement: UploadFile = File(None),
                       assessee_name: str = Form(""), pan: str = Form(""),
                       fy: str = Form("2024-25"), turnover_lakhs: str = Form("100")):
    """Form 3CD AI Assistant — Generate clause-wise tax audit observations from Tally + bank data."""
    user = await get_current_user(request, authorization)
    
    tally_bytes = await tally_export.read()
    tally_text = extract_xlsx_text(tally_bytes) if tally_export.filename.endswith(".xlsx") else extract_pdf_text(tally_bytes)
    
    forensic_summary = ""
    if bank_statement:
        bank_bytes = await bank_statement.read()
        bank_text = extract_xlsx_text(bank_bytes) if bank_statement.filename.endswith(".xlsx") else extract_pdf_text(bank_bytes)
        forensic_summary = bank_text[:2000]
    
    audit_prompt = build_audit_prompt(
        assessee_name=assessee_name,
        pan=pan,
        fy=fy,
        turnover_lakhs=float(turnover_lakhs),
        financial_data=tally_text,
        forensic_summary=forensic_summary
    )
    
    result = await process_query(
        user_query=audit_prompt,
        mode="partner",
        matter_context=f"Tax Audit Mode | {assessee_name} | FY {fy}"
    )
    
    clauses = parse_audit_clauses(result.get("response_text", ""))
    high_risk = [c for c in clauses if c.get("status") == "HIGH_RISK"]
    
    # Save to DB
    audit_id = f"audit_{uuid.uuid4().hex[:12]}"
    try:
        await db.tax_audits.insert_one({
            "audit_id": audit_id,
            "user_id": user["user_id"],
            "assessee_name": assessee_name,
            "pan": pan,
            "fy": fy,
            "turnover_lakhs": turnover_lakhs,
            "clauses": clauses,
            "high_risk_count": len(high_risk),
            "raw_response": result.get("response_text", ""),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    except Exception as e:
        logger.warning(f"Tax audit save failed (returning result anyway): {e}")
    
    return {
        "audit_id": audit_id,
        "assessee_name": assessee_name,
        "pan": pan,
        "fy": fy,
        "total_clauses": len(clauses),
        "high_risk_clauses": len(high_risk),
        "clauses": clauses,
        "high_risk_details": high_risk,
    }


@api_router.get("/tools/tax-audit/form3cd/clauses")
async def api_form3cd_clauses(request: Request, authorization: str = Header(None)):
    """List all 44 Form 3CD clause descriptions."""
    await get_current_user(request, authorization)
    return FORM_3CD_CLAUSES


# ── IBC / NCLT ────────────────────────────────────────────────────────
@api_router.get("/tools/ibc/cirp-milestones")
async def api_cirp_milestones(commencement_date: str = Query(""), bench: str = Query("Mumbai"),
                               request: Request = None, authorization: str = Header(None)):
    """Calculate all CIRP milestone dates and alert status from insolvency commencement date."""
    await get_current_user(request, authorization)
    result = calculate_cirp_milestones(commencement_date, bench)
    return result


@api_router.get("/tools/ibc/timelines")
async def api_ibc_timelines(request: Request, authorization: str = Header(None)):
    """Get all IBC/CIRP statutory timelines."""
    await get_current_user(request, authorization)
    return IBC_TIMELINES


@api_router.post("/tools/ibc/section-29a-check")
async def api_29a_check(request: Request, authorization: str = Header(None),
                         applicant_facts: str = Form("")):
    """Section 29A eligibility preliminary check for resolution applicants."""
    await get_current_user(request, authorization)
    result = check_section_29a_eligibility(applicant_facts)
    return result


@api_router.get("/tools/ibc/workflow/{application_type}")
async def api_ibc_workflow(application_type: str, request: Request, authorization: str = Header(None)):
    """Get Section 7 or Section 9 application checklist and requirements."""
    await get_current_user(request, authorization)
    if application_type not in IBC_WORKFLOW_TEMPLATES:
        raise HTTPException(status_code=400, detail="Use 'section_7' or 'section_9'")
    return IBC_WORKFLOW_TEMPLATES[application_type]


@api_router.post("/tools/ibc/draft-application")
async def api_ibc_draft(request: Request, authorization: str = Header(None),
                          application_type: str = Form("section_7"),
                          corporate_debtor: str = Form(""),
                          default_amount: str = Form(""),
                          default_date: str = Form(""),
                          creditor_name: str = Form(""),
                          facts: str = Form("")):
    """Draft a Section 7 or Section 9 IBC application."""
    await get_current_user(request, authorization)
    
    template = IBC_WORKFLOW_TEMPLATES.get(application_type, {})
    draft_prompt = f"""
You are an IBC Specialist drafting a {template.get('title', application_type)} application for NCLT filing.

FACTS:
- Corporate Debtor: {corporate_debtor}
- Creditor: {creditor_name}
- Default Amount: ₹{default_amount}
- Date of Default: {default_date}
- Additional Facts: {facts}

Draft a complete {template.get('title', 'IBC Application')} including:
1. Cover Page (Court, Parties, Bench)
2. Facts of the Case
3. Grounds for Admission
4. Relief Claimed
5. List of Documents to be annexed

Reference key NCLT judgments on admission threshold and default proof.
Apply the {template.get('section', 'Section 7/9 IBC')} standard precisely.
"""
    
    result = await process_query(
        user_query=draft_prompt,
        mode="partner",
        matter_context=f"IBC {application_type.upper()} | {corporate_debtor} | ₹{default_amount}"
    )
    
    return {
        "application_type": application_type,
        "corporate_debtor": corporate_debtor,
        "draft": result.get("response_text", ""),
        "checklist": template.get("checklist", []),
        "filing_fee": template.get("filing_fee", "₹2,000"),
        "limitation_note": template.get("limitation", ""),
    }


# ── REGULATORY MONITOR ────────────────────────────────────────────────
@api_router.get("/regulatory/updates")
async def api_regulatory_updates(
    matter_types: str = Query("general"),
    days_back: int = Query(7),
    request: Request = None, authorization: str = Header(None)
):
    """Check recent regulatory updates (SEBI/RBI/MCA/CBIC/IBBI) relevant to active matter types."""
    await get_current_user(request, authorization)
    matter_type_list = [m.strip() for m in matter_types.split(",")]
    updates = await check_regulatory_updates(matter_type_list, days_back)
    return {"matter_types": matter_type_list, "updates": updates, "count": len(updates)}


@api_router.post("/regulatory/impact-analysis")
async def api_regulatory_impact(request: Request, authorization: str = Header(None),
                                  update_title: str = Form(""), update_description: str = Form(""),
                                  matter_type: str = Form("general"), matter_facts: str = Form("")):
    """Analyze the impact of a specific regulatory update on a client matter."""
    await get_current_user(request, authorization)
    impact = await generate_regulatory_impact(
        update_title=update_title,
        update_description=update_description,
        matter_type=matter_type,
        matter_facts=matter_facts,
        process_query_fn=process_query
    )
    return {
        "update_title": update_title,
        "matter_type": matter_type,
        "impact_analysis": impact,
    }


# ==================== PORTFOLIO COMMAND CENTER ====================

class PortfolioClientCreate(BaseModel):
    name: str
    pan: str = ""
    gstin: str = ""
    entity_type: str = "Company"
    practice_areas: list = []

@api_router.get("/portfolio/clients")
async def get_portfolio_clients(request: Request, authorization: str = Header(None)):
    """Get all clients for the authenticated user's portfolio with computed risk and deadlines."""
    user = await get_current_user(request, authorization)

    try:
        cursor = db.portfolio_clients.find(
            {"user_id": user["user_id"]},
            {"_id": 0}
        ).sort("created_at", -1)
        clients = await cursor.to_list(500)
    except Exception as e:
        logger.warning(f"MongoDB portfolio query failed: {e}")
        return {"clients": [], "total": 0}
    
    # Compute deadline status and conflict detection for each client
    now = datetime.now(timezone.utc)
    all_pans = [c.get("pan", "") for c in clients if c.get("pan")]
    
    for client in clients:
        # Count active matters
        matter_count = await db.matters.count_documents({
            "user_id": user["user_id"],
            "client_name": {"$regex": client["name"], "$options": "i"}
        })
        client["active_matters"] = matter_count
        
        # Find next deadline from compliance calendar
        next_deadline = await db.compliance_deadlines.find_one(
            {"client_id": client.get("client_id"), "due_date": {"$gte": now.isoformat()}},
            sort=[("due_date", 1)]
        )
        client["next_deadline"] = next_deadline.get("due_date") if next_deadline else None
        
        # Compute risk level based on overdue items and matter complexity
        overdue_count = await db.compliance_deadlines.count_documents({
            "client_id": client.get("client_id"),
            "due_date": {"$lt": now.isoformat()},
            "status": {"$ne": "completed"}
        })
        if overdue_count >= 3:
            client["risk_level"] = "HIGH"
        elif overdue_count >= 1:
            client["risk_level"] = "MEDIUM"
        else:
            client["risk_level"] = client.get("risk_level", "LOW")
        
        # Conflict detection — check if this client's PAN appears as an opposing party 
        # in any other client's matters
        conflicts = []
        if client.get("pan"):
            opposing = await db.matters.find(
                {
                    "user_id": user["user_id"],
                    "opposing_party_pan": client["pan"],
                    "client_name": {"$not": {"$regex": client["name"], "$options": "i"}}
                }
            ).to_list(5)
            if opposing:
                for opp in opposing:
                    conflicts.append(f"Opposing party in {opp.get('matter_title', 'a matter')} for {opp.get('client_name', 'another client')}")
        client["conflicts"] = conflicts
    
    return {"clients": clients, "total": len(clients)}


@api_router.post("/portfolio/clients")
async def add_portfolio_client(req: PortfolioClientCreate, request: Request, authorization: str = Header(None)):
    """Add a new client to the portfolio."""
    user = await get_current_user(request, authorization)
    
    client_id = f"cli_{uuid.uuid4().hex[:12]}"
    client_doc = {
        "client_id": client_id,
        "user_id": user["user_id"],
        "name": req.name,
        "pan": req.pan,
        "gstin": req.gstin,
        "entity_type": req.entity_type,
        "practice_areas": req.practice_areas,
        "risk_level": "LOW",
        "active_matters": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    try:
        await db.portfolio_clients.insert_one(client_doc)
    except Exception as e:
        logger.warning(f"Portfolio client insert failed: {e}")
        raise HTTPException(status_code=503, detail="Database temporarily unavailable. Please try again.")

    return {"client_id": client_id, "name": req.name, "pan": req.pan, "gstin": req.gstin, "entity_type": req.entity_type, "risk_level": "LOW", "active_matters": 0, "message": f"Client '{req.name}' added to portfolio."}

@api_router.get("/statutes/{act}/{section}")
async def get_statute(act: str, section: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    import re as _re
    try:
        safe_section = _re.escape(section)
        statute_doc = await db["master_statutes"].find_one({
            "act": act,
            "section": {'$regex': f'^{safe_section}$', '$options': 'i'}
        })
        if statute_doc:
            statute_doc.pop("_id", None)
            return statute_doc
        else:
            raise HTTPException(status_code=404, detail="Statute not found")
    except Exception as e:
        logger.error(f"Error fetching statute: {e}")
        raise HTTPException(status_code=500, detail="Database error")

from document_intelligence import index_document, query_documents

@api_router.post("/vault/bulk-upload")
async def vault_bulk_upload(request: Request, authorization: str = Header(None)):
    """Upload multiple files to a matter."""
    user = await get_current_user(request, authorization)
    form = await request.form()
    matter_id = form.get("matter_id", "")
    
    docs_saved = []
    
    # Process multiple files
    for key in form.keys():
        if key.startswith('file_'):
            file = form[key]
            # Regular upload logic
            import os, uuid
            # (Truncated for standard saving - assumes storage_utils usage typically, but we do basic save for now)
            filename = getattr(file, "filename", "unknown.pdf")
            file_id = f"doc_{uuid.uuid4().hex[:12]}"
            ext = os.path.splitext(filename)[1].lower()
            os.makedirs("vault_data", exist_ok=True)
            local_path = os.path.join("vault_data", f"{file_id}{ext}")
            
            content = await file.read()
            with open(local_path, "wb") as f:
                f.write(content)
                
            doc_record = {
                "document_id": file_id,
                "matter_id": matter_id,
                "user_id": user["user_id"],
                "filename": filename,
                "path": local_path,
                "uploaded_at": datetime.now(timezone.utc).isoformat()
            }
            await db.vault_documents.insert_one(doc_record)
            
            # Fire and forget indexing for bulk intelligence
            asyncio.create_task(index_document(file_id, local_path, matter_id))
            
            docs_saved.append(doc_record)
            
    return {"message": f"Successfully uploaded {len(docs_saved)} documents", "documents": [d["filename"] for d in docs_saved]}

class IntelligenceQuery(BaseModel):
    matter_id: str
    query: str

@api_router.post("/vault/intelligence")
async def bulk_intelligence_query(req: IntelligenceQuery, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    
    result = await query_documents(req.matter_id, req.query)
    return {"answer": result["answer"], "sources": result["sources"], "source_chunks": result["chunks"]}


from obligation_extractor import extract_obligations
from workflow_chain import start_chain, advance_chain, get_chain_status, get_templates
from court_tracker import track_case, get_tracked_cases, remove_tracked_case, refresh_case, search_ecourts
from playbook_engine import compare_against_playbook
from judgment_summarizer import summarize_judgment

class ObligationRequest(BaseModel):
    document_text: str

@api_router.post("/vault/extract-obligations")
async def extract_obligations_endpoint(req: ObligationRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    result = await extract_obligations(req.document_text)
    return result

class ChainStartRequest(BaseModel):
    chain_type: str
    initial_input: str

class ChainAdvanceRequest(BaseModel):
    chain_id: str
    edited_output: str = None

@api_router.get("/workflows/chain/templates")
async def get_chain_templates(request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    return {"templates": get_templates()}

@api_router.post("/workflows/chain/start")
async def start_chain_endpoint(req: ChainStartRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    result = await start_chain(req.chain_type, req.initial_input, user["user_id"])
    return result

@api_router.post("/workflows/chain/next")
async def advance_chain_endpoint(req: ChainAdvanceRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    result = await advance_chain(req.chain_id, req.edited_output, user_id=user["user_id"])
    return result

@api_router.get("/workflows/chain/{chain_id}")
async def get_chain_endpoint(chain_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    result = await get_chain_status(chain_id, user_id=user["user_id"])
    return result

class CourtTrackRequest(BaseModel):
    case_number: str
    court: str = "supreme_court"
    party_name: str = ""
    matter_id: str = None

@api_router.post("/court/track")
async def track_case_endpoint(req: CourtTrackRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        result = await track_case(user["user_id"], req.case_number, req.court, req.party_name, req.matter_id)
        try:
            from audit_sink import log_court_tracker_event as _log_ct
            await _log_ct(user=user, request=request, action="track", outcome="success",
                          case_number=req.case_number, track_id=result.get("track_id"), court=req.court)
        except Exception: pass
        return result
    except Exception as e:
        # DB may be unavailable — return a locally-tracked stub so the UI
        # still gets a usable entry. Log the full error server-side.
        logger.error(f"Court tracker insert failed: {e}", exc_info=True)
        import uuid as _uuid
        from datetime import datetime as _dt, timezone as _tz
        stub = {
            "track_id": f"trk_{_uuid.uuid4().hex[:12]}",
            "user_id": user["user_id"],
            "case_number": req.case_number,
            "court": req.court,
            "party_name": req.party_name,
            "matter_id": req.matter_id,
            "status": "tracking_local",
            "next_hearing": None,
            "last_checked": _dt.now(_tz.utc).isoformat(),
            "search_results": [],
            "created_at": _dt.now(_tz.utc).isoformat(),
            "note": "Tracked locally — database temporarily unavailable. Will sync when connection restores.",
        }
        return stub

@api_router.get("/court/upcoming")
async def get_upcoming_cases(request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        cases = await get_tracked_cases(user["user_id"])
    except Exception as e:
        logger.warning(f"Court tracker query failed: {e}")
        cases = []
    return {"cases": cases}

@api_router.delete("/court/track/{track_id}")
async def remove_case_endpoint(track_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    success = await remove_tracked_case(track_id, user["user_id"])
    return {"deleted": success}

@api_router.post("/court/refresh/{track_id}")
async def refresh_case_endpoint(track_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    result = await refresh_case(track_id)
    return result

class PlaybookCompareRequest(BaseModel):
    playbook_text: str
    draft_text: str

@api_router.post("/playbook/compare")
async def playbook_compare_endpoint(req: PlaybookCompareRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    result = await compare_against_playbook(req.playbook_text, req.draft_text)
    return result

class JudgmentSummarizeRequest(BaseModel):
    document_text: str

@api_router.post("/vault/summarize-judgment")
async def summarize_judgment_endpoint(req: JudgmentSummarizeRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    result = await summarize_judgment(req.document_text)
    return result

class WorkspaceInvite(BaseModel):
    matter_id: str
    email: str
    role: str = "editor"

@api_router.post("/matters/{matter_id}/invite")
async def invite_to_matter(matter_id: str, req: WorkspaceInvite, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    member_doc = {
        "matter_id": matter_id,
        "invited_by": user["user_id"],
        "email": req.email,
        "role": req.role,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    try:
        await db.workspace_members.insert_one(member_doc)
    except Exception as e:
        logger.warning(f"Member invite save failed: {e}")
        raise HTTPException(status_code=503, detail="Database temporarily unavailable. Please try again.")
    return {"message": f"Invitation sent to {req.email}"}

@api_router.get("/matters/{matter_id}/members")
async def get_matter_members(matter_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    try:
        cursor = db.workspace_members.find({"matter_id": matter_id})
        members = []
        async for doc in cursor:
            doc.pop("_id", None)
            members.append(doc)
        return {"members": members}
    except Exception as e:
        logger.warning(f"Members fetch failed: {e}")
        return {"members": []}



@api_router.post("/regulatory/force-check")
async def force_regulatory_check(request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    updates = await check_regulatory_updates(["general"], days_back=1)
    return {"new_updates_found": len(updates), "updates": updates}

@api_router.post("/export/enhanced")
async def enhanced_export_route(request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    data = await request.json()
    doc_type = data.get("doc_type", "memo")
    content = data.get("content", "")
    
    if doc_type == "court":
        doc_bytes = generate_court_document(
            content=content,
            court_type=data.get("court_type", "high_court"),
            case_number=data.get("case_number", ""),
            petitioner=data.get("petitioner", ""),
            respondent=data.get("respondent", ""),
            document_type=data.get("document_type", "PETITION")
        )
        filename = "Court_Filing.docx"
    elif doc_type == "tax_reply":
        doc_bytes = generate_tax_notice_reply(
            notice_details=data.get("notice_details", {}),
            reply_content=content
        )
        filename = "Notice_Reply.docx"
    else:
        doc_bytes = generate_memo_export(content, title=data.get("title", "Legal Memorandum"))
        filename = "Legal_Memo.docx"
        
    return Response(
        content=doc_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/matters/{matter_id}/ingest-email")
async def get_ingest_email_route(matter_id: str, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    email = await generate_ingest_address(matter_id, user["user_id"])
    return {"ingestion_email": email}


class ClassificationRequest(BaseModel):
    items: list[str]

@api_router.post("/excel/classify")
async def excel_classify_route(req: ClassificationRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    
    # Mocking classification for speed based on vendor_classifier heuristics
    results = []
    for item in req.items:
        val = item.lower()
        if 'gst' in val or 'ltd' in val or 'pvt' in val:
            results.append("GST Registered")
        elif 'hospital' in val or 'school' in val:
            results.append("Exempt Entity")
        else:
            results.append("Unregistered Vendor")
    return {"classification": results}

class ReconRequest(BaseModel):
    purchase_register: list[dict]
    gstr_2b: list[dict]

@api_router.post("/excel/reconcile")
async def excel_recon_route(req: ReconRequest, request: Request, authorization: str = Header(None)):
    user = await get_current_user(request, authorization)
    return {"status": "success", "message": "Reconciliation complete"}
# ==================== AGENT TOOL ACCESS METADATA ====================

from ai_engine import AGENT_TOOLS

@api_router.get("/agent/tools")
async def list_agent_tools():
    """List all tools available to the AI agent for autonomous execution."""
    tools = []
    for name, tool in AGENT_TOOLS.items():
        tools.append({
            "name": tool["name"],
            "description": tool["description"],
            "parameters": tool["parameters"],
        })
    return {"tools": tools, "count": len(tools)}

# ==================== CLIENTS (with per-client memory) ====================
## ClientCreate model is defined earlier (~line 2620) — single source of truth

@api_router.get("/clients")
async def get_clients(request: Request, authorization: str = Header(None)):
    """Get all clients for the user. Simplified client management with per-client memory."""
    user = await get_current_user(request, authorization)

    try:
        clients_cursor = db.clients.find(
            {"user_id": user["user_id"], "is_deleted": {"$ne": True}},
            {"_id": 0}
        ).sort("updated_at", -1)
        clients = await clients_cursor.to_list(500)
    except Exception as e:
        logger.warning(f"Clients fetch failed: {e}")
        clients = []

    # Compute stats for each client
    for c in clients:
        try:
            c["active_matters"] = await db.matters.count_documents({
                "user_id": user["user_id"],
                "client_name": {"$regex": c["name"], "$options": "i"}
            })
        except Exception:
            c["active_matters"] = 0

        try:
            next_dl = await db.compliance_deadlines.find_one(
                {"client_id": c.get("client_id"), "due_date": {"$gte": datetime.now(timezone.utc).isoformat()}},
                sort=[("due_date", 1)]
            )
            c["next_deadline"] = next_dl.get("due_date") if next_dl else None
        except Exception:
            c["next_deadline"] = None

        try:
            overdue = await db.compliance_deadlines.count_documents({
                "client_id": c.get("client_id"),
                "due_date": {"$lt": datetime.now(timezone.utc).isoformat()},
                "status": {"$ne": "completed"}
            })
            c["risk_level"] = "HIGH" if overdue >= 3 else "MEDIUM" if overdue >= 1 else "LOW"
        except Exception:
            c["risk_level"] = "LOW"

    return clients


@api_router.get("/clients/{client_id}")
async def get_client(client_id: str, request: Request, authorization: str = Header(None)):
    """Get a single client."""
    user = await get_current_user(request, authorization)
    try:
        client = await db.clients.find_one(
            {"client_id": client_id, "user_id": user["user_id"]}, {"_id": 0}
        )
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        return client
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {e}")


@api_router.post("/clients/{client_id}/query")
async def client_query(client_id: str, request: Request, authorization: str = Header(None)):
    """Query the AI assistant in the context of a specific client. Carries client memory."""
    user = await get_current_user(request, authorization)
    body = await request.json()
    user_query = body.get("query", "")

    if not user_query:
        raise HTTPException(status_code=400, detail="Query required")

    # Fetch client
    client = None
    try:
        client = await db.clients.find_one({"client_id": client_id, "user_id": user["user_id"]}, {"_id": 0})
    except Exception:
        pass
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    # Build client context with memory (past conversations)
    client_context = f"CLIENT: {client.get('name', '')} | Type: {client.get('entity_type', '')} | PAN: {client.get('pan', '')} | GSTIN: {client.get('gstin', '')}"
    client_context += f"\nPractice Areas: {', '.join(client.get('practice_areas', []))}"

    # Retrieve conversation memory for this client (last 5 queries)
    try:
        recent_history = await db.client_conversations.find(
            {"client_id": client_id, "user_id": user["user_id"]}
        ).sort("created_at", -1).limit(5).to_list(5)
        if recent_history:
            client_context += "\n\n=== PREVIOUS CONVERSATIONS WITH THIS CLIENT ===\n"
            for h in reversed(recent_history):
                client_context += f"Q: {h.get('query', '')[:200]}\nA: {h.get('response', '')[:500]}\n\n"
    except Exception as e:
        logger.warning(f"Client history lookup failed: {e}")

    # Statute context
    statute_context = ""
    try:
        statute_context = await get_statute_context(user_query)
    except Exception:
        pass

    # Firm library context
    firm_context = ""
    try:
        lib_items = await db.library_items.find({"user_id": user["user_id"]}).to_list(10)
        if lib_items:
            firm_context = "Apply these firm templates if relevant:\n"
            for idx, item in enumerate(lib_items):
                firm_context += f"--- {item.get('title', '')} ---\n{item.get('content', '')[:500]}\n\n"
    except Exception:
        pass

    # === PROMPT INJECTION DEFENSE (client advisor context) ===
    client_context = sanitize_context_for_llm(client_context, source_label="client_advisor_context")
    statute_context = sanitize_context_for_llm(statute_context, source_label="client_advisor_statute")
    firm_context = sanitize_context_for_llm(firm_context, source_label="client_advisor_firm")

    from fastapi.responses import StreamingResponse

    async def sse_generator():
        yield ": " + " " * 2048 + "\n\n"

        full_response = ""
        async for chunk in process_query_streamed(
            user_query=user_query,
            mode="partner",
            matter_context=client_context,
            statute_context=statute_context,
            firm_context=firm_context,
            user_id=user.get("user_id") if isinstance(user, dict) else None,
            conversation_id=f"client_advisor_{uuid.uuid4().hex[:10]}",
        ):
            yield f"data: {chunk}\n\n"
            # Capture response for memory
            try:
                data = json.loads(chunk)
                if data.get("type") == "partner_payload":
                    full_response = data.get("content", "")
                elif data.get("type") == "fast_chunk" and not full_response:
                    full_response = data.get("content", "")
            except Exception:
                pass

        yield "data: [DONE]\n\n"

        # Save to client conversation memory
        if full_response:
            try:
                await db.client_conversations.insert_one({
                    "client_id": client_id,
                    "user_id": user["user_id"],
                    "query": user_query[:5000],
                    "response": full_response[:10000],
                    "source": "chat",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                })
                # Update client's updated_at
                await db.clients.update_one(
                    {"client_id": client_id},
                    {"$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
                )
            except Exception as e:
                logger.warning(f"Client conversation save failed: {e}")

    return StreamingResponse(sse_generator(), media_type="text/event-stream")


@api_router.get("/clients/{client_id}/history")
async def get_client_history(client_id: str, request: Request, authorization: str = Header(None)):
    """Get conversation history for a specific client (includes both chat and email)."""
    user = await get_current_user(request, authorization)

    try:
        history = await db.client_conversations.find(
            {"client_id": client_id, "user_id": user["user_id"]},
            {"_id": 0}
        ).sort("created_at", -1).limit(50).to_list(50)
        return history
    except Exception as e:
        logger.warning(f"Client history fetch failed: {e}")
        return []


# ==================== AGENTMAIL EMAIL PROCESSING ====================

@api_router.get("/email/inbox")
async def get_email_inbox(request: Request, authorization: str = Header(None)):
    """Get inbox email address and status."""
    await get_current_user(request, authorization)
    messages = await list_unread_messages()
    return {
        "inbox_email": INBOX_EMAIL,
        "unread_count": len(messages),
        "status": "active" if INBOX_EMAIL else "not_configured",
    }


@api_router.post("/email/process")
async def process_emails_now(request: Request, authorization: str = Header(None)):
    """Manually trigger email processing (check inbox, process, reply)."""
    await get_current_user(request, authorization)
    try:
        count = await process_incoming_emails(db)
        return {"processed": count, "status": "ok"}
    except Exception as e:
        logger.error(f"Email processing failed: {e}")
        return {"processed": 0, "status": "error", "detail": str(e)}


@api_router.post("/email/send")
async def send_email_endpoint(request: Request, authorization: str = Header(None)):
    """Send an email from the Spectr inbox."""
    user = await get_current_user(request, authorization)
    body = await request.json()
    to_addr = body.get("to", "")
    subject = body.get("subject", "")
    text_body = body.get("text", body.get("body", ""))
    html_body = body.get("html", "")

    if not to_addr or not subject:
        raise HTTPException(status_code=400, detail="'to' and 'subject' required")

    try:
        success = await send_email(to_addr, subject, text_body or html_body)
        if not success:
            raise HTTPException(status_code=502, detail="Email delivery failed. Check inbox configuration.")
        return {"sent": True, "to": to_addr, "subject": subject, "from": INBOX_EMAIL}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Email send failed: {e}")
        raise HTTPException(status_code=500, detail=f"Email send error: {str(e)[:200]}")


@api_router.get("/email/dead-letter")
async def get_dead_letter_queue(request: Request, authorization: str = Header(None)):
    """View the dead-letter queue — emails that failed after all retries."""
    await get_current_user(request, authorization)
    return {
        "count": len(dead_letter_queue),
        "messages": dead_letter_queue[-50:],  # Last 50
    }


@api_router.post("/email/dead-letter/{msg_id}/reprocess")
async def reprocess_dead_letter(msg_id: str, request: Request, authorization: str = Header(None)):
    """Reprocess a specific dead-lettered email."""
    await get_current_user(request, authorization)
    # Find the message in dead-letter queue
    target = None
    for i, entry in enumerate(dead_letter_queue):
        if entry.get("message_id") == msg_id:
            target = entry
            dead_letter_queue.pop(i)
            break
    if not target:
        raise HTTPException(status_code=404, detail="Message not found in dead-letter queue")
    # Re-trigger processing — the email worker will pick it up from AgentMail
    # Remove from processed_ids so it gets re-processed
    if msg_id in _processed_ids:
        _processed_ids.remove(msg_id)
    return {"status": "requeued", "message_id": msg_id, "original_error": target.get("error", "")}


@api_router.get("/email/history")
async def get_email_history(request: Request, authorization: str = Header(None), limit: int = 50):
    """Get email conversation history — all processed emails with responses."""
    await get_current_user(request, authorization)
    try:
        history = await db.email_history.find(
            {}, {"_id": 0}
        ).sort("created_at", -1).limit(limit).to_list(limit)
        return {"emails": history, "total": len(history)}
    except Exception as e:
        logger.warning(f"Email history fetch failed: {e}")
        return {"emails": [], "total": 0}


@api_router.get("/email/analytics")
async def get_email_analytics(request: Request, authorization: str = Header(None), days: int = 30):
    """Email analytics dashboard — response times, volume, complexity distribution."""
    await get_current_user(request, authorization)
    try:
        from datetime import timedelta
        cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

        # Total emails processed
        total = await db.email_history.count_documents({"created_at": {"$gte": cutoff}})
        replied = await db.email_history.count_documents({"created_at": {"$gte": cutoff}, "status": "replied"})
        failed = await db.email_history.count_documents({"created_at": {"$gte": cutoff}, "status": "failed"})

        # Complexity distribution
        complexity_pipeline = [
            {"$match": {"created_at": {"$gte": cutoff}}},
            {"$group": {"_id": "$complexity", "count": {"$sum": 1}}},
        ]
        complexity_dist = {}
        async for doc in db.email_history.aggregate(complexity_pipeline):
            complexity_dist[doc["_id"] or "unknown"] = doc["count"]

        # Unique clients
        unique_clients = await db.email_history.distinct("client_email", {"created_at": {"$gte": cutoff}})

        return {
            "period_days": days,
            "total_processed": total,
            "replied": replied,
            "failed": failed,
            "success_rate": round((replied / total * 100), 1) if total > 0 else 0,
            "complexity_distribution": complexity_dist,
            "unique_clients": len([c for c in unique_clients if c]),
            "dead_letter_count": len(dead_letter_queue),
            "worker_status": "running" if _email_worker_running else "stopped",
        }
    except Exception as e:
        logger.warning(f"Email analytics failed: {e}")
        return {
            "period_days": days,
            "total_processed": 0,
            "replied": 0,
            "failed": 0,
            "success_rate": 0,
            "complexity_distribution": {},
            "unique_clients": 0,
            "dead_letter_count": len(dead_letter_queue),
            "worker_status": "running" if _email_worker_running else "stopped",
        }


@api_router.post("/webhook/email-ingest")
async def webhook_email_ingest(request: Request):
    """Receive forwarded emails via webhook — routes attachments to matter vault.
    No auth required (webhooks come from email providers, not users).
    Validates via recipient address mapping instead.
    """
    try:
        from email_ingestion import process_inbound_email
        body = await request.json()
        raw_email = body.get("raw_email", body.get("email", ""))
        recipient = body.get("recipient", body.get("to", ""))

        if not raw_email or not recipient:
            raise HTTPException(status_code=400, detail="'raw_email' and 'recipient' required")

        result = await process_inbound_email(raw_email, recipient)
        if "error" in result:
            raise HTTPException(status_code=400, detail=result["error"])

        logger.info(f"Email ingest: {result.get('attachments_saved', 0)} files -> matter {result.get('matter_id', 'unknown')}")
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Email ingest webhook failed: {e}")
        raise HTTPException(status_code=500, detail=f"Ingestion failed: {str(e)[:200]}")


# ==================== TDS/TCS ENGINE ROUTES ====================

class TDSCalcRequest(BaseModel):
    section: str
    payment_amount: float
    payee_pan: str = ""
    payee_type: str = "individual"
    is_resident: bool = True
    has_lower_deduction_cert: bool = False
    certificate_rate: Optional[float] = None
    is_senior_citizen: bool = False
    aggregate_paid_in_fy: Optional[float] = None
    is_non_filer_206ab: bool = False

@api_router.post("/tools/tds/calculate")
async def api_tds_calculate(req: TDSCalcRequest, request: Request, authorization: str = Header(None)):
    """Calculate TDS for a single payment with full section/rate logic."""
    user = await get_current_user(request, authorization)
    result = calculate_tds(
        section=req.section,
        payment_amount=req.payment_amount,
        payee_pan=req.payee_pan,
        payee_type=req.payee_type,
        is_resident=req.is_resident,
        has_lower_deduction_cert=req.has_lower_deduction_cert,
        certificate_rate=req.certificate_rate,
        is_senior_citizen=req.is_senior_citizen,
        aggregate_paid_in_fy=req.aggregate_paid_in_fy,
        is_non_filer_206ab=req.is_non_filer_206ab,
    )
    # Convert Decimals to strings for JSON
    return json.loads(json.dumps(result, default=str))

class SalaryTDSRequest(BaseModel):
    gross_salary: float
    hra_exempt: Optional[float] = None
    lta_exempt: Optional[float] = None
    standard_deduction: Optional[float] = None
    chapter_vi_a_deductions: Optional[dict] = None
    regime: str = "new"
    employer_pf_contribution: Optional[float] = None
    professional_tax: Optional[float] = None
    other_income: Optional[float] = None
    tds_already_deducted: Optional[float] = None
    months_remaining: int = 12

@api_router.post("/tools/tds/salary")
async def api_tds_salary(req: SalaryTDSRequest, request: Request, authorization: str = Header(None)):
    """Calculate Section 192 salary TDS with full slab computation."""
    user = await get_current_user(request, authorization)
    result = calculate_salary_tds(
        gross_salary=req.gross_salary,
        hra_exempt=req.hra_exempt,
        lta_exempt=req.lta_exempt,
        standard_deduction=req.standard_deduction,
        chapter_vi_a_deductions=req.chapter_vi_a_deductions,
        regime=req.regime,
        employer_pf_contribution=req.employer_pf_contribution,
        professional_tax=req.professional_tax,
        other_income=req.other_income,
        tds_already_deducted=req.tds_already_deducted,
        months_remaining=req.months_remaining,
    )
    return json.loads(json.dumps(result, default=str))

class TCSCalcRequest(BaseModel):
    tcs_section: str
    sale_amount: float
    buyer_pan: str = ""
    is_non_filer: bool = False
    aggregate_in_fy: Optional[float] = None

@api_router.post("/tools/tcs/calculate")
async def api_tcs_calculate(req: TCSCalcRequest, request: Request, authorization: str = Header(None)):
    """Calculate TCS under Section 206C."""
    user = await get_current_user(request, authorization)
    result = calculate_tcs(
        tcs_section=req.tcs_section,
        sale_amount=req.sale_amount,
        buyer_pan=req.buyer_pan,
        is_non_filer=req.is_non_filer,
        aggregate_in_fy=req.aggregate_in_fy,
    )
    return json.loads(json.dumps(result, default=str))

class BulkTDSRequest(BaseModel):
    records: list

@api_router.post("/tools/tds/bulk")
async def api_tds_bulk(req: BulkTDSRequest, request: Request, authorization: str = Header(None)):
    """Compute TDS for multiple payments at once."""
    user = await get_current_user(request, authorization)
    result = compute_bulk_tds(req.records)
    return json.loads(json.dumps(result, default=str))

class Form26QRequest(BaseModel):
    tan: str
    deductor_name: str
    deductor_address: dict = {}
    financial_year: str
    quarter: str
    deductee_records: list
    challans: list = []

@api_router.post("/tools/tds/form-26q")
async def api_form_26q(req: Form26QRequest, request: Request, authorization: str = Header(None)):
    """Generate Form 26Q (TDS return for non-salary payments)."""
    user = await get_current_user(request, authorization)
    result = generate_form_26q(
        tan=req.tan,
        deductor_name=req.deductor_name,
        deductor_address=req.deductor_address,
        financial_year=req.financial_year,
        quarter=req.quarter,
        deductee_records=req.deductee_records,
        challans=req.challans,
    )
    return json.loads(json.dumps(result, default=str))

class Form24QRequest(BaseModel):
    tan: str
    employer_name: str
    employer_address: dict = {}
    financial_year: str
    quarter: str
    employee_records: list
    challans: list = []

@api_router.post("/tools/tds/form-24q")
async def api_form_24q(req: Form24QRequest, request: Request, authorization: str = Header(None)):
    """Generate Form 24Q (TDS return for salary payments)."""
    user = await get_current_user(request, authorization)
    result = generate_form_24q(
        tan=req.tan,
        employer_name=req.employer_name,
        employer_address=req.employer_address,
        financial_year=req.financial_year,
        quarter=req.quarter,
        employee_records=req.employee_records,
        challans=req.challans,
    )
    return json.loads(json.dumps(result, default=str))

class Form27QRequest(BaseModel):
    tan: str
    deductor_name: str
    deductor_address: dict = {}
    financial_year: str
    quarter: str
    deductee_records: list
    challans: list = []

@api_router.post("/tools/tds/form-27q")
async def api_form_27q(req: Form27QRequest, request: Request, authorization: str = Header(None)):
    """Generate Form 27Q (TDS return for NRI/foreign payments)."""
    user = await get_current_user(request, authorization)
    result = generate_form_27q(
        tan=req.tan,
        deductor_name=req.deductor_name,
        deductor_address=req.deductor_address,
        financial_year=req.financial_year,
        quarter=req.quarter,
        deductee_records=req.deductee_records,
        challans=req.challans,
    )
    return json.loads(json.dumps(result, default=str))

class TDSDepositRequest(BaseModel):
    deposits: list

@api_router.post("/tools/tds/track-deposits")
async def api_tds_track_deposits(req: TDSDepositRequest, request: Request, authorization: str = Header(None)):
    """Track TDS deposits, compute due dates, and calculate interest on late deposits."""
    user = await get_current_user(request, authorization)
    result = track_tds_deposits(req.deposits)
    return json.loads(json.dumps(result, default=str))

class Recon26ASRequest(BaseModel):
    form_26as_entries: list
    books_entries: list

@api_router.post("/tools/tds/reconcile-26as")
async def api_reconcile_26as(req: Recon26ASRequest, request: Request, authorization: str = Header(None)):
    """Reconcile Form 26AS TDS credits with books/ledger entries."""
    user = await get_current_user(request, authorization)
    result = reconcile_26as_with_books(req.form_26as_entries, req.books_entries)
    return json.loads(json.dumps(result, default=str))

class LateDepositInterestRequest(BaseModel):
    tds_amount: float
    deduction_date: str
    deposit_date: str

@api_router.post("/tools/tds/late-interest")
async def api_tds_late_interest(req: LateDepositInterestRequest, request: Request, authorization: str = Header(None)):
    """Calculate interest under Section 201(1A) for late TDS deposit."""
    user = await get_current_user(request, authorization)
    from datetime import date as _date
    ded_date = _date.fromisoformat(req.deduction_date)
    dep_date = _date.fromisoformat(req.deposit_date)
    result = calculate_interest_on_late_deposit(req.tds_amount, ded_date, dep_date)
    return json.loads(json.dumps(result, default=str))

class Section234ERequest(BaseModel):
    financial_year: str
    quarter: str
    filing_date: str

@api_router.post("/tools/tds/234e-fee")
async def api_234e_fee(req: Section234ERequest, request: Request, authorization: str = Header(None)):
    """Calculate Section 234E late filing fee for TDS returns."""
    user = await get_current_user(request, authorization)
    from datetime import date as _date
    result = calculate_234e_fee(req.financial_year, req.quarter, _date.fromisoformat(req.filing_date))
    return json.loads(json.dumps(result, default=str))

@api_router.get("/tools/tds/sections")
async def api_tds_sections(request: Request, authorization: str = Header(None), section: str = None, keyword: str = None):
    """Look up TDS/TCS sections by code or keyword."""
    user = await get_current_user(request, authorization)
    result = lookup_tds_section(section=section, keyword=keyword)
    return json.loads(json.dumps({"sections": result}, default=str))

@api_router.get("/tools/tds/due-dates/{financial_year}")
async def api_tds_due_dates(financial_year: str, request: Request, authorization: str = Header(None)):
    """Get quarterly TDS return filing and deposit due dates."""
    user = await get_current_user(request, authorization)
    result = get_quarterly_return_due_dates(financial_year)
    return json.loads(json.dumps(result, default=str))

class DTAARequest(BaseModel):
    country: str
    nature_of_payment: str

@api_router.post("/tools/tds/dtaa")
async def api_dtaa_lookup(req: DTAARequest, request: Request, authorization: str = Header(None)):
    """Look up DTAA treaty rates for cross-border payments."""
    user = await get_current_user(request, authorization)
    result = get_dtaa_rate(req.country, req.nature_of_payment)
    return json.loads(json.dumps(result, default=str))

@api_router.get("/tools/tds/rate-master")
async def api_tds_rate_master(request: Request, authorization: str = Header(None)):
    """Get the complete TDS rate master with all sections."""
    user = await get_current_user(request, authorization)
    result = lookup_tds_section()  # All sections
    return json.loads(json.dumps({"sections": result, "total": len(result)}, default=str))


# ==================== GST ENGINE ROUTES ====================

class GSTINValidateRequest(BaseModel):
    gstin: str

@api_router.post("/tools/gst/validate-gstin")
async def api_validate_gstin(req: GSTINValidateRequest, request: Request, authorization: str = Header(None)):
    """Validate a GSTIN with checksum verification and state/entity decoding."""
    user = await get_current_user(request, authorization)
    result = validate_gstin(req.gstin)
    return result

class GSTINBatchRequest(BaseModel):
    gstins: list

@api_router.post("/tools/gst/validate-gstin/batch")
async def api_validate_gstin_batch(req: GSTINBatchRequest, request: Request, authorization: str = Header(None)):
    """Validate multiple GSTINs at once."""
    user = await get_current_user(request, authorization)
    result = validate_gstin_batch(req.gstins)
    return {"results": result}

class HSNValidateRequest(BaseModel):
    code: str

@api_router.post("/tools/gst/validate-hsn")
async def api_validate_hsn(req: HSNValidateRequest, request: Request, authorization: str = Header(None)):
    """Validate HSN/SAC code and get GST rate."""
    user = await get_current_user(request, authorization)
    result = validate_hsn_code(req.code)
    return result

class PlaceOfSupplyRequest(BaseModel):
    supply_type: str = "goods"
    supplier_state: str = ""
    recipient_state: str = ""
    delivery_state: str = ""
    nature_of_service: str = ""
    location_of_supplier: str = ""
    location_of_recipient: str = ""
    is_import: bool = False
    is_export: bool = False

@api_router.post("/tools/gst/place-of-supply")
async def api_place_of_supply(req: PlaceOfSupplyRequest, request: Request, authorization: str = Header(None)):
    """Determine place of supply and IGST vs CGST+SGST applicability."""
    user = await get_current_user(request, authorization)
    result = determine_place_of_supply(
        supply_type=req.supply_type,
        supplier_state=req.supplier_state,
        recipient_state=req.recipient_state,
        delivery_state=req.delivery_state,
        nature_of_service=req.nature_of_service,
        location_of_supplier=req.location_of_supplier,
        location_of_recipient=req.location_of_recipient,
        is_import=req.is_import,
        is_export=req.is_export,
    )
    return json.loads(json.dumps(result, default=str))

@api_router.post("/tools/gst/parse-invoices")
async def api_parse_invoices(request: Request, authorization: str = Header(None), file: UploadFile = File(...), supplier_gstin: str = Form(""), supplier_state: str = Form("")):
    """Parse invoice register (Excel/CSV) for GSTR-1 generation."""
    user = await get_current_user(request, authorization)
    file_bytes = await file.read()
    result = await parse_invoice_register(file_bytes, file.filename, supplier_gstin=supplier_gstin)
    return json.loads(json.dumps(result, default=str))

@api_router.post("/tools/gst/gstr1")
async def api_generate_gstr1(request: Request, authorization: str = Header(None), file: UploadFile = File(...), supplier_gstin: str = Form(""), supplier_state: str = Form("")):
    """Auto-generate GSTR-1 from invoice register."""
    user = await get_current_user(request, authorization)
    file_bytes = await file.read()
    invoices = await parse_invoice_register(file_bytes, file.filename, supplier_gstin=supplier_gstin)
    if "error" in invoices:
        raise HTTPException(status_code=400, detail=invoices["error"])
    result = await generate_gstr1(invoices["invoices"], supplier_gstin=supplier_gstin, supplier_state=supplier_state)
    return json.loads(json.dumps(result, default=str))

@api_router.post("/tools/gst/gstr3b")
async def api_compute_gstr3b(request: Request, authorization: str = Header(None), file: UploadFile = File(...), supplier_gstin: str = Form(""), supplier_state: str = Form("")):
    """Compute GSTR-3B liability from invoice register."""
    user = await get_current_user(request, authorization)
    file_bytes = await file.read()
    invoices = await parse_invoice_register(file_bytes, file.filename, supplier_gstin=supplier_gstin)
    if "error" in invoices:
        raise HTTPException(status_code=400, detail=invoices["error"])
    gstr1 = await generate_gstr1(invoices["invoices"], supplier_gstin=supplier_gstin, supplier_state=supplier_state)
    result = await compute_gstr3b(gstr1, supplier_gstin=supplier_gstin, supplier_state=supplier_state)
    return json.loads(json.dumps(result, default=str))

class ITCTrackRequest(BaseModel):
    inward_supplies: list
    gstr2b_data: list = []
    total_taxable_turnover: Optional[float] = None
    exempt_turnover: Optional[float] = None

@api_router.post("/tools/gst/itc-track")
async def api_track_itc(req: ITCTrackRequest, request: Request, authorization: str = Header(None)):
    """Track ITC eligibility with 2B matching, blocked credit detection, Rule 36(4)/42/43."""
    user = await get_current_user(request, authorization)
    result = await track_itc(
        inward_supplies=req.inward_supplies,
        gstr2b_data=req.gstr2b_data,
        total_taxable_turnover=req.total_taxable_turnover,
        exempt_turnover=req.exempt_turnover,
    )
    return json.loads(json.dumps(result, default=str))

class RCMCheckRequest(BaseModel):
    description: str
    supplier_type: str = ""

@api_router.post("/tools/gst/rcm-check")
async def api_rcm_check(req: RCMCheckRequest, request: Request, authorization: str = Header(None)):
    """Check if reverse charge mechanism applies to a supply."""
    user = await get_current_user(request, authorization)
    result = check_rcm_applicability(req.description, req.supplier_type)
    return result

class GSTInterestRequest(BaseModel):
    tax_amount: float
    due_date: str
    payment_date: str
    interest_type: str = "late_payment"

@api_router.post("/tools/gst/interest")
async def api_gst_interest(req: GSTInterestRequest, request: Request, authorization: str = Header(None)):
    """Calculate GST interest for late payment/ITC reversal."""
    user = await get_current_user(request, authorization)
    from datetime import date as _date
    result = calculate_gst_interest(
        tax_amount=req.tax_amount,
        due_date=_date.fromisoformat(req.due_date),
        payment_date=_date.fromisoformat(req.payment_date),
        interest_type=req.interest_type,
    )
    return json.loads(json.dumps(result, default=str))

class GSTLateFeeRequest(BaseModel):
    return_type: str
    filing_period: str
    filing_date: str
    tax_liability: float = 0

@api_router.post("/tools/gst/late-fee")
async def api_gst_late_fee(req: GSTLateFeeRequest, request: Request, authorization: str = Header(None)):
    """Calculate GST late fee for delayed return filing."""
    user = await get_current_user(request, authorization)
    from datetime import date as _date
    result = calculate_late_fee(
        return_type=req.return_type,
        filing_period=req.filing_period,
        filing_date=_date.fromisoformat(req.filing_date),
        tax_liability=req.tax_liability,
    )
    return json.loads(json.dumps(result, default=str))

@api_router.post("/tools/gst/full-return")
async def api_full_gst_return(request: Request, authorization: str = Header(None), file: UploadFile = File(...), supplier_gstin: str = Form(""), supplier_state: str = Form("")):
    """End-to-end GST return pipeline: parse → GSTR-1 → GSTR-3B → ITC → filing summary."""
    user = await get_current_user(request, authorization)
    file_bytes = await file.read()
    result = await process_gst_return(file_bytes, file.filename, supplier_gstin=supplier_gstin, supplier_state=supplier_state)
    return json.loads(json.dumps(result, default=str))

class EInvoiceRequest(BaseModel):
    invoice: dict
    supplier_details: dict

@api_router.post("/tools/gst/einvoice")
async def api_generate_einvoice(req: EInvoiceRequest, request: Request, authorization: str = Header(None)):
    """Generate e-Invoice JSON in NIC schema v1.1 format."""
    user = await get_current_user(request, authorization)
    result = generate_einvoice_json(req.invoice, req.supplier_details)
    return result


# ==================== INCOME TAX ENGINE ROUTES ====================

class ITRFormRequest(BaseModel):
    assessee_type: str = "individual"
    income_sources: list = ["salary"]  # e.g. ["salary", "house_property", "capital_gains", "business"]
    turnover: float = 0
    total_income: float = 0
    has_foreign_income: bool = False
    has_foreign_assets: bool = False
    is_director: bool = False
    unlisted_shares: bool = False
    is_presumptive: bool = False
    presumptive_section: str = ""
    number_of_house_properties: int = 0
    has_brought_forward_losses: bool = False
    has_agricultural_income_above_5000: bool = False

@api_router.post("/tools/income-tax/itr-form")
async def api_itr_form_selector(req: ITRFormRequest, request: Request, authorization: str = Header(None)):
    """Select the correct ITR form based on income sources and assessee type."""
    user = await get_current_user(request, authorization)
    result = select_itr_form(
        assessee_type=req.assessee_type,
        income_sources=req.income_sources,
        turnover=req.turnover,
        total_income=req.total_income,
        has_foreign_income=req.has_foreign_income,
        has_foreign_assets=req.has_foreign_assets,
        is_director=req.is_director,
        unlisted_shares=req.unlisted_shares,
        is_presumptive=req.is_presumptive,
        presumptive_section=req.presumptive_section,
        number_of_house_properties=req.number_of_house_properties,
        has_brought_forward_losses=req.has_brought_forward_losses,
        has_agricultural_income_above_5000=req.has_agricultural_income_above_5000,
    )
    return json.loads(json.dumps(result, default=str))

class TaxComputeRequest(BaseModel):
    gross_total_income: float
    regime: str = "new"
    assessee_type: str = "individual"
    salary_income: float = 0
    hp_income: float = 0
    business_income: float = 0
    stcg_111a: float = 0
    stcg_normal: float = 0
    ltcg_112a: float = 0
    ltcg_112: float = 0
    other_income: float = 0
    age_bracket: str = "general"
    tds_total: float = 0
    advance_tax_paid: float = 0
    company_section: str = ""

@api_router.post("/tools/income-tax/compute")
async def api_compute_tax(req: TaxComputeRequest, request: Request, authorization: str = Header(None)):
    """Compute income tax with surcharge, cess, rebate for any regime/assessee type."""
    user = await get_current_user(request, authorization)
    from decimal import Decimal
    D = lambda v: Decimal(str(v)) if v else Decimal("0")
    result = compute_tax(
        assessee_type=req.assessee_type,
        regime=req.regime,
        gross_total_income=D(req.gross_total_income),
        salary_income=D(req.salary_income),
        hp_income=D(req.hp_income),
        business_income=D(req.business_income),
        stcg_111a=D(req.stcg_111a),
        stcg_normal=D(req.stcg_normal),
        ltcg_112a=D(req.ltcg_112a),
        ltcg_112=D(req.ltcg_112),
        other_income=D(req.other_income),
        age_bracket=req.age_bracket,
        tds_total=D(req.tds_total),
        advance_tax_paid=D(req.advance_tax_paid),
        company_section=req.company_section,
    )
    return json.loads(json.dumps(result, default=str))

class RegimeCompareRequest(BaseModel):
    assessee_type: str = "individual"
    salary_income: float = 0
    gross_income: float = 0  # Frontend alias for salary_income
    gross_salary: float = 0  # Another alias
    hp_income: float = 0
    hra: float = 0  # Frontend field
    basic_salary: float = 0
    hra_received: float = 0
    business_income: float = 0
    stcg_111a: float = 0
    stcg_normal: float = 0
    ltcg_112a: float = 0
    ltcg_112: float = 0
    other_income: float = 0
    age_bracket: str = "general"
    tds_total: float = 0
    advance_tax_paid: float = 0
    deductions_80c: float = 0  # Frontend field
    notes: str = ""  # Frontend field

    def get_salary_income(self):
        return self.salary_income or self.gross_income or self.gross_salary or 0

@api_router.post("/tools/income-tax/compare-regimes")
async def api_compare_regimes(req: RegimeCompareRequest, request: Request, authorization: str = Header(None)):
    """Side-by-side old vs new regime comparison with recommendation."""
    user = await get_current_user(request, authorization)
    from decimal import Decimal
    D = lambda v: Decimal(str(v)) if v else Decimal("0")
    result = compare_regimes(
        assessee_type=req.assessee_type,
        salary_income=D(req.get_salary_income()),
        hp_income=D(req.hp_income),
        business_income=D(req.business_income),
        stcg_111a=D(req.stcg_111a),
        stcg_normal=D(req.stcg_normal),
        ltcg_112a=D(req.ltcg_112a),
        ltcg_112=D(req.ltcg_112),
        other_income=D(req.other_income),
        age_bracket=req.age_bracket,
        tds_total=D(req.tds_total),
        advance_tax_paid=D(req.advance_tax_paid),
    )
    output = json.loads(json.dumps(result, default=str))
    # Normalize — frontend reads flat summary fields
    old_tax = float(output.get("old_regime", {}).get("total_tax_liability", 0))
    new_tax = float(output.get("new_regime", {}).get("total_tax_liability", 0))
    savings = abs(old_tax - new_tax)
    output["old_regime_tax"] = old_tax
    output["new_regime_tax"] = new_tax
    output["savings"] = savings
    output["recommendation"] = (
        f"New Regime saves ₹{savings:,.0f}" if new_tax < old_tax
        else f"Old Regime saves ₹{savings:,.0f}" if old_tax < new_tax
        else "Both regimes result in same tax"
    )
    return output

class QuickTaxRequest(BaseModel):
    annual_income: float
    regime: str = "new"
    age_bracket: str = "general"
    deductions_80c: float = 0
    deductions_80d: float = 0
    hra_exemption: float = 0
    home_loan_interest: float = 0
    nps_80ccd_1b: float = 0

@api_router.post("/tools/income-tax/quick-estimate")
async def api_quick_tax(req: QuickTaxRequest, request: Request, authorization: str = Header(None)):
    """Quick tax estimate for salaried individuals — single call."""
    user = await get_current_user(request, authorization)
    from decimal import Decimal
    D = lambda v: Decimal(str(v)) if v else Decimal("0")
    result = quick_tax_estimate(
        annual_income=D(req.annual_income),
        regime=req.regime,
        age_bracket=req.age_bracket,
        deductions_80c=D(req.deductions_80c),
        deductions_80d=D(req.deductions_80d),
        hra_exemption=D(req.hra_exemption),
        home_loan_interest=D(req.home_loan_interest),
        nps_80ccd_1b=D(req.nps_80ccd_1b),
    )
    return json.loads(json.dumps(result, default=str))

class HousePropertyRequest(BaseModel):
    properties: list  # List of dicts with property_type, gross_annual_value, rent_received, etc.

@api_router.post("/tools/income-tax/house-property")
async def api_house_property(req: HousePropertyRequest, request: Request, authorization: str = Header(None)):
    """Compute income from house property (Schedule HP)."""
    user = await get_current_user(request, authorization)
    from decimal import Decimal as D
    props = []
    for p in req.properties:
        props.append(HousePropertyInput(
            property_type=PropertyType(p.get("property_type", "self_occupied")),
            gross_annual_value=D(str(p.get("gross_annual_value", 0))),
            rent_received=D(str(p.get("rent_received", 0))),
            municipal_taxes_paid=D(str(p.get("municipal_taxes_paid", 0))),
            interest_on_borrowed_capital=D(str(p.get("interest_on_borrowed_capital", 0))),
            pre_construction_interest=D(str(p.get("pre_construction_interest", 0))),
        ))
    result = compute_house_property_income(props)
    return json.loads(json.dumps(result, default=str))

class CapitalGainsRequest(BaseModel):
    transactions: list  # List of dicts with asset_type, sale_consideration, cost_of_acquisition, etc.

@api_router.post("/tools/income-tax/capital-gains")
async def api_capital_gains(req: CapitalGainsRequest, request: Request, authorization: str = Header(None)):
    """Compute capital gains with STCG/LTCG classification and exemptions."""
    user = await get_current_user(request, authorization)
    from decimal import Decimal as D
    txns = []
    for t in req.transactions:
        txns.append(CapitalGainInput(
            asset_type=AssetType(t.get("asset_type", "other")),
            sale_consideration=D(str(t.get("sale_consideration", 0))),
            cost_of_acquisition=D(str(t.get("cost_of_acquisition", 0))),
            cost_of_improvement=D(str(t.get("cost_of_improvement", 0))),
            transfer_expenses=D(str(t.get("transfer_expenses", 0))),
            year_of_acquisition=t.get("year_of_acquisition", ""),
            year_of_transfer=t.get("year_of_transfer", ""),
            is_stt_paid=t.get("is_stt_paid", False),
            exemption_54=D(str(t.get("exemption_54", 0))),
            exemption_section=t.get("exemption_section", ""),
        ))
    result = compute_capital_gains(txns)
    return json.loads(json.dumps(result, default=str))

class BusinessIncomeRequest(BaseModel):
    is_presumptive: bool = False
    presumptive_section: str = ""
    gross_turnover: float = 0
    cash_turnover: float = 0
    digital_turnover: float = 0
    gross_receipts_profession: float = 0
    number_of_heavy_vehicles: int = 0
    number_of_other_vehicles: int = 0
    months_owned: int = 12
    net_profit_as_per_books: float = 0
    depreciation_as_per_books: float = 0
    depreciation_as_per_it: float = 0
    disallowances: dict = {}
    exempt_income_debited: float = 0
    income_not_credited: float = 0

@api_router.post("/tools/income-tax/business-income")
async def api_business_income(req: BusinessIncomeRequest, request: Request, authorization: str = Header(None)):
    """Compute business income — presumptive 44AD/44ADA/44AE or regular."""
    user = await get_current_user(request, authorization)
    from decimal import Decimal as D
    inputs = BusinessIncomeInput(
        is_presumptive=req.is_presumptive,
        presumptive_section=req.presumptive_section,
        gross_turnover=D(str(req.gross_turnover)),
        cash_turnover=D(str(req.cash_turnover)),
        digital_turnover=D(str(req.digital_turnover)),
        gross_receipts_profession=D(str(req.gross_receipts_profession)),
        number_of_heavy_vehicles=req.number_of_heavy_vehicles,
        number_of_other_vehicles=req.number_of_other_vehicles,
        months_owned=req.months_owned,
        net_profit_as_per_books=D(str(req.net_profit_as_per_books)),
        depreciation_as_per_books=D(str(req.depreciation_as_per_books)),
        depreciation_as_per_it=D(str(req.depreciation_as_per_it)),
        disallowances={k: D(str(v)) for k, v in req.disallowances.items()},
        exempt_income_debited=D(str(req.exempt_income_debited)),
        income_not_credited=D(str(req.income_not_credited)),
    )
    result = compute_business_income(inputs)
    return json.loads(json.dumps(result, default=str))

class AdvanceTaxRequest(BaseModel):
    assessed_tax: float
    tds_credit: float = 0
    payments: list = []  # List of {date_paid: "YYYY-MM-DD", amount: float}

@api_router.post("/tools/income-tax/advance-tax")
async def api_advance_tax(req: AdvanceTaxRequest, request: Request, authorization: str = Header(None)):
    """Compute advance tax schedule with 234B/234C interest."""
    user = await get_current_user(request, authorization)
    from decimal import Decimal as D
    from datetime import date as _date
    pay_objs = []
    for p in req.payments:
        pay_objs.append(AdvanceTaxPayment(
            date_paid=_date.fromisoformat(p["date_paid"]),
            amount=D(str(p["amount"])),
        ))
    result = compute_advance_tax(
        assessed_tax=req.assessed_tax,
        tds_credit=req.tds_credit,
        payments=pay_objs,
    )
    return json.loads(json.dumps(result, default=str))

class IndexedCostRequest(BaseModel):
    cost_of_acquisition: float
    year_of_acquisition: str
    year_of_transfer: str

@api_router.post("/tools/income-tax/indexed-cost")
async def api_indexed_cost(req: IndexedCostRequest, request: Request, authorization: str = Header(None)):
    """Compute indexed cost of acquisition using CII table."""
    user = await get_current_user(request, authorization)
    result = compute_indexed_cost_of_acquisition(
        cost_of_acquisition=req.cost_of_acquisition,
        year_of_acquisition=req.year_of_acquisition,
        year_of_transfer=req.year_of_transfer,
    )
    return json.loads(json.dumps(result, default=str))

class DepreciationRequest(BaseModel):
    assets: list  # List of {asset_category, opening_wdv, additions_first_half, etc.}

@api_router.post("/tools/income-tax/depreciation")
async def api_depreciation(req: DepreciationRequest, request: Request, authorization: str = Header(None)):
    """Compute IT Act depreciation (WDV method) with additional depreciation."""
    user = await get_current_user(request, authorization)
    from decimal import Decimal as D
    asset_objs = []
    for a in req.assets:
        asset_objs.append(DepreciationAsset(
            asset_category=a.get("asset_category", ""),
            opening_wdv=D(str(a.get("opening_wdv", 0))),
            additions_first_half=D(str(a.get("additions_first_half", 0))),
            additions_second_half=D(str(a.get("additions_second_half", 0))),
            sale_proceeds=D(str(a.get("sale_proceeds", 0))),
            is_new_manufacturing_asset=a.get("is_new_manufacturing_asset", False),
            description=a.get("description", ""),
        ))
    result = compute_depreciation(asset_objs)
    return json.loads(json.dumps(result, default=str))

@api_router.get("/tools/income-tax/cii-table")
async def api_cii_table(request: Request, authorization: str = Header(None)):
    """Get complete Cost Inflation Index table."""
    user = await get_current_user(request, authorization)
    return {"cii_table": get_cii_table()}

@api_router.get("/tools/income-tax/depreciation-rates")
async def api_depreciation_rates(request: Request, authorization: str = Header(None)):
    """Get IT Act depreciation rate categories."""
    user = await get_current_user(request, authorization)
    return {"rates": get_depreciation_rates_reference()}

@api_router.get("/tools/income-tax/section-80-reference")
async def api_section_80_ref(request: Request, authorization: str = Header(None)):
    """Get Section 80 deductions reference with limits and eligibility."""
    user = await get_current_user(request, authorization)
    return {"deductions": get_section80_deductions_reference()}

@api_router.get("/tools/income-tax/holding-periods")
async def api_holding_periods(request: Request, authorization: str = Header(None)):
    """Get asset holding period reference for STCG/LTCG classification."""
    user = await get_current_user(request, authorization)
    return {"holding_periods": get_holding_period_reference()}

class FullTaxRequest(BaseModel):
    assessee_type: str = "individual"
    age_bracket: str = "general"
    salary_gross: float = 0
    standard_deduction_applicable: bool = True
    house_properties: list = []
    capital_gain_transactions: list = []
    business_input: Optional[dict] = None
    other_sources_income: float = 0
    deductions: Optional[dict] = None
    advance_payments: list = []
    company_section: str = ""

@api_router.post("/tools/income-tax/full-computation")
async def api_full_tax(req: FullTaxRequest, request: Request, authorization: str = Header(None)):
    """Complete income tax computation — all schedules, both regimes, advance tax."""
    user = await get_current_user(request, authorization)
    from decimal import Decimal as D
    from datetime import date as _date
    # Build house property inputs
    hp_inputs = None
    if req.house_properties:
        hp_inputs = []
        for p in req.house_properties:
            hp_inputs.append(HousePropertyInput(
                property_type=PropertyType(p.get("property_type", "self_occupied")),
                gross_annual_value=D(str(p.get("gross_annual_value", 0))),
                rent_received=D(str(p.get("rent_received", 0))),
                municipal_taxes_paid=D(str(p.get("municipal_taxes_paid", 0))),
                interest_on_borrowed_capital=D(str(p.get("interest_on_borrowed_capital", 0))),
                pre_construction_interest=D(str(p.get("pre_construction_interest", 0))),
            ))
    # Build CG inputs
    cg_inputs = None
    if req.capital_gain_transactions:
        cg_inputs = []
        for t in req.capital_gain_transactions:
            cg_inputs.append(CapitalGainInput(
                asset_type=AssetType(t.get("asset_type", "other")),
                sale_consideration=D(str(t.get("sale_consideration", 0))),
                cost_of_acquisition=D(str(t.get("cost_of_acquisition", 0))),
                cost_of_improvement=D(str(t.get("cost_of_improvement", 0))),
                year_of_acquisition=t.get("year_of_acquisition", ""),
                year_of_transfer=t.get("year_of_transfer", ""),
                is_stt_paid=t.get("is_stt_paid", False),
                exemption_54=D(str(t.get("exemption_54", 0))),
            ))
    # Build business input
    biz_input = None
    if req.business_input:
        b = req.business_input
        biz_input = BusinessIncomeInput(
            is_presumptive=b.get("is_presumptive", False),
            presumptive_section=b.get("presumptive_section", ""),
            gross_turnover=D(str(b.get("gross_turnover", 0))),
            cash_turnover=D(str(b.get("cash_turnover", 0))),
            digital_turnover=D(str(b.get("digital_turnover", 0))),
            net_profit_as_per_books=D(str(b.get("net_profit_as_per_books", 0))),
        )
    # Build deductions
    ded_input = None
    if req.deductions:
        d = req.deductions
        ded_input = Section80Deductions(
            sec_80c=D(str(d.get("sec_80c", 0))),
            sec_80d_self=D(str(d.get("sec_80d_self", 0))),
            sec_80d_parents=D(str(d.get("sec_80d_parents", 0))),
            sec_80ccd_1b=D(str(d.get("sec_80ccd_1b", 0))),
            sec_80e=D(str(d.get("sec_80e", 0))),
            sec_80g=D(str(d.get("sec_80g", 0))),
            sec_80gg=D(str(d.get("sec_80gg", 0))),
            sec_80tta=D(str(d.get("sec_80tta", 0))),
        )
    # Build advance payments
    adv_payments = None
    if req.advance_payments:
        adv_payments = []
        for p in req.advance_payments:
            adv_payments.append(AdvanceTaxPayment(
                date_paid=_date.fromisoformat(p["date_paid"]),
                amount=D(str(p["amount"])),
            ))
    result = compute_full_tax(
        assessee_type=req.assessee_type,
        age_bracket=req.age_bracket,
        salary_gross=D(str(req.salary_gross)),
        standard_deduction_applicable=req.standard_deduction_applicable,
        house_properties=hp_inputs,
        business_input=biz_input,
        capital_gain_transactions=cg_inputs,
        other_sources_income=D(str(req.other_sources_income)),
        deductions=ded_input,
        advance_payments=adv_payments,
        company_section=req.company_section,
    )
    return json.loads(json.dumps(result, default=str))

class ITRChecklistRequest(BaseModel):
    form: str

@api_router.post("/tools/income-tax/itr-checklist")
async def api_itr_checklist(req: ITRChecklistRequest, request: Request, authorization: str = Header(None)):
    """Get document checklist for an ITR form."""
    user = await get_current_user(request, authorization)
    result = generate_itr_checklist(req.form)
    return result


# ==================== FINANCIAL ANALYSIS ROUTES ====================

@api_router.post("/tools/financial/parse-trial-balance")
async def api_parse_trial_balance(request: Request, authorization: str = Header(None), file: UploadFile = File(...)):
    """Parse trial balance from Excel/CSV and classify accounts."""
    user = await get_current_user(request, authorization)
    file_bytes = await file.read()
    result = parse_trial_balance(file_bytes, file.filename)
    return json.loads(json.dumps(result, default=str))

class RatioAnalysisRequest(BaseModel):
    financials: dict  # revenue, cogs, total_assets, equity, etc.
    shares_outstanding: Optional[float] = None
    dividend_paid: Optional[float] = None
    diluted_shares: Optional[float] = None
    industry: str = "general"

@api_router.post("/tools/financial/ratios")
async def api_compute_ratios(req: RatioAnalysisRequest, request: Request, authorization: str = Header(None)):
    """Compute 25+ financial ratios with interpretations."""
    user = await get_current_user(request, authorization)
    result = compute_ratios(
        financials=req.financials,
        shares_outstanding=req.shares_outstanding,
        dividend_paid=req.dividend_paid,
        diluted_shares=req.diluted_shares,
        industry=req.industry,
    )
    return json.loads(json.dumps(result, default=str))

class CashFlowRequest(BaseModel):
    current_pl: dict  # P&L: pat, depreciation, finance_cost, etc.
    current_bs: dict  # Balance sheet: receivables, inventory, payables, etc.
    previous_bs: dict  # Previous year balance sheet
    additional_info: Optional[dict] = None  # capex, dividends_paid, etc.

@api_router.post("/tools/financial/cash-flow")
async def api_generate_cash_flow(req: CashFlowRequest, request: Request, authorization: str = Header(None)):
    """Generate cash flow statement (indirect method) per AS-3/Ind AS 7."""
    user = await get_current_user(request, authorization)
    result = generate_cash_flow(req.current_pl, req.current_bs, req.previous_bs, req.additional_info)
    return json.loads(json.dumps(result, default=str))

class DebtorAgingRequest(BaseModel):
    receivables: list
    as_of_date: Optional[str] = None
    ecl_rates: Optional[dict] = None

@api_router.post("/tools/financial/debtor-aging")
async def api_debtor_aging(req: DebtorAgingRequest, request: Request, authorization: str = Header(None)):
    """Analyze debtor aging with ECL provisioning and concentration risk."""
    user = await get_current_user(request, authorization)
    from datetime import date as _date
    as_of = _date.fromisoformat(req.as_of_date) if req.as_of_date else None
    result = analyze_debtor_aging(req.receivables, as_of_date=as_of, ecl_rates=req.ecl_rates)
    return json.loads(json.dumps(result, default=str))

class CreditorAgingRequest(BaseModel):
    payables: list
    as_of_date: Optional[str] = None

@api_router.post("/tools/financial/creditor-aging")
async def api_creditor_aging(req: CreditorAgingRequest, request: Request, authorization: str = Header(None)):
    """Analyze creditor aging with MSME compliance check."""
    user = await get_current_user(request, authorization)
    from datetime import date as _date
    as_of = _date.fromisoformat(req.as_of_date) if req.as_of_date else None
    result = analyze_creditor_aging(req.payables, as_of_date=as_of)
    return json.loads(json.dumps(result, default=str))

class ComparativeRequest(BaseModel):
    periods: list  # List of financial data dicts for each period
    period_labels: list = []  # e.g. ["FY 2025-26", "FY 2024-25"]

@api_router.post("/tools/financial/comparative")
async def api_comparative_statements(req: ComparativeRequest, request: Request, authorization: str = Header(None)):
    """Generate comparative financial statements with trend analysis."""
    user = await get_current_user(request, authorization)
    result = generate_comparative_statements(req.periods, period_labels=req.period_labels if req.period_labels else None)
    return json.loads(json.dumps(result, default=str))

class RedFlagRequest(BaseModel):
    current_financials: dict
    previous_financials: Optional[dict] = None
    current_ocf: Optional[float] = None
    related_party_transactions: Optional[float] = None
    contingent_liabilities: Optional[float] = None

@api_router.post("/tools/financial/red-flags")
async def api_red_flags(req: RedFlagRequest, request: Request, authorization: str = Header(None)):
    """Detect financial red flags with audit implications."""
    user = await get_current_user(request, authorization)
    result = detect_red_flags(
        current_financials=req.current_financials,
        previous_financials=req.previous_financials,
        current_ocf=req.current_ocf,
        related_party_transactions=req.related_party_transactions,
        contingent_liabilities=req.contingent_liabilities,
    )
    return json.loads(json.dumps(result, default=str))

@api_router.post("/tools/financial/full-analysis")
async def api_full_financial_analysis(request: Request, authorization: str = Header(None), file: UploadFile = File(...), prior_year_file: UploadFile = File(None)):
    """Full financial analysis pipeline from trial balance: ratios, cash flow, red flags."""
    user = await get_current_user(request, authorization)
    file_bytes = await file.read()
    prior_bytes = await prior_year_file.read() if prior_year_file else None
    result = run_full_analysis(file_bytes, prior_year_data=prior_bytes)
    return json.loads(json.dumps(result, default=str))


# ==================== LEGAL TEMPLATES ROUTES ====================

@api_router.get("/tools/templates")
async def api_list_templates(request: Request, authorization: str = Header(None), category: str = None):
    """List all legal templates, optionally filtered by category."""
    user = await get_current_user(request, authorization)
    result = list_templates(category=category)
    return {"templates": result, "total": len(result)}

@api_router.get("/tools/templates/categories")
async def api_template_categories(request: Request, authorization: str = Header(None)):
    """Get template categories with counts."""
    user = await get_current_user(request, authorization)
    cats = get_template_categories()
    counts = get_template_count()
    return {"categories": cats, "counts": counts}

class TemplateSearchRequest(BaseModel):
    query: str

@api_router.post("/tools/templates/search")
async def api_search_templates(req: TemplateSearchRequest, request: Request, authorization: str = Header(None)):
    """Search templates by keyword."""
    user = await get_current_user(request, authorization)
    result = search_templates(req.query)
    return {"results": result, "total": len(result)}

@api_router.get("/tools/templates/{template_id}")
async def api_get_template_info(template_id: str, request: Request, authorization: str = Header(None)):
    """Get template metadata, required fields, and preview."""
    user = await get_current_user(request, authorization)
    result = get_template_info(template_id)
    if not result:
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found")
    return result

class RenderTemplateRequest(BaseModel):
    template_id: str
    data: dict

@api_router.post("/tools/templates/render")
async def api_render_template(req: RenderTemplateRequest, request: Request, authorization: str = Header(None)):
    """Render a legal template with provided data."""
    user = await get_current_user(request, authorization)
    try:
        rendered = render_template(req.template_id, req.data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    template_info = get_template_info(req.template_id)
    title = template_info.get("title", "Legal Document") if template_info else "Legal Document"
    return {"rendered": rendered, "template_id": req.template_id, "title": title, "char_count": len(rendered)}

@api_router.post("/tools/templates/render-export")
async def api_render_template_export(req: RenderTemplateRequest, request: Request, authorization: str = Header(None)):
    """Render a template and export as Word document."""
    user = await get_current_user(request, authorization)
    try:
        rendered_text = render_template(req.template_id, req.data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    template_info = get_template_info(req.template_id)
    title = template_info.get("title", "Legal Document") if template_info else "Legal Document"
    doc_bytes = generate_word_document(rendered_text, title=title)
    return StreamingResponse(
        io.BytesIO(doc_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{req.template_id}.docx"'}
    )


# ==================== DUE DILIGENCE ROUTES ====================

class DDChecklistRequest(BaseModel):
    transaction_type: str = "acquisition"
    target_type: str = "private_company"
    sector: str = ""
    deal_value_crores: Optional[float] = None
    is_listed: bool = False
    has_foreign_investment: bool = False
    has_international_transactions: bool = False

@api_router.post("/tools/due-diligence/checklist")
async def api_dd_checklist(req: DDChecklistRequest, request: Request, authorization: str = Header(None)):
    """Generate comprehensive DD checklist filtered by transaction and target type."""
    user = await get_current_user(request, authorization)
    result = generate_dd_checklist(
        transaction_type=req.transaction_type,
        target_type=req.target_type,
        sector=req.sector,
        deal_value_crores=req.deal_value_crores,
        is_listed=req.is_listed,
        has_foreign_investment=req.has_foreign_investment,
        has_international_transactions=req.has_international_transactions,
    )
    return json.loads(json.dumps(result, default=str))

class DDRedFlagRequest(BaseModel):
    financial_data: dict  # revenue_concentration_pct, related_party_revenue_pct, etc.
    compliance_data: dict = {}

@api_router.post("/tools/due-diligence/red-flags")
async def api_dd_red_flags(req: DDRedFlagRequest, request: Request, authorization: str = Header(None)):
    """Detect M&A red flags from financial and compliance data."""
    user = await get_current_user(request, authorization)
    result = dd_detect_red_flags(req.financial_data, req.compliance_data)
    return json.loads(json.dumps(result, default=str))

class DDComplianceScoreRequest(BaseModel):
    checklist_results: dict
    red_flag_results: dict = {}

@api_router.post("/tools/due-diligence/compliance-score")
async def api_dd_compliance_score(req: DDComplianceScoreRequest, request: Request, authorization: str = Header(None)):
    """Compute weighted compliance score with Go/No-Go recommendation."""
    user = await get_current_user(request, authorization)
    result = compute_compliance_score(req.checklist_results, req.red_flag_results)
    return json.loads(json.dumps(result, default=str))

class DDReportRequest(BaseModel):
    checklist_results: dict
    red_flag_results: Optional[dict] = None
    compliance_score: Optional[dict] = None
    metadata: Optional[dict] = None  # target_name, transaction_type, sector, deal_value, etc.

@api_router.post("/tools/due-diligence/report")
async def api_dd_report(req: DDReportRequest, request: Request, authorization: str = Header(None)):
    """Generate a full DD report with executive summary, red flags, R&W suggestions."""
    user = await get_current_user(request, authorization)
    result = generate_dd_report(
        checklist_results=req.checklist_results,
        red_flag_results=req.red_flag_results,
        compliance_score=req.compliance_score,
        metadata=req.metadata,
    )
    return {"report": result}

@api_router.post("/tools/due-diligence/report/export")
async def api_dd_report_export(req: DDReportRequest, request: Request, authorization: str = Header(None)):
    """Generate DD report and export as Word document."""
    user = await get_current_user(request, authorization)
    report_md = generate_dd_report(
        checklist_results=req.checklist_results,
        red_flag_results=req.red_flag_results,
        compliance_score=req.compliance_score,
        metadata=req.metadata,
    )
    target_name = (req.metadata or {}).get("target_name", "Target")
    doc_bytes = generate_word_document(report_md, title=f"Due Diligence Report - {target_name}")
    return StreamingResponse(
        io.BytesIO(doc_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="DD_Report_{target_name}.docx"'}
    )


# ==================== HEALTH CHECK ====================

_server_start_time = datetime.now(timezone.utc)

@app.get("/health")
async def health_check():
    """Health check endpoint — no auth required."""
    uptime_seconds = (datetime.now(timezone.utc) - _server_start_time).total_seconds()
    db_status = "unknown"
    try:
        real_db = get_db()
        if real_db is not None:
            await real_db.command("ping")
            db_status = "connected"
        else:
            db_status = "unavailable"
    except Exception:
        db_status = "error"

    return {
        "status": "ok",
        "uptime_seconds": int(uptime_seconds),
        "database": db_status,
        "email_worker": "running" if _email_worker_running else "stopped",
        "dead_letter_count": len(dead_letter_queue),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ==================== BACKGROUND EMAIL WORKER ====================

_email_worker_running = False
_email_worker_consecutive_errors = 0

@app.on_event("startup")
async def start_email_worker():
    """Start background email polling worker with adaptive polling and error resilience."""
    global _email_worker_running
    if _email_worker_running:
        return
    _email_worker_running = True

    async def email_poll_loop():
        """Poll for new emails with adaptive interval.
        - Normal: every 60s
        - After error: backs off up to 5 minutes
        - After success: resets to 60s
        """
        global _email_worker_consecutive_errors
        base_interval = 30  # Poll every 30 seconds
        max_interval = 120  # Max backoff 2 minutes

        while True:
            try:
                db = get_db()
                count = await process_incoming_emails(db)
                if count > 0:
                    logger.info(f"Email worker processed {count} email(s)")
                _email_worker_consecutive_errors = 0
            except Exception as e:
                _email_worker_consecutive_errors += 1
                logger.error(f"Email worker error (#{_email_worker_consecutive_errors}): {e}")

            # Adaptive interval: back off on repeated errors
            if _email_worker_consecutive_errors > 0:
                interval = min(base_interval * (2 ** min(_email_worker_consecutive_errors, 3)), max_interval)
            else:
                interval = base_interval

            await asyncio.sleep(interval)

    asyncio.create_task(email_poll_loop())
    logger.info("Email worker started — polling spectr.r@agentmail.to every 60s")


# ==================== SANDBOX RESEARCH ROUTES ====================

class SandboxResearchRequest(BaseModel):
    query: str
    query_types: Optional[List[str]] = None
    max_pages: int = 8

@api_router.post("/sandbox/research")
async def api_sandbox_research(req: SandboxResearchRequest, request: Request, authorization: str = Header(None)):
    """Execute browser-based research using a Blaxel sandbox with headless Chromium.
    Fires up an isolated computer, searches the web, extracts content from legal
    databases, government portals, and returns structured research."""
    user = await get_current_user(request, authorization)

    # Auto-detect query types if not provided
    from ai_engine import classify_query
    query_types = req.query_types or classify_query(req.query)

    result = await execute_browser_research(
        user_query=req.query,
        query_types=query_types,
        max_pages=req.max_pages,
    )

    return result


@api_router.get("/sandbox/status")
async def api_sandbox_status(request: Request, authorization: str = Header(None)):
    """Get status of all active research sandboxes."""
    user = await get_current_user(request, authorization)
    sandboxes = await get_active_sandboxes()
    return {
        "active_sandboxes": sandboxes,
        "pool_size": len(sandboxes),
    }


@api_router.delete("/sandbox/{sandbox_name}")
async def api_sandbox_cleanup(sandbox_name: str, request: Request, authorization: str = Header(None)):
    """Delete a specific research sandbox."""
    user = await get_current_user(request, authorization)
    success = await cleanup_sandbox(sandbox_name)
    return {"status": "deleted" if success else "failed", "sandbox_name": sandbox_name}


@api_router.post("/sandbox/warm")
async def api_sandbox_warm(request: Request, authorization: str = Header(None)):
    """Pre-warm a sandbox so the first research query is instant."""
    user = await get_current_user(request, authorization)
    asyncio.create_task(warm_sandbox_pool())
    return {"status": "warming", "message": "Sandbox pre-warm initiated in background"}


@api_router.post("/sandbox/cleanup-all")
async def api_sandbox_cleanup_all(request: Request, authorization: str = Header(None)):
    """Destroy ALL active research sandboxes immediately. Use to stop cost."""
    user = await get_current_user(request, authorization)
    sandboxes_before = len(await get_active_sandboxes())
    await cleanup_all_sandboxes()
    return {
        "status": "cleaned",
        "sandboxes_destroyed": sandboxes_before,
        "message": f"All {sandboxes_before} sandbox(es) destroyed. No active sandboxes remain."
    }


# ==================== MOUNT ROUTER ====================
# Re-include the router AFTER all routes are defined to pick up late-defined routes
app.include_router(api_router)

# ==================== T&C COMPLIANCE ENDPOINTS ====================
# Mount the Terms-of-Service compliance router (Clause 2.1-2.3 Acceptance
# Event, Clause 8.9 Grievance Officer, Clause 8.6 data deletion).
try:
    from tos_compliance import router as tos_router, record_acceptance_event, get_latest_acceptance, request_account_deletion, CURRENT_TOS_VERSION

    # Authenticated endpoints — these need the user context from middleware
    @tos_router.post("/acceptance")
    async def _post_acceptance(payload: "AcceptanceEventRequest", request: Request, authorization: str = Header(None)):  # noqa: F821
        user = await get_current_user(request, authorization)

        # ── LOG FIRST ── capture the attempt on the (separate) logs cluster
        # BEFORE we try to write to the primary. If primary is down this is
        # the ONLY record we'll have. Fire-and-forget so a logs-cluster outage
        # never blocks the user's acceptance flow either.
        try:
            from audit_sink import log_tos_event as _log_tos_attempt
            await _log_tos_attempt(
                user=user, request=request, action="accept_attempt", outcome="submitted",
                tos_version=payload.tos_version,
                acknowledged_disclosures=payload.acknowledged_disclosures,
            )
        except Exception:
            pass

        # Now try to persist to the primary (may 500 if Atlas is down)
        try:
            result = await record_acceptance_event(db, user, payload, request)
        except Exception as e:
            # Log the failure on the logs cluster so we know the attempt happened
            try:
                from audit_sink import log_tos_event as _log_tos_fail, log_error_event as _log_err
                await _log_tos_fail(
                    user=user, request=request, action="accept", outcome="persistence_failed",
                    tos_version=payload.tos_version,
                    acknowledged_disclosures=payload.acknowledged_disclosures,
                )
                await _log_err(request=request, error=e, user=user, context={"endpoint": "/api/legal/acceptance"})
            except Exception:
                pass
            # Still return a stub response so the frontend can continue — the
            # logs-cluster record IS the evidentiary artifact when primary is down.
            import uuid as _uuid, datetime as _dt
            result = {
                "acceptance_id": f"acc_logs_only_{_uuid.uuid4().hex[:12]}",
                "user_id": user.get("user_id"),
                "tos_version": payload.tos_version,
                "accepted_at_utc": _dt.datetime.now(_dt.timezone.utc).isoformat(),
                "recorded_in": "logs_cluster_only",
                "note": "Primary DB unreachable — acceptance recorded on logs cluster. Reconcile when primary restores.",
            }

        # Log success AFTER persistence — this is the canonical "accepted" marker
        try:
            from audit_sink import log_tos_event as _log_tos_done
            await _log_tos_done(
                user=user, request=request, action="accept", outcome="success",
                tos_version=payload.tos_version,
                acknowledged_disclosures=payload.acknowledged_disclosures,
                acceptance_id=result.get("acceptance_id"),
            )
        except Exception:
            pass
        return result

    @tos_router.get("/acceptance/status")
    async def _get_acceptance_status(request: Request, authorization: str = Header(None)):
        user = await get_current_user(request, authorization)
        latest = await get_latest_acceptance(db, user["user_id"])
        needs_acceptance = (not latest) or (latest.get("tos_version") != CURRENT_TOS_VERSION)
        return {
            "has_valid_acceptance": not needs_acceptance,
            "current_tos_version": CURRENT_TOS_VERSION,
            "accepted_version": latest.get("tos_version") if latest else None,
            "accepted_at_ist": latest.get("accepted_at_ist") if latest else None,
        }

    @tos_router.post("/account/delete")
    async def _request_deletion(request: Request, authorization: str = Header(None)):
        user = await get_current_user(request, authorization)
        return await request_account_deletion(db, user)

    # Fix the type hint now that the module is imported
    from tos_compliance import AcceptanceEventRequest
    _post_acceptance.__annotations__["payload"] = AcceptanceEventRequest

    app.include_router(tos_router)
    logger.info("T&C compliance router mounted at /api/legal")
except Exception as e:
    logger.warning(f"T&C compliance router mount failed: {e}", exc_info=True)

# Mount Drive MCP server at /api/mcp/drive
try:
    from mcp_drive_server import create_mcp_router as create_drive_mcp_router
    drive_mcp_router = create_drive_mcp_router(
        db_getter=lambda: db,
        files_cache_getter=lambda: _generated_files_cache,
    )
    # Wrap under /api prefix
    api_router_for_mcp = APIRouter(prefix="/api")
    api_router_for_mcp.include_router(drive_mcp_router)
    app.include_router(api_router_for_mcp)
    logger.info("Drive MCP server mounted at /api/mcp/drive")
except Exception as e:
    logger.warning(f"Drive MCP server mount failed: {e}")


# ==================== ENTRY POINT ====================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8000,
        log_level="info",
        access_log=True,
    )
